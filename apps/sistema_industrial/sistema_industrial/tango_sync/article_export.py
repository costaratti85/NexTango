"""Export de artículos ERPNext → plantilla de actualización masiva de Tango (STA11).

Espejo inverso de `article_push.py` (que trae Tango→ERPNext). Acá vamos
ERPNext→Tango, para el caso de uso del OCR de proveedores: los artículos de
**ferretería** (prefijo `06-`) que el flujo OCR crea en ERPNext a partir de las
facturas, se bajan a Tango con la plantilla masiva.

ALCANCE Y SEGURIDAD (leer):
- Esta capa es **pura y de solo lectura**: arma las FILAS de datos a partir de los
  Items. **No escribe en Tango, no hace submit, no genera el .xlsx final.**
- El paso final —pegar estas filas en la **plantilla oficial de Tango** (que trae
  81 hojas + `_metadata` con claves que Tango valida) y subirla— es un paso
  **separado y gateado**: es **zona fiscal**, requiere OK de Constantino y depende
  de la prueba "¿Tango pisa o respeta celdas vacías?" (caso A/B) y de si el alta
  de artículo nuevo exige columnas fiscales. Ver
  `coordination/reports/FORGE_PROPUESTA_EXPORT_FERRETERIA_A_TANGO.md`.
- Entrada flexible: acepta un **conjunto puntual de item_codes** (los nuevos del
  OCR) o un filtro por Item Group (default "Ferretería" = todo `06-`).

Contrato de salida por Item:
    {
        "item_code": str,
        "filled": {  <encabezado exacto de la columna Tango>: <valor>, ... },
        "gaps":   [ <encabezados fiscales/TBD que quedan vacíos>, ... ],
    }
"""
from __future__ import annotations

import io
import os

import frappe

# Plantilla oficial de Tango (bundleada en el repo). El generador la carga y solo
# pega las filas en la hoja "Artículos", preservando las 80 hojas de lookup + la
# hoja oculta `_metadata` con las claves que Tango valida. NO se fabrica de cero.
_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "tango_templates", "articulos_update_template.xlsx"
)
_TEMPLATE_SHEET = "Artículos"

# --- Mapeos de ferretería (uniformes, verificados sobre los 50 items 06-) ---

# stock_uom (ERPNext) -> código MEDIDA de Tango. Ferretería es por unidad.
# (KG/METRO/M² no aplican a ferretería; si aparece algo distinto -> gap.)
_UOM_TO_TANGO: dict[str, str] = {
    "Nos": "UNIDAD",
    "Unidad": "UNIDAD",
    "Unit": "UNIDAD",
}

# Encabezados exactos de las columnas de la hoja "Artículos" de la plantilla Tango.
COL_CODIGO = "Código"
COL_DESCRIPCION = "Descripción"
COL_PERFIL = "Perfil"
COL_UM_STOCK1 = "Código de UM de stock 1 (Precios y costos)"
COL_LLEVA_STOCK = "Lleva stock"
COL_ELIMINAR = "Eliminar"

# Columnas fiscales/comerciales que ERPNext no tiene hoy → gaps (se completan
# según la decisión de alcance de Constantino / caso A-B de la plantilla).
_FISCAL_GAPS = [
    "Código de IVA Ventas",
    "Código de IVA Compras",
    "Código de base",
]

_ITEM_FIELDS = [
    "item_code",
    "item_name",
    "item_group",
    "stock_uom",
    "is_stock_item",
    "is_sales_item",
    "is_purchase_item",
    "si_tango_id",
]


def _perfil(is_sales: int, is_purchase: int) -> str:
    """Perfil Tango: A=Compra-Venta, V=Venta, C=Compra, N=Inhabilitado."""
    if is_sales and is_purchase:
        return "A"
    if is_sales:
        return "V"
    if is_purchase:
        return "C"
    return "N"


def select_export_items(
    item_codes: list[str] | None = None,
    item_group: str = "Ferretería",
) -> list[dict]:
    """Resuelve el conjunto de Items a exportar.

    Args:
        item_codes: lista puntual de códigos (los nuevos del OCR). Tiene prioridad.
        item_group: si no se pasan item_codes, filtra por este grupo (default
            "Ferretería" = todo `06-`).

    Returns:
        list[dict] con los campos de `_ITEM_FIELDS`.
    """
    if item_codes:
        filters: dict = {"item_code": ["in", list(item_codes)]}
    else:
        filters = {"item_group": item_group}

    return frappe.get_all(
        "Item",
        filters=filters,
        fields=_ITEM_FIELDS,
        order_by="item_code asc",
        limit_page_length=0,
    )


def build_tango_article_rows(
    item_codes: list[str] | None = None,
    item_group: str = "Ferretería",
) -> list[dict]:
    """Arma las filas Tango para un conjunto de Items (o un grupo).

    PURO / SOLO LECTURA: no escribe nada. Devuelve las filas listas para que el
    paso gateado las pegue en la plantilla oficial de Tango.

    Uso desde el flujo OCR (tras la confirmación humana de los artículos nuevos):
        from sistema_industrial.tango_sync.article_export import build_tango_article_rows
        filas = build_tango_article_rows(item_codes=nuevos_codigos_ocr)
    """
    items = select_export_items(item_codes=item_codes, item_group=item_group)
    rows: list[dict] = []

    for it in items:
        uom_tango = _UOM_TO_TANGO.get(it.get("stock_uom") or "")
        filled: dict = {
            COL_CODIGO: it["item_code"],
            COL_DESCRIPCION: it.get("item_name") or it["item_code"],
            COL_PERFIL: _perfil(it.get("is_sales_item") or 0, it.get("is_purchase_item") or 0),
            # "Lleva stock": la plantilla usa codificación true/false
            COL_LLEVA_STOCK: "true" if it.get("is_stock_item") else "false",
            COL_ELIMINAR: "No",
        }

        gaps = list(_FISCAL_GAPS)
        if uom_tango:
            filled[COL_UM_STOCK1] = uom_tango
        else:
            # unidad inesperada para ferretería → queda como gap explícito
            gaps.append(f"{COL_UM_STOCK1} (unidad ERPNext '{it.get('stock_uom')}' sin mapeo Tango)")

        rows.append({"item_code": it["item_code"], "filled": filled, "gaps": gaps})

    return rows


def _render_template_xlsx(rows: list[dict]) -> bytes:
    """Carga la plantilla oficial de Tango y pega las filas en la hoja "Artículos".

    Preserva el resto del workbook (lookups + `_metadata`). Devuelve el .xlsx en bytes.
    openpyxl se importa de forma perezosa (dep del bench, ver deps del OCR).
    """
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError(
            "Falta openpyxl en el entorno del bench (dep del módulo OCR). "
            "Instalar: bench pip install openpyxl."
        ) from exc

    if not os.path.exists(_TEMPLATE_PATH):
        raise FileNotFoundError(f"No se encuentra la plantilla de Tango: {_TEMPLATE_PATH}")

    wb = openpyxl.load_workbook(_TEMPLATE_PATH)
    if _TEMPLATE_SHEET not in wb.sheetnames:
        raise ValueError(f"La plantilla no tiene la hoja '{_TEMPLATE_SHEET}'")
    ws = wb[_TEMPLATE_SHEET]

    # Mapa encabezado (fila 1) -> índice de columna.
    header_to_col = {
        ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
    }

    # Escribir cada fila a partir de la fila 2, solo en las columnas conocidas.
    for offset, row in enumerate(rows):
        excel_row = 2 + offset
        for header, value in row["filled"].items():
            col = header_to_col.get(header)
            if col:
                ws.cell(row=excel_row, column=col, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_tango_import_excel(
    item_codes: list[str] | None = None,
    item_group: str = "Ferretería",
) -> dict:
    """Genera el Excel de importación masiva de Tango para un conjunto de Items.

    Es el punto que **Atlas invoca después de crear los Items nuevos del OCR**:
    le pasa los `item_codes` recién creados (ferretería `06-`, con el código que
    ingresó el humano) y recibe el `file_url` del .xlsx listo para que Constantino
    lo revise y lo suba a Tango.

    IMPORTANTE (Regla 8 / fiscal): esto genera un **archivo descargable**, NO sube
    nada a Tango. La subida la hace el humano. El `_metadata`/claves de la plantilla
    se preservan, pero la validación final del importador de Tango debe probarse con
    Constantino (idealmente un ALTA, no un update — ver §6-bis del reporte de diseño).

    Returns:
        {
            "file_url": str,        # URL del File de Frappe (privado)
            "file_name": str,
            "count": int,           # cantidad de artículos exportados
            "item_codes": list[str],
            "gaps": list[str],      # columnas fiscales sin datos en ERPNext (unión)
        }
    """
    rows = build_tango_article_rows(item_codes=item_codes, item_group=item_group)
    if not rows:
        frappe.throw("No hay Items para exportar con los filtros dados.")

    content = _render_template_xlsx(rows)

    from frappe.utils.file_manager import save_file

    stamp = frappe.utils.now().replace(" ", "_").replace(":", "").replace("-", "")
    file_name = f"tango_articulos_{len(rows)}items_{stamp}.xlsx"
    file_doc = save_file(file_name, content, None, None, is_private=1)

    gaps: list[str] = sorted({g for row in rows for g in row["gaps"]})
    return {
        "file_url": file_doc.file_url,
        "file_name": file_doc.file_name,
        "count": len(rows),
        "item_codes": [row["item_code"] for row in rows],
        "gaps": gaps,
    }


@frappe.whitelist()
def generate_tango_import_excel_api(item_codes=None, item_group: str = "Ferretería") -> dict:
    """Wrapper whitelisted de `generate_tango_import_excel`.

    Atlas puede llamar la función in-process; este endpoint es para disparo desde
    la UI (botón) o pruebas. `item_codes` acepta CSV o JSON.
    """
    codes = _parse_codes(item_codes)
    return generate_tango_import_excel(item_codes=codes, item_group=item_group)


def _parse_codes(item_codes) -> list[str] | None:
    """Normaliza item_codes que puede venir como lista, CSV o JSON string (HTTP)."""
    if not item_codes:
        return None
    if isinstance(item_codes, str):
        item_codes = item_codes.strip()
        if item_codes.startswith("["):
            return frappe.parse_json(item_codes)
        return [c.strip() for c in item_codes.split(",") if c.strip()]
    return list(item_codes)


@frappe.whitelist()
def preview_tango_export(item_codes=None, item_group: str = "Ferretería") -> dict:
    """Wrapper whitelisted de PREVIEW (debug). No escribe nada.

    item_codes puede venir como CSV o JSON por HTTP.
    """
    rows = build_tango_article_rows(item_codes=_parse_codes(item_codes), item_group=item_group)
    return {
        "count": len(rows),
        "rows": rows,
        "note": (
            "PREVIEW de solo lectura. Generar el .xlsx sobre la plantilla oficial de "
            "Tango y subirlo es un paso separado y gateado (fiscal, OK Constantino)."
        ),
    }
