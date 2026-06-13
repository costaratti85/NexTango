import os
import re
import json
import sqlite3
import tempfile
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from dataclasses import dataclass
from typing import List
from difflib import SequenceMatcher
from datetime import datetime
from PIL import Image, ImageDraw, ImageOps, ImageEnhance
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
import fitz
import pytesseract

try:
    import cv2
except ImportError:
    cv2 = None


pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

APP_DB = "facturas_tango_v8.db"
CHECK_ON = "☑"
CHECK_OFF = "☐"


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    page: int


@dataclass
class LineBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    words: List[WordBox]
    score: int = 0


class BaseLocal:
    def __init__(self, db_path=APP_DB):
        self.db_path = db_path
        self.inicializar()

    def conectar(self):
        return sqlite3.connect(self.db_path)

    def inicializar(self):
        con = self.conectar()
        cur = con.cursor()

        cur.execute("""
            CREATE TABLE IF NOT EXISTS facturas_procesadas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clave TEXT UNIQUE,
                proveedor TEXT,
                cuit TEXT,
                tipo TEXT,
                punto_venta TEXT,
                numero TEXT,
                numero_completo TEXT,
                fecha TEXT,
                archivo TEXT,
                ruta TEXT,
                fecha_proceso TEXT,
                archivo_resumen TEXT,
                archivo_alta TEXT
            )
        """)

        cur.execute("""
            CREATE TABLE IF NOT EXISTS equivalencias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cuit TEXT,
                proveedor TEXT,
                codigo_proveedor TEXT,
                descripcion_proveedor TEXT,
                codigo_tango TEXT,
                descripcion_tango TEXT,
                codigo_barras TEXT,
                veces_usado INTEGER DEFAULT 1,
                ultima_fecha TEXT,
                UNIQUE(cuit, codigo_proveedor)
            )
        """)

        con.commit()
        con.close()

    def factura_existe(self, clave):
        con = self.conectar()
        cur = con.cursor()
        cur.execute("SELECT fecha_proceso, archivo_resumen FROM facturas_procesadas WHERE clave = ?", (clave,))
        row = cur.fetchone()
        con.close()
        return row

    def guardar_factura_procesada(self, datos, archivo_resumen="", archivo_alta=""):
        con = self.conectar()
        cur = con.cursor()

        cur.execute("""
            INSERT OR REPLACE INTO facturas_procesadas (
                clave, proveedor, cuit, tipo, punto_venta, numero, numero_completo,
                fecha, archivo, ruta, fecha_proceso, archivo_resumen, archivo_alta
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datos.get("clave", ""),
            datos.get("proveedor", ""),
            datos.get("cuit_proveedor", ""),
            datos.get("tipo", ""),
            datos.get("punto_venta", ""),
            datos.get("numero", ""),
            datos.get("numero_completo", ""),
            datos.get("fecha", ""),
            datos.get("archivo", ""),
            datos.get("ruta", ""),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            archivo_resumen,
            archivo_alta
        ))

        con.commit()
        con.close()

    def buscar_equivalencia(self, cuit, codigo_proveedor):
        if not cuit or not codigo_proveedor:
            return None

        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            SELECT codigo_tango, descripcion_tango, codigo_barras
            FROM equivalencias
            WHERE cuit = ? AND codigo_proveedor = ?
        """, (str(cuit).strip(), str(codigo_proveedor).strip()))

        row = cur.fetchone()
        con.close()

        if row:
            return {
                "codigo_tango": row[0],
                "descripcion_tango": row[1],
                "codigo_barras": row[2],
            }

        return None

    def guardar_equivalencia(self, cuit, proveedor, codigo_proveedor, descripcion_proveedor,
                             codigo_tango, descripcion_tango, codigo_barras):
        if not cuit or not codigo_proveedor or not codigo_tango:
            return

        con = self.conectar()
        cur = con.cursor()

        cur.execute("""
            INSERT INTO equivalencias (
                cuit, proveedor, codigo_proveedor, descripcion_proveedor,
                codigo_tango, descripcion_tango, codigo_barras,
                veces_usado, ultima_fecha
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(cuit, codigo_proveedor)
            DO UPDATE SET
                proveedor = excluded.proveedor,
                descripcion_proveedor = excluded.descripcion_proveedor,
                codigo_tango = excluded.codigo_tango,
                descripcion_tango = excluded.descripcion_tango,
                codigo_barras = excluded.codigo_barras,
                veces_usado = veces_usado + 1,
                ultima_fecha = excluded.ultima_fecha
        """, (
            str(cuit).strip(),
            str(proveedor).strip(),
            str(codigo_proveedor).strip(),
            str(descripcion_proveedor).strip(),
            str(codigo_tango).strip(),
            str(descripcion_tango).strip(),
            str(codigo_barras).strip(),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

        con.commit()
        con.close()


class CatalogoTango:
    def __init__(self):
        self.articulos = []

    def quitar_acentos(self, texto):
        texto = unicodedata.normalize("NFD", texto)
        texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
        return texto

    def normalizar(self, texto):
        if texto is None:
            return ""

        texto = str(texto).upper().strip()
        texto = self.quitar_acentos(texto)

        reemplazos = {
            "P/ACERO": "ACERO",
            "P ACERO": "ACERO",
            "PACERO": "ACERO",
            "P/ ACERO": "ACERO",
            "P/INOX": "INOX",
            "P INOX": "INOX",
            "P/ INOX": "INOX",
            "INOXIDABLE": "INOX",
            "C/": "CON ",
            "S/": "SIN ",
            "MM.": "MM",
            "M.M.": "MM",
        }

        for viejo, nuevo in reemplazos.items():
            texto = texto.replace(viejo, nuevo)

        for ch in ["/", "-", ",", ".", "(", ")", "[", "]", "_"]:
            texto = texto.replace(ch, " ")

        texto = re.sub(r"\s+", " ", texto).strip()
        return texto

    def tokens(self, texto):
        texto = self.normalizar(texto)
        return [t for t in texto.split() if len(t) >= 2]

    def cargar(self, ruta_excel):
        wb = load_workbook(ruta_excel, data_only=True)

        if "Artículos" not in wb.sheetnames:
            raise Exception("El Excel no tiene una hoja llamada 'Artículos'.")

        ws = wb["Artículos"]
        self.articulos = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = row[0] if len(row) > 0 else ""
            descripcion = row[1] if len(row) > 1 else ""
            desc_adic = row[2] if len(row) > 2 else ""
            sinonimo = row[3] if len(row) > 3 else ""
            codigo_barras = row[4] if len(row) > 4 else ""

            if not codigo:
                continue

            art = {
                "codigo": str(codigo).strip(),
                "descripcion": str(descripcion or "").strip(),
                "desc_adic": str(desc_adic or "").strip(),
                "sinonimo": str(sinonimo or "").strip(),
                "codigo_barras": str(codigo_barras or "").strip(),
            }

            art["_codigo_n"] = self.normalizar(art["codigo"])
            art["_descripcion_n"] = self.normalizar(art["descripcion"])
            art["_desc_adic_n"] = self.normalizar(art["desc_adic"])
            art["_sinonimo_n"] = self.normalizar(art["sinonimo"])
            art["_codigo_barras_n"] = self.normalizar(art["codigo_barras"])
            art["_texto_completo_n"] = self.normalizar(
                f"{art['codigo']} {art['descripcion']} {art['desc_adic']} {art['sinonimo']} {art['codigo_barras']}"
            )

            self.articulos.append(art)

    def similitud(self, a, b):
        a = self.normalizar(a)
        b = self.normalizar(b)

        if not a or not b:
            return 0

        return int(SequenceMatcher(None, a, b).ratio() * 100)

    def score_tokens(self, texto_ocr, texto_tango):
        tokens_ocr = set(self.tokens(texto_ocr))
        tokens_tango = set(self.tokens(texto_tango))

        if not tokens_ocr or not tokens_tango:
            return 0

        comunes = tokens_ocr.intersection(tokens_tango)
        return int((len(comunes) / len(tokens_ocr)) * 100)

    def buscar(self, codigo_proveedor="", codigo_barras="", descripcion=""):
        codigo_proveedor_n = self.normalizar(codigo_proveedor)
        codigo_barras_n = self.normalizar(codigo_barras)
        descripcion_n = self.normalizar(descripcion)

        if codigo_barras_n:
            for art in self.articulos:
                if art["_codigo_barras_n"] and art["_codigo_barras_n"] == codigo_barras_n:
                    return art, "Código de barras exacto", 100

        if codigo_proveedor_n:
            for art in self.articulos:
                if art["_codigo_n"] == codigo_proveedor_n:
                    return art, "Código Tango exacto", 100

            for art in self.articulos:
                if art["_sinonimo_n"] and art["_sinonimo_n"] == codigo_proveedor_n:
                    return art, "Sinónimo exacto", 100

            for art in self.articulos:
                if codigo_proveedor_n in art["_descripcion_n"]:
                    return art, "Código proveedor en descripción", 96

            for art in self.articulos:
                if codigo_proveedor_n in art["_texto_completo_n"]:
                    return art, "Código proveedor contenido", 94

        mejor_art = None
        mejor_metodo = ""
        mejor_score = 0

        if descripcion_n:
            for art in self.articulos:
                score_desc = self.similitud(descripcion_n, art["_descripcion_n"])
                score_tokens = self.score_tokens(descripcion_n, art["_descripcion_n"])
                score_final = max(score_desc, score_tokens)

                if score_final > mejor_score:
                    mejor_score = score_final
                    mejor_art = art
                    mejor_metodo = "Similitud descripción"

            if mejor_art and mejor_score >= 82:
                return mejor_art, mejor_metodo, mejor_score

        return None, "", 0


class FacturaTableReader:
    def __init__(self):
        self.stop_words = [
            "total", "subtotal", "cae", "iva 21", "iva 10", "iva 27",
            "son pesos", "observaciones", "importe neto", "fecha de vto"
        ]

    def is_cuit(self, text):
        clean = text.strip()

        if re.fullmatch(r"\d{2}-\d{8}-\d", clean):
            return True

        only_digits = re.sub(r"\D", "", clean)

        if len(only_digits) == 11 and only_digits[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
            return True

        return False

    def normalizar_cuit(self, cuit):
        return re.sub(r"\D", "", str(cuit or ""))

    def abrir_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()
        if ext == ".pdf":
            return self.leer_pdf(ruta)
        return self.leer_imagen(ruta)

    def leer_pdf(self, ruta):
        doc = fitz.open(ruta)
        paginas = []

        for page_index, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(tempfile.gettempdir(), f"pagina_factura_{page_index}.png")
            pix.save(img_path)
            img = Image.open(img_path)

            palabras = []
            escala = 2

            for w in page.get_text("words"):
                x0, y0, x1, y1, text = w[:5]
                text = str(text).strip()

                if text:
                    palabras.append(WordBox(x0 * escala, y0 * escala, x1 * escala, y1 * escala, text, page_index))

            paginas.append((img, palabras))

        return paginas

    def leer_imagen(self, ruta):
        img = Image.open(ruta)
        img_ocr = self.preparar_imagen(img)

        data = pytesseract.image_to_data(
            img_ocr,
            lang="spa",
            output_type=pytesseract.Output.DICT,
            config="--psm 6"
        )

        palabras = []

        for i, text in enumerate(data["text"]):
            text = str(text).strip()
            if not text:
                continue

            try:
                conf = int(float(data["conf"][i]))
            except Exception:
                conf = -1

            if conf < 25:
                continue

            x = data["left"][i]
            y = data["top"][i]
            w = data["width"][i]
            h = data["height"][i]

            palabras.append(WordBox(x, y, x + w, y + h, text, 1))

        return [(img, palabras)]

    def preparar_imagen(self, img):
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.2)
        return img

    def agrupar_lineas(self, palabras):
        palabras = sorted(palabras, key=lambda w: (w.y0, w.x0))
        lineas = []
        actual = []

        for palabra in palabras:
            if not actual:
                actual = [palabra]
                continue

            y_prom = sum(w.y0 for w in actual) / len(actual)

            if abs(palabra.y0 - y_prom) <= 8:
                actual.append(palabra)
            else:
                lineas.append(self.crear_linea(actual))
                actual = [palabra]

        if actual:
            lineas.append(self.crear_linea(actual))

        return lineas

    def crear_linea(self, palabras):
        palabras = sorted(palabras, key=lambda w: w.x0)
        return LineBox(
            x0=min(w.x0 for w in palabras),
            y0=min(w.y0 for w in palabras),
            x1=max(w.x1 for w in palabras),
            y1=max(w.y1 for w in palabras),
            text=" ".join(w.text for w in palabras),
            words=palabras
        )

    def parse_decimal(self, t):
        t = str(t).replace("$", "").replace(" ", "").strip()

        if not t:
            return None

        if re.match(r"^-?\d{1,3}(\.\d{3})*,\d+$", t):
            t = t.replace(".", "").replace(",", ".")
        elif re.match(r"^-?\d+,\d+$", t):
            t = t.replace(",", ".")
        else:
            try:
                return float(t)
            except Exception:
                return None

        try:
            return float(t)
        except Exception:
            return None

    def es_numero(self, t):
        t = t.replace("$", "").strip()

        if self.is_cuit(t):
            return False

        return (
            bool(re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2,4}$", t))
            or bool(re.match(r"^-?\d+([,.]\d+)?$", t))
        )

    def cantidad_columnas(self, linea):
        if len(linea.words) < 3:
            return 0

        gaps = 0
        for a, b in zip(linea.words, linea.words[1:]):
            if b.x0 - a.x1 > 25:
                gaps += 1

        return gaps + 1

    def puntuar_linea(self, linea):
        texto = linea.text.lower()

        if any(self.is_cuit(w.text) for w in linea.words):
            linea.score = -10
            return linea.score

        if any(p in texto for p in self.stop_words):
            linea.score = -5
            return linea.score

        numeros = sum(1 for w in linea.words if self.es_numero(w.text))
        columnas = self.cantidad_columnas(linea)

        score = 0

        if columnas >= 3:
            score += 2
        if columnas >= 5:
            score += 2
        if numeros >= 2:
            score += 2
        if numeros >= 4:
            score += 2
        if len(linea.words) >= 5:
            score += 1

        linea.score = score
        return score

    def detectar_zonas(self, lineas):
        for linea in lineas:
            self.puntuar_linea(linea)

        candidatas = [l for l in lineas if l.score >= 4]

        if not candidatas:
            return []

        grupos = []
        grupo = [candidatas[0]]

        for linea in candidatas[1:]:
            if linea.y0 - grupo[-1].y1 < 75:
                grupo.append(linea)
            else:
                grupos.append(grupo)
                grupo = [linea]

        grupos.append(grupo)

        zonas = []

        for grupo in grupos:
            if len(grupo) < 2:
                continue

            zonas.append((
                min(l.x0 for l in grupo),
                min(l.y0 for l in grupo),
                max(l.x1 for l in grupo),
                max(l.y1 for l in grupo),
                grupo
            ))

        return zonas

    def detectar_items(self, lineas):
        zonas = self.detectar_zonas(lineas)
        items = []

        for x0, y0, x1, y1, grupo in zonas:
            for linea in grupo:
                if linea.score < 4:
                    continue

                if any(self.is_cuit(w.text) for w in linea.words):
                    continue

                item = self.linea_a_item(linea)

                if item:
                    items.append(item)

        return items, zonas

    def linea_a_item(self, linea):
        words = linea.words

        codigo_proveedor = ""
        descripcion_parts = []
        cantidad = ""
        precio = ""
        importe = ""
        iva = ""
        advertencia = ""

        numeros_detectados = []

        for w in words:
            t = w.text.strip()

            if self.is_cuit(t):
                return None

            if self.es_numero(t):
                numeros_detectados.append(t)
                continue

            if not codigo_proveedor and re.match(r"^[A-Z0-9][A-Z0-9.\-/]{2,}$", t, re.IGNORECASE):
                codigo_proveedor = t[:15]
                continue

            if t.upper() in ["U", "UN", "KG", "M", "L", "%"]:
                continue

            if re.match(r"^-?\d+([,.]\d+)?%?$", t):
                continue

            descripcion_parts.append(t)

        valores = [(n, self.parse_decimal(n)) for n in numeros_detectados]
        valores_validos = [(n, v) for n, v in valores if v is not None]

        # Detectar IVA: si hay muchos números, valores como 21,00 / 10,50 / 27,00 / 0,00 probablemente son IVA
        iva_posibles = []
        if len(valores_validos) >= 4:
            for n, v in valores_validos:
                if abs(v - 21.0) < 0.01 or abs(v - 10.5) < 0.01 or abs(v - 27.0) < 0.01 or abs(v - 0.0) < 0.01:
                    iva_posibles.append((n, v))

        if iva_posibles:
            # preferimos 21, 10.5 o 27 antes que 0
            no_cero = [x for x in iva_posibles if abs(x[1]) > 0.01]
            elegido = no_cero[0] if no_cero else iva_posibles[0]
            iva = elegido[0]
            valores_validos = [x for x in valores_validos if x[0] != elegido[0]]

        if valores_validos:
            cantidad = valores_validos[0][0]

        if len(valores_validos) >= 2:
            precio = valores_validos[-2][0]

        if len(valores_validos) >= 1:
            importe = valores_validos[-1][0]

        descripcion = " ".join(descripcion_parts)
        descripcion = re.sub(r"\s+", " ", descripcion).strip()
        descripcion = descripcion.replace(" /", "").replace("/", "")

        if not descripcion and not codigo_proveedor:
            return None

        advertencia = self.validar_calculo(cantidad, precio, importe)

        return {
            "usar": True,
            "existe_tango": "No",
            "accion": "Alta artículo",
            "codigo_tango": "",
            "descripcion": descripcion[:50],
            "desc_adic": "",
            "sinonimo": codigo_proveedor[:15],
            "codigo_barras": "",
            "cantidad": cantidad,
            "precio": precio,
            "importe": importe,
            "iva": iva,
            "advertencia": advertencia,
            "tipo": "Simple",
            "escala": "No usa",
            "linea_detectada": linea.text,
            "score": linea.score,
            "coincidencia": "",
            "match_score": 0,
        }

    def validar_calculo(self, cantidad, precio, importe):
        c = self.parse_decimal(cantidad)
        p = self.parse_decimal(precio)
        i = self.parse_decimal(importe)

        if c is None or p is None or i is None:
            return ""

        calculado = c * p

        # tolerancia: 1 peso o 1%
        tolerancia = max(1.0, abs(i) * 0.01)

        if abs(calculado - i) > tolerancia:
            return "Revisar cantidad/precio/importe"

        return "OK"

    def extraer_texto_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()

        if ext == ".pdf":
            try:
                doc = fitz.open(ruta)
                texto = ""
                for page in doc:
                    texto += page.get_text("text") + "\n"
                return texto
            except Exception:
                return ""

        try:
            img = Image.open(ruta)
            img = self.preparar_imagen(img)
            try:
                return pytesseract.image_to_string(img, lang="spa")
            except Exception:
                return pytesseract.image_to_string(img)
        except Exception:
            return ""

    def detectar_proveedor(self, texto):
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]

        for l in lineas[:20]:
            may = l.upper()
            if "S.R.L" in may or "SRL" in may or "S.A" in may or " SA" in may:
                return l

        for l in lineas[:10]:
            if len(l) > 4 and not re.search(r"\d{2}/\d{2}/\d{2,4}", l):
                return l

        return ""

    def detectar_cuit_proveedor(self, texto):
        patrones = [
            r"CUIT\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
            r"C\.?U\.?I\.?T\.?\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
        ]

        for patron in patrones:
            matches = re.findall(patron, texto, re.IGNORECASE)
            for m in matches:
                cuit = self.normalizar_cuit(m)
                if len(cuit) == 11 and cuit[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
                    return cuit

        m2 = re.search(r"\b((?:20|23|24|27|30|33|34)\d{9})\b", texto)
        if m2:
            return m2.group(1)

        return ""

    def detectar_tipo_comprobante(self, texto):
        may = texto.upper()

        if "FACTURA" in may:
            if re.search(r"FACTURA\s*A", may) or re.search(r"\bCOD\.?\s*0?1\b", may):
                return "FACTURA A"
            if re.search(r"FACTURA\s*B", may) or re.search(r"\bCOD\.?\s*0?6\b", may):
                return "FACTURA B"
            if re.search(r"FACTURA\s*C", may) or re.search(r"\bCOD\.?\s*0?11\b", may):
                return "FACTURA C"
            return "FACTURA"

        if "REMITO" in may:
            return "REMITO"

        return ""

    def detectar_numero_comprobante(self, texto):
        patrones = [
            r"Punto\s+de\s+Venta\s*[:\-]?\s*(\d{4,5}).{0,80}?Comp\.?\s*Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"Pto\.?\s*Vta\.?\s*[:\-]?\s*(\d{4,5}).{0,80}?Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"(?:FACTURA|REMITO)\s*(?:N[°º]?)?\s*[:\-]?\s*(\d{4,5})[-\s]?(\d{6,8})",
            r"\b(\d{4,5})[-\s](\d{6,8})\b",
        ]

        for patron in patrones:
            m = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
            if m:
                punto = m.group(1).zfill(4)
                numero = m.group(2).zfill(8)
                return punto, numero, f"{punto}-{numero}"

        return "", "", ""

    def detectar_fecha(self, texto):
        m = re.search(
            r"(?:Fecha\s+de\s+Emisi[oó]n|Fecha)\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})",
            texto,
            re.IGNORECASE
        )

        if m:
            return m.group(1)

        m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", texto)
        if m:
            return m.group(1)

        return ""

    def detectar_total(self, texto):
        posibles = re.findall(r"(?:TOTAL|Total)\s*[:$ ]+\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})", texto)
        if posibles:
            return posibles[-1]
        return ""

    def clave_factura(self, datos):
        cuit = self.normalizar_cuit(datos.get("cuit_proveedor", ""))
        proveedor = str(datos.get("proveedor", "")).strip().upper()
        tipo = str(datos.get("tipo", "")).strip().upper()
        punto = str(datos.get("punto_venta", "")).strip().zfill(4) if datos.get("punto_venta") else ""
        numero = str(datos.get("numero", "")).strip().zfill(8) if datos.get("numero") else ""

        if cuit and punto and numero:
            return f"CUIT:{cuit}|TIPO:{tipo}|PV:{punto}|N:{numero}"

        if proveedor and punto and numero:
            proveedor_n = re.sub(r"\s+", " ", proveedor)
            return f"PROV:{proveedor_n}|TIPO:{tipo}|PV:{punto}|N:{numero}"

        archivo = datos.get("archivo", "")
        return f"ARCHIVO:{archivo}"

    def detectar_datos_factura(self, ruta):
        texto = self.extraer_texto_documento(ruta)

        archivo = os.path.basename(ruta)
        proveedor = self.detectar_proveedor(texto)
        cuit = self.detectar_cuit_proveedor(texto)
        tipo = self.detectar_tipo_comprobante(texto)
        punto, numero, numero_completo = self.detectar_numero_comprobante(texto)
        fecha = self.detectar_fecha(texto)
        total = self.detectar_total(texto)

        etiqueta = archivo
        partes = []

        if proveedor:
            partes.append(proveedor[:30])
        if numero_completo:
            partes.append(numero_completo)
        if fecha:
            partes.append(fecha)

        if partes:
            etiqueta = " | ".join(partes)

        datos = {
            "archivo": archivo,
            "ruta": ruta,
            "proveedor": proveedor,
            "cuit_proveedor": cuit,
            "tipo": tipo,
            "punto_venta": punto,
            "numero": numero,
            "numero_completo": numero_completo,
            "fecha": fecha,
            "total": total,
            "etiqueta": etiqueta,
        }

        datos["clave"] = self.clave_factura(datos)
        return datos

    def analizar(self, ruta):
        paginas = self.abrir_documento(ruta)
        resultados = []
        todos_items = []

        datos_factura = self.detectar_datos_factura(ruta)

        for page_num, (img, palabras) in enumerate(paginas, start=1):
            lineas = self.agrupar_lineas(palabras)
            items, zonas = self.detectar_items(lineas)
            debug_path = self.dibujar_debug(img, lineas, zonas, page_num, datos_factura["archivo"])

            for item in items:
                item["pagina"] = page_num
                item["factura_archivo"] = datos_factura["archivo"]
                item["factura_etiqueta"] = datos_factura["etiqueta"]
                item["proveedor_factura"] = datos_factura["proveedor"]
                item["cuit_proveedor"] = datos_factura["cuit_proveedor"]
                item["tipo_factura"] = datos_factura["tipo"]
                item["punto_venta"] = datos_factura["punto_venta"]
                item["nro_factura"] = datos_factura["numero"]
                item["numero_completo"] = datos_factura["numero_completo"]
                item["fecha_factura"] = datos_factura["fecha"]
                item["total_factura"] = datos_factura["total"]
                item["ruta_factura"] = datos_factura["ruta"]
                item["clave_factura"] = datos_factura["clave"]
                todos_items.append(item)

            resultados.append((datos_factura["archivo"], page_num, debug_path, lineas, zonas))

        return todos_items, resultados, datos_factura

    def dibujar_debug(self, img, lineas, zonas, page_num, nombre_archivo):
        debug = img.convert("RGB")
        draw = ImageDraw.Draw(debug)

        for linea in lineas:
            if linea.score >= 4:
                draw.rectangle([linea.x0, linea.y0, linea.x1, linea.y1], outline="blue", width=2)
                draw.text((linea.x0, max(0, linea.y0 - 14)), f"S{linea.score}", fill="blue")

        for x0, y0, x1, y1, grupo in zonas:
            draw.rectangle([x0, y0, x1, y1], outline="red", width=4)

        base = re.sub(r"[^A-Za-z0-9_-]+", "_", nombre_archivo)
        out = os.path.join(tempfile.gettempdir(), f"debug_{base}_pagina_{page_num}.png")
        debug.save(out)
        return out


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Facturas múltiples → Tango V8")
        self.root.geometry("1760x860")

        self.base = BaseLocal()
        self.reader = FacturaTableReader()
        self.catalogo = CatalogoTango()

        self.ruta_modelo_tango = None
        self.ruta_catalogo_tango = None
        self.resultados_debug = []
        self.facturas_cargadas = []
        self.claves_facturas = set()
        self._todos_items = []
        self.filtro_actual = tk.StringVar(value="Ver todos")


        self.crear_ui()

    def crear_ui(self):
        self.configurar_estilo()

        main = ttk.Frame(self.root, padding=8)
        main.pack(fill="both", expand=True)

        top = ttk.Frame(main)
        top.pack(fill="x", pady=(0, 8))

        ttk.Button(top, text="1. Artículos Tango", command=self.abrir_catalogo).pack(side="left", padx=3)
        ttk.Button(top, text="2. Modelo Tango", command=self.abrir_modelo).pack(side="left", padx=3)
        ttk.Button(top, text="3. Agregar factura(s)", command=self.agregar_facturas).pack(side="left", padx=3)
        ttk.Button(top, text="Sacar foto", command=self.sacar_foto).pack(side="left", padx=3)
        ttk.Button(top, text="Ver debug", command=self.ver_debug).pack(side="left", padx=3)
        ttk.Button(top, text="Abrir factura", command=self.abrir_factura_original).pack(side="left", padx=3)
        ttk.Button(top, text="Guardar sesión", command=self.guardar_sesion).pack(side="left", padx=3)
        ttk.Button(top, text="Cargar sesión", command=self.cargar_sesion).pack(side="left", padx=3)
        ttk.Button(top, text="Aprender equivalencia", command=self.aprender_equivalencia_seleccionada).pack(side="left", padx=3)

        filtro_frame = ttk.Frame(main)
        filtro_frame.pack(fill="x", pady=(0, 6))

        ttk.Label(filtro_frame, text="Filtro:").pack(side="left")
        combo = ttk.Combobox(
            filtro_frame,
            textvariable=self.filtro_actual,
            state="readonly",
            width=24,
            values=[
                "Ver todos",
                "Solo seleccionados",
                "Solo nuevos",
                "Solo existentes",
                "Sin código Tango",
                "Baja confianza",
                "Destildados",
                "Con advertencias",
            ]
        )
        combo.pack(side="left", padx=5)
        combo.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtro())

        self.lbl_catalogo = ttk.Label(main, text="Artículos Tango: no seleccionado")
        self.lbl_catalogo.pack(fill="x")

        self.lbl_modelo = ttk.Label(main, text="Modelo Tango: no seleccionado")
        self.lbl_modelo.pack(fill="x")

        self.lbl_facturas = ttk.Label(main, text="Facturas cargadas: 0")
        self.lbl_facturas.pack(fill="x", pady=(0, 6))

        columns = (
            "usar", "factura", "proveedor", "cuit", "tipo", "numero", "fecha",
            "existe", "accion", "codigo_tango", "descripcion", "sinonimo",
            "codigo_barras", "cantidad", "precio", "importe", "iva", "advertencia",
            "match", "match_score", "tipo_art", "escala", "pagina", "ruta", "linea"
        )

        headers = {
            "usar": "✓",
            "factura": "Factura",
            "proveedor": "Proveedor",
            "cuit": "CUIT",
            "tipo": "Tipo",
            "numero": "N°",
            "fecha": "Fecha",
            "existe": "Existe",
            "accion": "Acción",
            "codigo_tango": "Código Tango",
            "descripcion": "Descripción",
            "sinonimo": "Cod. proveedor",
            "codigo_barras": "Cód. barras",
            "cantidad": "Cantidad",
            "precio": "Precio",
            "importe": "Importe",
            "iva": "IVA %",
            "advertencia": "Advertencia",
            "match": "Coincidencia",
            "match_score": "Conf.",
            "tipo_art": "Tipo art.",
            "escala": "Escala",
            "pagina": "Pág.",
            "ruta": "Ruta factura",
            "linea": "Línea detectada",
        }

        widths = {
            "usar": 45,
            "factura": 160,
            "proveedor": 200,
            "cuit": 105,
            "tipo": 90,
            "numero": 110,
            "fecha": 85,
            "existe": 65,
            "accion": 145,
            "codigo_tango": 125,
            "descripcion": 320,
            "sinonimo": 130,
            "codigo_barras": 125,
            "cantidad": 85,
            "precio": 90,
            "importe": 95,
            "iva": 70,
            "advertencia": 190,
            "match": 210,
            "match_score": 60,
            "tipo_art": 80,
            "escala": 80,
            "pagina": 50,
            "ruta": 260,
            "linea": 420,
        }

        frame_tabla = ttk.Frame(main)
        frame_tabla.pack(fill="both", expand=True, pady=6)

        scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")
        scroll_y.pack(side="right", fill="y")

        scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        self.tabla = ttk.Treeview(
            frame_tabla,
            columns=columns,
            show="headings",
            height=25,
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        scroll_y.config(command=self.tabla.yview)
        scroll_x.config(command=self.tabla.xview)

        for c in columns:
            self.tabla.heading(c, text=headers[c])
            self.tabla.column(c, width=widths[c], stretch=False)

        self.tabla.tag_configure("verde", background="#d9f5d6")
        self.tabla.tag_configure("amarillo", background="#fff3bf")
        self.tabla.tag_configure("rojo", background="#ffd6d6")
        self.tabla.tag_configure("gris", background="#e9ecef")
        self.tabla.tag_configure("azul", background="#d6eaff")
        self.tabla.tag_configure("advertencia", background="#ffd8a8")

        self.tabla.pack(side="left", fill="both", expand=True)
        self.tabla.bind("<Double-1>", self.doble_click)

        bottom = ttk.Frame(main)
        bottom.pack(fill="x", pady=(8, 0))

        ttk.Button(bottom, text="Seleccionar todo", command=self.seleccionar_todo).pack(side="left", padx=3)
        ttk.Button(bottom, text="Borrar selección", command=self.borrar_seleccion).pack(side="left", padx=3)
        ttk.Button(bottom, text="Invertir selección", command=self.invertir_seleccion).pack(side="left", padx=3)
        ttk.Button(bottom, text="Agregar manual", command=self.agregar_manual).pack(side="left", padx=3)
        ttk.Button(bottom, text="Eliminar fila", command=self.eliminar_fila).pack(side="left", padx=3)
        ttk.Button(bottom, text="Limpiar todo", command=self.limpiar_todo).pack(side="left", padx=3)
        ttk.Button(bottom, text="Recomparar", command=self.recomparar).pack(side="left", padx=3)

        ttk.Button(bottom, text="Reporte revisión", command=self.generar_reporte_revision).pack(side="right", padx=3)
        ttk.Button(bottom, text="Alta artículos", command=self.generar_excel_alta_articulos).pack(side="right", padx=3)
        ttk.Button(bottom, text="Resumen compra/stock", command=self.generar_resumen_compra_stock).pack(side="right", padx=3)

        self.lbl_estado = ttk.Label(
            main,
            text="V8: historial, equivalencias, filtros, colores, sesión, IVA y validación cantidad × precio.",
            foreground="#0b5ed7"
        )
        self.lbl_estado.pack(fill="x", pady=(6, 0))

    def configurar_estilo(self):
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("TButton", padding=5)
        style.configure("Treeview", rowheight=24)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("TLabel", font=("Segoe UI", 9))

    def abrir_catalogo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Artículos.xlsx exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not ruta:
            return

        try:
            self.catalogo.cargar(ruta)
            self.ruta_catalogo_tango = ruta
            self.lbl_catalogo.config(text=f"Artículos Tango: {ruta} ({len(self.catalogo.articulos)} artículos)")
            self.recomparar(silencioso=True)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def abrir_modelo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel modelo exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def factura_duplicada_en_sesion(self, datos_factura):
        clave = datos_factura.get("clave", "")
        return clave in self.claves_facturas

    def confirmar_factura_duplicada(self, datos, fuente):
        mensaje = (
            f"Esta factura ya fue cargada ({fuente}).\n\n"
            f"Proveedor: {datos.get('proveedor', '')}\n"
            f"CUIT: {datos.get('cuit_proveedor', '')}\n"
            f"Tipo: {datos.get('tipo', '')}\n"
            f"Número: {datos.get('numero_completo', '')}\n"
            f"Fecha: {datos.get('fecha', '')}\n"
            f"Archivo: {datos.get('archivo', '')}\n\n"
            "¿Querés cargarla igualmente?"
        )

        return messagebox.askyesno("Factura duplicada", mensaje)

    def agregar_facturas(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar una o varias facturas",
            filetypes=[
                ("Facturas", "*.pdf *.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                ("Todos", "*.*")
            ]
        )

        if not rutas:
            return

        total_items = 0
        cargadas = 0
        salteadas = 0

        for ruta in rutas:
            try:
                datos_previos = self.reader.detectar_datos_factura(ruta)

                if self.factura_duplicada_en_sesion(datos_previos):
                    if not self.confirmar_factura_duplicada(datos_previos, "en esta sesión"):
                        salteadas += 1
                        continue

                historial = self.base.factura_existe(datos_previos.get("clave", ""))
                if historial:
                    fecha_proc, archivo_resumen = historial
                    datos_previos["historial_info"] = f"Procesada el {fecha_proc}. Archivo: {archivo_resumen or ''}"
                    if not self.confirmar_factura_duplicada(datos_previos, "historial permanente"):
                        salteadas += 1
                        continue

                items, debug, datos_factura = self.reader.analizar(ruta)

                self.resultados_debug.extend(debug)
                self.facturas_cargadas.append(datos_factura)
                self.claves_facturas.add(datos_factura["clave"])

                for item in items:
                    self.insertar_item(item)

                total_items += len(items)
                cargadas += 1

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo procesar:\n{ruta}\n\n{e}")

        if self.catalogo.articulos:
            self.recomparar(silencioso=True)

        self.actualizar_label_facturas()
        self.aplicar_filtro()

        messagebox.showinfo(
            "Lectura terminada",
            f"Facturas agregadas: {cargadas}\n"
            f"Facturas salteadas por duplicadas: {salteadas}\n"
            f"Líneas candidatas agregadas: {total_items}"
        )

    def sacar_foto(self):
        if cv2 is None:
            messagebox.showerror("Error", "Falta instalar opencv-python")
            return

        cap = cv2.VideoCapture(0)

        if not cap.isOpened():
            messagebox.showerror("Error", "No se pudo abrir la cámara.")
            return

        messagebox.showinfo("Cámara", "ESPACIO captura / ESC cancela")
        ruta_foto = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            cv2.imshow("Capturar factura", frame)
            key = cv2.waitKey(1)

            if key == 27:
                break

            if key == 32:
                ruta_foto = os.path.join(tempfile.gettempdir(), f"factura_capturada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
                cv2.imwrite(ruta_foto, frame)
                break

        cap.release()
        cv2.destroyAllWindows()

        if ruta_foto:
            self.procesar_ruta_unica(ruta_foto)

    def procesar_ruta_unica(self, ruta):
        try:
            datos_previos = self.reader.detectar_datos_factura(ruta)

            if self.factura_duplicada_en_sesion(datos_previos):
                if not self.confirmar_factura_duplicada(datos_previos, "en esta sesión"):
                    return

            historial = self.base.factura_existe(datos_previos.get("clave", ""))
            if historial:
                if not self.confirmar_factura_duplicada(datos_previos, "historial permanente"):
                    return

            items, debug, datos_factura = self.reader.analizar(ruta)

            self.resultados_debug.extend(debug)
            self.facturas_cargadas.append(datos_factura)
            self.claves_facturas.add(datos_factura["clave"])

            for item in items:
                self.insertar_item(item)

            if self.catalogo.articulos:
                self.recomparar(silencioso=True)

            self.actualizar_label_facturas()
            self.aplicar_filtro()

            messagebox.showinfo("Foto procesada", f"Líneas candidatas agregadas: {len(items)}")

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def insertar_item(self, item):
    valores = (
        CHECK_ON if item["usar"] else CHECK_OFF,
        item.get("factura_etiqueta", ""),
        item.get("proveedor_factura", ""),
        item.get("cuit_proveedor", ""),
        item.get("tipo_factura", ""),
        item.get("numero_completo", ""),
        item.get("fecha_factura", ""),
        item["existe_tango"],
        item["

        tag = self.tag_para_valores(valores)
        self.tabla.insert("", "end", values=valores, tags=(tag,))

    def tag_para_valores(self, v):
        usar = v[0]
        existe = v[7]
        codigo_tango = v[9]
        advertencia = v[17]
        match = v[18]
        conf = v[19]

        if usar == CHECK_OFF:
            return "gris"

        if advertencia and advertencia != "OK":
            return "advertencia"

        if match == "Equivalencia aprendida":
            return "azul"

        try:
            conf_int = int(conf)
        except Exception:
            conf_int = 0

        if existe == "Sí" and conf_int >= 90:
            return "verde"

        if existe == "Sí" and conf_int < 90:
            return "amarillo"

        if existe == "No" or not codigo_tango:
            return "rojo"

        return ""

    def actualizar_tags(self):
        for item_id in self.tabla.get_children():
            v = self.tabla.item(item_id, "values")
            tag = self.tag_para_valores(v)
            self.tabla.item(item_id, tags=(tag,))

    def recomparar(self, silencioso=False):
        if not self.catalogo.articulos:
            if not silencioso:
                messagebox.showinfo("Catálogo no cargado", "Primero cargá el Excel Artículos.xlsx exportado desde Tango.")
            return

        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))

            cuit = v[3]
            codigo_proveedor = v[11]
            codigo_barras = v[12]
            descripcion = v[10]

            eq = self.base.buscar_equivalencia(cuit, codigo_proveedor)

            if eq:
                v[7] = "Sí"
                v[8] = "Cargar stock / compra"
                v[9] = eq["codigo_tango"]
                v[10] = eq["descripcion_tango"] or descripcion
                v[12] = eq["codigo_barras"]
                v[18] = "Equivalencia aprendida"
                v[19] = "100"
                self.tabla.item(item_id, values=v)
                continue

            art, metodo, score = self.catalogo.buscar(
                codigo_proveedor=codigo_proveedor,
                codigo_barras=codigo_barras,
                descripcion=descripcion
            )

            if art:
                v[7] = "Sí"
                v[8] = "Cargar stock / compra"
                v[9] = art["codigo"]
                v[10] = art["descripcion"]
                v[12] = art["codigo_barras"]
                v[18] = metodo
                v[19] = str(score)
            else:
                v[7] = "No"
                v[8] = "Alta artículo"
                v[18] = ""
                v[19] = ""

            self.tabla.item(item_id, values=v)

        self.actualizar_tags()
        self.aplicar_filtro()

        if not silencioso:
            messagebox.showinfo("Listo", "Comparación terminada.")

    def doble_click(self, event):
        item_id = self.tabla.focus()

        if not item_id:
            return

        col = self.tabla.identify_column(event.x)
        idx = int(col.replace("#", "")) - 1

        valores = list(self.tabla.item(item_id, "values"))

        if idx == 0:
            valores[0] = CHECK_OFF if valores[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=valores)
            self.actualizar_tags()
            self.aplicar_filtro()
            return

        nombres = [
            "Usar", "Factura", "Proveedor", "CUIT", "Tipo", "Número", "Fecha",
            "Existe", "Acción", "Código Tango", "Descripción", "Cod. proveedor",
            "Cód. barras", "Cantidad", "Precio", "Importe", "IVA %", "Advertencia",
            "Coincidencia", "Confianza", "Tipo art.", "Escala", "Página",
            "Ruta factura", "Línea detectada"
        ]

        actual = valores[idx]
        nuevo = simpledialog.askstring("Editar", nombres[idx], initialvalue=actual)

        if nuevo is not None:
            valores[idx] = nuevo
            self.tabla.item(item_id, values=valores)
            self.actualizar_tags()

    def seleccionar_todo(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_ON
            self.tabla.item(item_id, values=v)
        self.actualizar_tags()
        self.aplicar_filtro()

    def borrar_seleccion(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_OFF
            self.tabla.item(item_id, values=v)
        self.actualizar_tags()
        self.aplicar_filtro()

    def invertir_seleccion(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_OFF if v[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=v)
        self.actualizar_tags()
        self.aplicar_filtro()

    def agregar_manual(self):
        item = {
            "usar": True,
            "factura_etiqueta": "Manual",
            "proveedor_factura": "",
            "cuit_proveedor": "",
            "tipo_factura": "",
            "numero_completo": "",
            "fecha_factura": "",
            "existe_tango": "No",
            "accion": "Alta artículo",
            "codigo_tango": "",
            "descripcion": "",
            "desc_adic": "",
            "sinonimo": "",
            "codigo_barras": "",
            "cantidad": "",
            "precio": "",
            "importe": "",
            "iva": "",
            "advertencia": "",
            "tipo": "Simple",
            "escala": "No usa",
            "pagina": "",
            "ruta_factura": "",
            "linea_detectada": "Manual",
            "coincidencia": "",
            "match_score": "",
        }
        self.insertar_item(item)

    def eliminar_fila(self):
        for item_id in self.tabla.selection():
            self.tabla.delete(item_id)

    def limpiar_todo(self):
        if not messagebox.askyesno("Confirmar", "¿Querés borrar todas las facturas y filas cargadas?"):
            return

        for item_id in self.tabla.get_children():
            self.tabla.delete(item_id)

        self.facturas_cargadas = []
        self.resultados_debug = []
        self.claves_facturas = set()
        self.actualizar_label_facturas()

    def aplicar_filtro(self):
        filtro = self.filtro_actual.get()

        for item_id in self.tabla.get_children(""):
            self.tabla.detach(item_id)

        for item_id in self.tabla.get_children("detached"):
            pass

        # Tkinter no ofrece forma directa de listar detached.
        # Para simplificar, reconstruimos desde todos los hijos guardados internamente usando move.
        # En muchas versiones los detached siguen accesibles si tenemos sus ids: usamos una lista persistente.
        if not hasattr(self, "_todos_items"):
            self._todos_items = list(self.tabla.get_children(""))

        ids_actuales = list(self.tabla.get_children("")) + getattr(self, "_todos_items", [])
        ids_unicos = []
        for i in ids_actuales:
            if i not in ids_unicos:
                ids_unicos.append(i)

        self._todos_items = ids_unicos

        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
            except Exception:
                continue

            if not v:
                continue

            mostrar = self.pasa_filtro(v, filtro)

            if mostrar:
                try:
                    self.tabla.move(item_id, "", "end")
                except Exception:
                    pass
            else:
                try:
                    self.tabla.detach(item_id)
                except Exception:
                    pass

    def pasa_filtro(self, v, filtro):
        usar = v[0]
        existe = v[7]
        codigo = v[9]
        advertencia = v[17]
        conf = v[19]

        try:
            conf_int = int(conf)
        except Exception:
            conf_int = 0

        if filtro == "Ver todos":
            return True
        if filtro == "Solo seleccionados":
            return usar == CHECK_ON
        if filtro == "Solo nuevos":
            return existe == "No"
        if filtro == "Solo existentes":
            return existe == "Sí"
        if filtro == "Sin código Tango":
            return not codigo
        if filtro == "Baja confianza":
            return existe == "Sí" and conf_int < 90
        if filtro == "Destildados":
            return usar == CHECK_OFF
        if filtro == "Con advertencias":
            return bool(advertencia and advertencia != "OK")

        return True

    def registrar_item_en_lista_filtro(self):
        self._todos_items = list(self.tabla.get_children(""))

    def actualizar_label_facturas(self):
        nombres = []

        for f in self.facturas_cargadas:
            etiqueta = f.get("numero_completo") or f.get("archivo", "")
            proveedor = f.get("proveedor", "")
            if proveedor:
                etiqueta = f"{proveedor[:25]} {etiqueta}"
            nombres.append(etiqueta)

        if len(nombres) <= 4:
            detalle = " | ".join(nombres)
        else:
            detalle = " | ".join(nombres[:4]) + f" | ... (+{len(nombres) - 4})"

        self.lbl_facturas.config(text=f"Facturas cargadas: {len(self.facturas_cargadas)}   {detalle}")

    def ver_debug(self):
        if not self.resultados_debug:
            messagebox.showerror("Error", "Primero agregá una factura.")
            return

        for archivo, pagina, debug_path, lineas, zonas in self.resultados_debug:
            try:
                os.startfile(debug_path)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return

    def abrir_factura_original(self):
        item_id = self.tabla.focus()

        if not item_id:
            messagebox.showerror("Error", "Seleccioná una fila.")
            return

        v = self.tabla.item(item_id, "values")
        ruta = v[23]

        if not ruta or not os.path.exists(ruta):
            messagebox.showerror("Error", "No se encontró la ruta de la factura original.")
            return

        os.startfile(ruta)

    def leer_filas(self, solo_visibles=False):
        if solo_visibles:
            ids = self.tabla.get_children("")
        else:
            ids = getattr(self, "_todos_items", None) or self.tabla.get_children("")

        filas = []

        for item_id in ids:
            try:
                v = self.tabla.item(item_id, "values")
            except Exception:
                continue

            if not v:
                continue

            filas.append({
                "usar": v[0],
                "factura": v[1],
                "proveedor": v[2],
                "cuit": v[3],
                "tipo_comprobante": v[4],
                "numero_comprobante": v[5],
                "fecha": v[6],
                "existe": v[7],
                "accion": v[8],
                "codigo_tango": v[9].strip(),
                "descripcion": v[10].strip(),
                "sinonimo": v[11].strip(),
                "codigo_barras": v[12].strip(),
                "cantidad": v[13].strip(),
                "precio": v[14].strip(),
                "importe": v[15].strip(),
                "iva": v[16].strip(),
                "advertencia": v[17].strip(),
                "coincidencia": v[18].strip(),
                "match_score": v[19].strip(),
                "tipo": v[20].strip() or "Simple",
                "escala": v[21].strip() or "No usa",
                "pagina": v[22],
                "ruta": v[23],
                "linea": v[24],
            })

        return filas

    def deduplicar_nuevos(self, nuevos):
        normalizador = CatalogoTango()
        vistos = {}
        resultado = []
        duplicados = []

        for f in nuevos:
            clave = (
                normalizador.normalizar(f["sinonimo"]),
                normalizador.normalizar(f["codigo_barras"]),
                normalizador.normalizar(f["descripcion"]),
            )

            if clave in vistos:
                duplicados.append((f, vistos[clave]))
                continue

            vistos[clave] = f
            resultado.append(f)

        return resultado, duplicados

    def aprender_equivalencia_seleccionada(self):
        seleccion = self.tabla.selection()

        if not seleccion:
            messagebox.showerror("Error", "Seleccioná una fila para aprender la equivalencia.")
            return

        aprendidas = 0

        for item_id in seleccion:
            v = self.tabla.item(item_id, "values")

            cuit = v[3]
            proveedor = v[2]
            codigo_proveedor = v[11]
            descripcion_proveedor = v[24]
            codigo_tango = v[9]
            descripcion_tango = v[10]
            codigo_barras = v[12]

            if not cuit or not codigo_proveedor or not codigo_tango:
                continue

            self.base.guardar_equivalencia(
                cuit,
                proveedor,
                codigo_proveedor,
                descripcion_proveedor,
                codigo_tango,
                descripcion_tango,
                codigo_barras
            )

            aprendidas += 1

        messagebox.showinfo("Listo", f"Equivalencias aprendidas: {aprendidas}")
        self.recomparar(silencioso=True)

    def aprender_equivalencias_automaticas(self):
        filas = self.leer_filas()

        for f in filas:
            if f["usar"] != CHECK_ON:
                continue
            if not f["cuit"] or not f["sinonimo"] or not f["codigo_tango"]:
                continue

            self.base.guardar_equivalencia(
                f["cuit"],
                f["proveedor"],
                f["sinonimo"],
                f["linea"],
                f["codigo_tango"],
                f["descripcion"],
                f["codigo_barras"]
            )

    def generar_excel_alta_articulos(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo Tango.")
            return

        filas = self.leer_filas()

        nuevos = [
            f for f in filas
            if f["usar"] == CHECK_ON and f["existe"] == "No"
        ]

        if not nuevos:
            messagebox.showinfo("Sin artículos nuevos", "No hay artículos nuevos para dar de alta.")
            return

        faltan = [f for f in nuevos if not f["codigo_tango"]]

        if faltan:
            messagebox.showerror(
                "Faltan códigos Tango",
                "Hay artículos nuevos sin Código Tango.\n\nCompletalos antes de generar el Excel."
            )
            return

        nuevos_unicos, duplicados = self.deduplicar_nuevos(nuevos)

        if duplicados:
            messagebox.showinfo(
                "Duplicados agrupados",
                f"Se encontraron {len(duplicados)} artículos nuevos repetidos.\n"
                f"Se exportará solo una vez cada artículo nuevo."
            )

        salida = filedialog.asksaveasfilename(
            title="Guardar Excel Alta Artículos Tango",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Alta_Articulos_Tango.xlsx"
        )

        if not salida:
            return

        try:
            wb = load_workbook(self.ruta_modelo_tango)

            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja 'Artículos'.")
                return

            ws = wb["Artículos"]

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for fila, item in enumerate(nuevos_unicos, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = ""
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]

            wb.save(salida)
            self.aprender_equivalencias_automaticas()

            messagebox.showinfo(
                "Listo",
                f"Excel de alta de artículos generado:\n\n{salida}\n\n"
                f"Artículos nuevos únicos: {len(nuevos_unicos)}"
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def generar_resumen_compra_stock(self):
        filas = [
            f for f in self.leer_filas()
            if f["usar"] == CHECK_ON
        ]

        if not filas:
            messagebox.showerror("Error", "No hay filas seleccionadas.")
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar resumen compra/stock",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Resumen_Compra_Stock.xlsx"
        )

        if not salida:
            return

        self.crear_excel_resumen(salida, filas)

        for f in self.facturas_cargadas:
            self.base.guardar_factura_procesada(f, archivo_resumen=salida)

        self.aprender_equivalencias_automaticas()

        messagebox.showinfo(
            "Listo",
            f"Resumen de compra/stock generado:\n\n{salida}\n\n"
            f"Facturas registradas en historial permanente: {len(self.facturas_cargadas)}"
        )

    def crear_excel_resumen(self, salida, filas):
        wb = Workbook()
        bold = Font(bold=True)

        ws = wb.active
        ws.title = "Compra_Stock"

        headers = [
            "Factura", "Proveedor", "CUIT", "Tipo comprobante", "Número comprobante", "Fecha",
            "Existe en Tango", "Acción", "Código Tango", "Descripción Tango/OCR",
            "Código proveedor", "Código barras", "Cantidad", "Precio", "Importe", "IVA %",
            "Advertencia", "Coincidencia", "Confianza", "Línea detectada"
        ]

        for col, h in enumerate(headers, start=1):
            ws.cell(1, col).value = h
            ws.cell(1, col).font = bold

        for row, f in enumerate(filas, start=2):
            values = [
                f["factura"], f["proveedor"], f["cuit"], f["tipo_comprobante"],
                f["numero_comprobante"], f["fecha"], f["existe"], f["accion"],
                f["codigo_tango"], f["descripcion"], f["sinonimo"], f["codigo_barras"],
                f["cantidad"], f["precio"], f["importe"], f["iva"], f["advertencia"],
                f["coincidencia"], f["match_score"], f["linea"]
            ]

            for col, value in enumerate(values, start=1):
                ws.cell(row, col).value = value

        widths = [25, 35, 15, 18, 18, 12, 16, 22, 18, 50, 18, 18, 12, 12, 12, 10, 30, 30, 12, 80]

        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width

        ws2 = wb.create_sheet("Facturas_Cargadas")

        fact_headers = [
            "Clave", "Archivo", "Proveedor detectado", "CUIT proveedor", "Tipo",
            "Punto de venta", "Número", "Número completo", "Fecha", "Total", "Ruta"
        ]

        for col, h in enumerate(fact_headers, start=1):
            ws2.cell(1, col).value = h
            ws2.cell(1, col).font = bold

        for row, f in enumerate(self.facturas_cargadas, start=2):
            ws2.cell(row, 1).value = f["clave"]
            ws2.cell(row, 2).value = f["archivo"]
            ws2.cell(row, 3).value = f["proveedor"]
            ws2.cell(row, 4).value = f["cuit_proveedor"]
            ws2.cell(row, 5).value = f["tipo"]
            ws2.cell(row, 6).value = f["punto_venta"]
            ws2.cell(row, 7).value = f["numero"]
            ws2.cell(row, 8).value = f["numero_completo"]
            ws2.cell(row, 9).value = f["fecha"]
            ws2.cell(row, 10).value = f.get("total", "")
            ws2.cell(row, 11).value = f["ruta"]

        for col in range(1, 12):
            ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 25

        wb.save(salida)

    def generar_reporte_revision(self):
        filas = self.leer_filas()

        if not filas:
            messagebox.showerror("Error", "No hay filas para reportar.")
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar reporte de revisión",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Reporte_Revision.xlsx"
        )

        if not salida:
            return

        wb = Workbook()
        bold = Font(bold=True)

        ws = wb.active
        ws.title = "Resumen"

        total = len(filas)
        seleccionados = len([f for f in filas if f["usar"] == CHECK_ON])
        nuevos = len([f for f in filas if f["existe"] == "No"])
        existentes = len([f for f in filas if f["existe"] == "Sí"])
        sin_codigo = len([f for f in filas if not f["codigo_tango"]])
        advertencias = len([f for f in filas if f["advertencia"] and f["advertencia"] != "OK"])

        resumen = [
            ("Filas totales", total),
            ("Filas seleccionadas", seleccionados),
            ("Artículos existentes", existentes),
            ("Artículos nuevos", nuevos),
            ("Sin código Tango", sin_codigo),
            ("Advertencias cantidad/precio/importe", advertencias),
            ("Facturas cargadas", len(self.facturas_cargadas)),
        ]

        for row, (k, v) in enumerate(resumen, start=1):
            ws.cell(row, 1).value = k
            ws.cell(row, 1).font = bold
            ws.cell(row, 2).value = v

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20

        ws2 = wb.create_sheet("Detalle")
        headers = [
            "Usar", "Factura", "Proveedor", "CUIT", "Número", "Existe", "Acción",
            "Código Tango", "Descripción", "Código proveedor", "Cantidad", "Precio",
            "Importe", "IVA", "Advertencia", "Coincidencia", "Confianza"
        ]

        for col, h in enumerate(headers, start=1):
            ws2.cell(1, col).value = h
            ws2.cell(1, col).font = bold

        for row, f in enumerate(filas, start=2):
            vals = [
                f["usar"], f["factura"], f["proveedor"], f["cuit"], f["numero_comprobante"],
                f["existe"], f["accion"], f["codigo_tango"], f["descripcion"], f["sinonimo"],
                f["cantidad"], f["precio"], f["importe"], f["iva"], f["advertencia"],
                f["coincidencia"], f["match_score"]
            ]
            for col, val in enumerate(vals, start=1):
                ws2.cell(row, col).value = val

        for col in range(1, len(headers) + 1):
            ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 22

        wb.save(salida)

        messagebox.showinfo("Listo", f"Reporte de revisión generado:\n\n{salida}")

    def guardar_sesion(self):
        salida = filedialog.asksaveasfilename(
            title="Guardar sesión",
            defaultextension=".json",
            filetypes=[("JSON", "*.json")],
            initialfile=f"Sesion_Revision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        )

        if not salida:
            return

        data = {
            "version": "V8",
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ruta_modelo_tango": self.ruta_modelo_tango,
            "ruta_catalogo_tango": self.ruta_catalogo_tango,
            "facturas_cargadas": self.facturas_cargadas,
            "filas": self.leer_filas()
        }

        with open(salida, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        messagebox.showinfo("Listo", f"Sesión guardada:\n\n{salida}")

    def cargar_sesion(self):
        ruta = filedialog.askopenfilename(
            title="Cargar sesión",
            filetypes=[("JSON", "*.json")]
        )

        if not ruta:
            return

        with open(ruta, "r", encoding="utf-8") as f:
            data = json.load(f)

        self.limpiar_todo_sin_preguntar()

        self.ruta_modelo_tango = data.get("ruta_modelo_tango")
        self.ruta_catalogo_tango = data.get("ruta_catalogo_tango")

        if self.ruta_modelo_tango:
            self.lbl_modelo.config(text=f"Modelo Tango: {self.ruta_modelo_tango}")

        if self.ruta_catalogo_tango and os.path.exists(self.ruta_catalogo_tango):
            try:
                self.catalogo.cargar(self.ruta_catalogo_tango)
                self.lbl_catalogo.config(
                    text=f"Artículos Tango: {self.ruta_catalogo_tango} ({len(self.catalogo.articulos)} artículos)"
                )
            except Exception:
                pass

        self.facturas_cargadas = data.get("facturas_cargadas", [])
        self.claves_facturas = set(f.get("clave", "") for f in self.facturas_cargadas)

        for f in data.get("filas", []):
            item = {
                "usar": f.get("usar") == CHECK_ON,
                "factura_etiqueta": f.get("factura", ""),
                "proveedor_factura": f.get("proveedor", ""),
                "cuit_proveedor": f.get("cuit", ""),
                "tipo_factura": f.get("tipo_comprobante", ""),
                "numero_completo": f.get("numero_comprobante", ""),
                "fecha_factura": f.get("fecha", ""),
                "existe_tango": f.get("existe", "No"),
                "accion": f.get("accion", ""),
                "codigo_tango": f.get("codigo_tango", ""),
                "descripcion": f.get("descripcion", ""),
                "sinonimo": f.get("sinonimo", ""),
                "codigo_barras": f.get("codigo_barras", ""),
                "cantidad": f.get("cantidad", ""),
                "precio": f.get("precio", ""),
                "importe": f.get("importe", ""),
                "iva": f.get("iva", ""),
                "advertencia": f.get("advertencia", ""),
                "tipo": f.get("tipo", "Simple"),
                "escala": f.get("escala", "No usa"),
                "pagina": f.get("pagina", ""),
                "ruta_factura": f.get("ruta", ""),
                "linea_detectada": f.get("linea", ""),
                "coincidencia": f.get("coincidencia", ""),
                "match_score": f.get("match_score", ""),
            }
            self.insertar_item(item)

        self.actualizar_label_facturas()
        self.actualizar_tags()
        self._todos_items = list(self.tabla.get_children(""))
        self.aplicar_filtro()

        messagebox.showinfo("Listo", "Sesión cargada correctamente.")

    def limpiar_todo_sin_preguntar(self):
        for item_id in self.tabla.get_children():
            self.tabla.delete(item_id)

        self.facturas_cargadas = []
        self.resultados_debug = []
        self.claves_facturas = set()
        self._todos_items = []
        self.actualizar_label_facturas()

    def limpiar_todo(self):
        if not messagebox.askyesno("Confirmar", "¿Querés borrar todas las facturas y filas cargadas?"):
            return

        self.limpiar_todo_sin_preguntar()


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()