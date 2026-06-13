import os, re, json, sqlite3, tempfile, unicodedata
import tkinter as tk
import threading
import queue
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from dataclasses import dataclass
from difflib import SequenceMatcher
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from PIL import Image, ImageOps, ImageEnhance, ImageDraw
import fitz
import pytesseract
import numpy as np

try:
    import cv2
except Exception:
    cv2 = None

# Ajustar si Tesseract está en otra ruta
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

APP_DB = "facturas_tango_v8.db"
APP_VERSION = "V9.0"
CHECK_ON = "☑"
CHECK_OFF = "☐"
UMBRAL_CONFIANZA_DEFAULT = 82


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    page: int


@dataclass
class LineBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    words: list
    score: int = 0


class BaseLocal:
    def __init__(self, db_path=APP_DB):
        self.db_path = db_path
        self.inicializar()

    def conectar(self):
        return sqlite3.connect(self.db_path)

    def inicializar(self):
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS facturas_procesadas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clave TEXT UNIQUE,
                proveedor TEXT,
                cuit TEXT,
                tipo TEXT,
                punto_venta TEXT,
                numero TEXT,
                numero_completo TEXT,
                fecha TEXT,
                archivo TEXT,
                ruta TEXT,
                fecha_proceso TEXT,
                archivo_resumen TEXT,
                archivo_alta TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS equivalencias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cuit TEXT,
                proveedor TEXT,
                codigo_proveedor TEXT,
                descripcion_proveedor TEXT,
                codigo_tango TEXT,
                descripcion_tango TEXT,
                codigo_barras TEXT,
                veces_usado INTEGER DEFAULT 1,
                ultima_fecha TEXT,
                UNIQUE(cuit, codigo_proveedor)
            )
        """)
        # Layout aprendido por proveedor (CUIT -> zonas como % del tamaño de página)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS proveedor_layout (
                cuit TEXT PRIMARY KEY,
                proveedor TEXT,
                page_w REAL,
                page_h REAL,
                zona_items_y0_pct REAL,
                zona_items_y1_pct REAL,
                zona_items_x0_pct REAL,
                zona_items_x1_pct REAL,
                es_pdf_nativo INTEGER DEFAULT 0,
                necesita_ocr INTEGER DEFAULT 1,
                veces_procesado INTEGER DEFAULT 1,
                ultima_fecha TEXT
            )
        """)
        # Cache de datos del QR AFIP por CUIT
        cur.execute("""
            CREATE TABLE IF NOT EXISTS qr_cache (
                cuit TEXT PRIMARY KEY,
                proveedor TEXT,
                ultima_fecha TEXT,
                datos_json TEXT
            )
        """)
        con.commit()
        con.close()

    def factura_existe(self, clave):
        if not clave:
            return None
        con = self.conectar()
        cur = con.cursor()
        cur.execute("SELECT fecha_proceso, archivo_resumen FROM facturas_procesadas WHERE clave=?", (clave,))
        row = cur.fetchone()
        con.close()
        return row

    def guardar_factura_procesada(self, datos, archivo_resumen="", archivo_alta=""):
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            INSERT OR REPLACE INTO facturas_procesadas (
                clave, proveedor, cuit, tipo, punto_venta, numero, numero_completo,
                fecha, archivo, ruta, fecha_proceso, archivo_resumen, archivo_alta
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datos.get("clave", ""), datos.get("proveedor", ""), datos.get("cuit_proveedor", ""),
            datos.get("tipo", ""), datos.get("punto_venta", ""), datos.get("numero", ""),
            datos.get("numero_completo", ""), datos.get("fecha", ""), datos.get("archivo", ""),
            datos.get("ruta", ""), datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            archivo_resumen, archivo_alta
        ))
        con.commit()
        con.close()

    def guardar_layout(self, cuit, proveedor, page_w, page_h,
                       y0_pct, y1_pct, x0_pct, x1_pct,
                       es_pdf_nativo, necesita_ocr):
        if not cuit:
            return
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            INSERT INTO proveedor_layout (
                cuit, proveedor, page_w, page_h,
                zona_items_y0_pct, zona_items_y1_pct,
                zona_items_x0_pct, zona_items_x1_pct,
                es_pdf_nativo, necesita_ocr,
                veces_procesado, ultima_fecha
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(cuit) DO UPDATE SET
                proveedor=excluded.proveedor,
                page_w=excluded.page_w,
                page_h=excluded.page_h,
                zona_items_y0_pct=excluded.zona_items_y0_pct,
                zona_items_y1_pct=excluded.zona_items_y1_pct,
                zona_items_x0_pct=excluded.zona_items_x0_pct,
                zona_items_x1_pct=excluded.zona_items_x1_pct,
                es_pdf_nativo=excluded.es_pdf_nativo,
                necesita_ocr=excluded.necesita_ocr,
                veces_procesado=veces_procesado+1,
                ultima_fecha=excluded.ultima_fecha
        """, (
            str(cuit).strip(), str(proveedor).strip(),
            float(page_w), float(page_h),
            float(y0_pct), float(y1_pct),
            float(x0_pct), float(x1_pct),
            int(es_pdf_nativo), int(necesita_ocr),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        con.commit()
        con.close()

    def obtener_layout(self, cuit):
        if not cuit:
            return None
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            SELECT proveedor, page_w, page_h,
                   zona_items_y0_pct, zona_items_y1_pct,
                   zona_items_x0_pct, zona_items_x1_pct,
                   es_pdf_nativo, necesita_ocr, veces_procesado
            FROM proveedor_layout WHERE cuit=?
        """, (str(cuit).strip(),))
        row = cur.fetchone()
        con.close()
        if not row:
            return None
        return {
            "proveedor": row[0], "page_w": row[1], "page_h": row[2],
            "y0_pct": row[3], "y1_pct": row[4],
            "x0_pct": row[5], "x1_pct": row[6],
            "es_pdf_nativo": bool(row[7]), "necesita_ocr": bool(row[8]),
            "veces_procesado": row[9],
        }

    def guardar_qr_cache(self, cuit, proveedor, datos_json):
        if not cuit:
            return
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            INSERT INTO qr_cache (cuit, proveedor, ultima_fecha, datos_json)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(cuit) DO UPDATE SET
                proveedor=excluded.proveedor,
                ultima_fecha=excluded.ultima_fecha,
                datos_json=excluded.datos_json
        """, (
            str(cuit).strip(), str(proveedor).strip(),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            json.dumps(datos_json, ensure_ascii=False)
        ))
        con.commit()
        con.close()

    def obtener_qr_cache(self, cuit):
        if not cuit:
            return None
        con = self.conectar()
        cur = con.cursor()
        cur.execute("SELECT proveedor, datos_json FROM qr_cache WHERE cuit=?",
                    (str(cuit).strip(),))
        row = cur.fetchone()
        con.close()
        if not row:
            return None
        try:
            return {"proveedor": row[0], "datos": json.loads(row[1])}
        except Exception:
            return {"proveedor": row[0], "datos": {}}

    def buscar_equivalencia(self, cuit, codigo_proveedor):
        if not cuit or not codigo_proveedor:
            return None
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            SELECT codigo_tango, descripcion_tango, codigo_barras
            FROM equivalencias
            WHERE cuit=? AND codigo_proveedor=?
        """, (str(cuit).strip(), str(codigo_proveedor).strip()))
        row = cur.fetchone()
        con.close()
        if not row:
            return None
        return {"codigo_tango": row[0], "descripcion_tango": row[1], "codigo_barras": row[2]}

    def guardar_equivalencia(self, cuit, proveedor, codigo_proveedor, descripcion_proveedor,
                             codigo_tango, descripcion_tango, codigo_barras):
        if not cuit or not codigo_proveedor or not codigo_tango:
            return
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            INSERT INTO equivalencias (
                cuit, proveedor, codigo_proveedor, descripcion_proveedor,
                codigo_tango, descripcion_tango, codigo_barras, veces_usado, ultima_fecha
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(cuit, codigo_proveedor) DO UPDATE SET
                proveedor=excluded.proveedor,
                descripcion_proveedor=excluded.descripcion_proveedor,
                codigo_tango=excluded.codigo_tango,
                descripcion_tango=excluded.descripcion_tango,
                codigo_barras=excluded.codigo_barras,
                veces_usado=veces_usado+1,
                ultima_fecha=excluded.ultima_fecha
        """, (
            str(cuit).strip(), str(proveedor).strip(), str(codigo_proveedor).strip(),
            str(descripcion_proveedor).strip(), str(codigo_tango).strip(),
            str(descripcion_tango).strip(), str(codigo_barras).strip(),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        con.commit()
        con.close()


class Normalizador:
    def quitar_acentos(self, texto):
        texto = unicodedata.normalize("NFD", str(texto))
        return "".join(c for c in texto if unicodedata.category(c) != "Mn")

    def normalizar(self, texto):
        if texto is None:
            return ""
        texto = self.quitar_acentos(str(texto).upper().strip())
        reemplazos = {
            "P/ACERO": "ACERO", "P ACERO": "ACERO", "PACERO": "ACERO",
            "P/ ACERO": "ACERO", "P/INOX": "INOX", "P INOX": "INOX",
            "INOXIDABLE": "INOX", "C/": "CON ", "S/": "SIN ",
            "MM.": "MM", "M.M.": "MM"
        }
        for a, b in reemplazos.items():
            texto = texto.replace(a, b)
        for ch in ["/", "-", ",", ".", "(", ")", "[", "]", "_", ":"]:
            texto = texto.replace(ch, " ")
        return re.sub(r"\s+", " ", texto).strip()

    def tokens(self, texto):
        return [t for t in self.normalizar(texto).split() if len(t) >= 2]


class CatalogoTango(Normalizador):
    def __init__(self):
        self.articulos = []

    def cargar(self, ruta_excel):
        wb = load_workbook(ruta_excel, data_only=True)
        if "Artículos" not in wb.sheetnames:
            raise Exception("El Excel no tiene una hoja llamada 'Artículos'.")
        ws = wb["Artículos"]
        self.articulos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = row[0] if len(row) > 0 else ""
            if not codigo:
                continue
            descripcion = row[1] if len(row) > 1 else ""
            desc_adic = row[2] if len(row) > 2 else ""
            sinonimo = row[3] if len(row) > 3 else ""
            codigo_barras = row[4] if len(row) > 4 else ""
            art = {
                "codigo": str(codigo).strip(),
                "descripcion": str(descripcion or "").strip(),
                "desc_adic": str(desc_adic or "").strip(),
                "sinonimo": str(sinonimo or "").strip(),
                "codigo_barras": str(codigo_barras or "").strip(),
            }
            art["_codigo_n"] = self.normalizar(art["codigo"])
            art["_descripcion_n"] = self.normalizar(art["descripcion"])
            art["_sinonimo_n"] = self.normalizar(art["sinonimo"])
            art["_codigo_barras_n"] = self.normalizar(art["codigo_barras"])
            art["_texto_completo_n"] = self.normalizar(
                f"{art['codigo']} {art['descripcion']} {art['desc_adic']} {art['sinonimo']} {art['codigo_barras']}"
            )
            self.articulos.append(art)

    def similitud(self, a, b):
        a = self.normalizar(a)
        b = self.normalizar(b)
        if not a or not b:
            return 0
        return int(SequenceMatcher(None, a, b).ratio() * 100)

    def score_tokens(self, a, b):
        ta, tb = set(self.tokens(a)), set(self.tokens(b))
        if not ta or not tb:
            return 0
        return int((len(ta.intersection(tb)) / len(ta)) * 100)

    def buscar(self, codigo_proveedor="", codigo_barras="", descripcion="", umbral=UMBRAL_CONFIANZA_DEFAULT):
        """Busca el mejor articulo. Retorna (art, metodo, score)."""
        art, metodo, score = self._buscar_interno(codigo_proveedor, codigo_barras, descripcion, umbral)
        return art, metodo, score

    def buscar_candidatos(self, codigo_proveedor="", codigo_barras="", descripcion="", top=5):
        """Retorna una lista de hasta `top` candidatos ordenados por score descendente."""
        cp = self.normalizar(codigo_proveedor)
        cb = self.normalizar(codigo_barras)
        desc = self.normalizar(descripcion)

        resultados = []

        # Exactos primero
        if cb:
            for art in self.articulos:
                if art["_codigo_barras_n"] and art["_codigo_barras_n"] == cb:
                    resultados.append((art, "Codigo de barras exacto", 100))

        if cp:
            for art in self.articulos:
                if art["_codigo_n"] == cp:
                    resultados.append((art, "Codigo Tango exacto", 100))
            for art in self.articulos:
                if art["_sinonimo_n"] and art["_sinonimo_n"] == cp:
                    resultados.append((art, "Sinonimo exacto", 99))
            for art in self.articulos:
                if cp in art["_descripcion_n"]:
                    resultados.append((art, "Codigo proveedor en descripcion", 96))
                elif cp in art["_texto_completo_n"]:
                    resultados.append((art, "Codigo proveedor contenido", 94))

        # Por similitud de descripcion
        if desc:
            scored = []
            for art in self.articulos:
                score = max(
                    self.similitud(desc, art["_descripcion_n"]),
                    self.score_tokens(desc, art["_descripcion_n"])
                )
                if score >= 50:
                    scored.append((art, "Similitud descripcion", score))
            scored.sort(key=lambda x: -x[2])
            resultados.extend(scored[:top * 2])

        # Deduplicar por codigo Tango, quedarse con mejor score
        vistos = {}
        for art, metodo, score in resultados:
            cod = art["codigo"]
            if cod not in vistos or score > vistos[cod][2]:
                vistos[cod] = (art, metodo, score)

        ordenados = sorted(vistos.values(), key=lambda x: -x[2])
        return ordenados[:top]

    def _buscar_interno(self, codigo_proveedor="", codigo_barras="", descripcion="", umbral=UMBRAL_CONFIANZA_DEFAULT):
        cp = self.normalizar(codigo_proveedor)
        cb = self.normalizar(codigo_barras)
        desc = self.normalizar(descripcion)

        if cb:
            for art in self.articulos:
                if art["_codigo_barras_n"] and art["_codigo_barras_n"] == cb:
                    return art, "Codigo de barras exacto", 100

        if cp:
            for art in self.articulos:
                if art["_codigo_n"] == cp:
                    return art, "Codigo Tango exacto", 100
            for art in self.articulos:
                if art["_sinonimo_n"] and art["_sinonimo_n"] == cp:
                    return art, "Sinonimo exacto", 100
            for art in self.articulos:
                if cp in art["_descripcion_n"]:
                    return art, "Codigo proveedor en descripcion", 96
            for art in self.articulos:
                if cp in art["_texto_completo_n"]:
                    return art, "Codigo proveedor contenido", 94

        mejor, mejor_score = None, 0
        if desc:
            for art in self.articulos:
                score = max(self.similitud(desc, art["_descripcion_n"]), self.score_tokens(desc, art["_descripcion_n"]))
                if score > mejor_score:
                    mejor, mejor_score = art, score
            if mejor and mejor_score >= umbral:
                return mejor, "Similitud descripcion", mejor_score

        return None, "", 0


class FacturaTableReader:
    def __init__(self):
        self.stop_words = [
            "total", "subtotal", "cae", "iva 21", "iva 10", "iva 27",
            "son pesos", "observaciones", "importe neto", "fecha de vto"
        ]

    def is_cuit(self, text):
        clean = str(text).strip()
        if re.fullmatch(r"\d{2}-\d{8}-\d", clean):
            return True
        digits = re.sub(r"\D", "", clean)
        return len(digits) == 11 and digits[:2] in ["20", "23", "24", "27", "30", "33", "34"]

    def normalizar_cuit(self, cuit):
        return re.sub(r"\D", "", str(cuit or ""))

    def abrir_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()
        return self.leer_pdf(ruta) if ext == ".pdf" else self.leer_imagen(ruta)

    def _log(self, msg):
        """Escribe al archivo de log para diagnostico."""
        try:
            log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "facturas_tango_debug.log")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")
        except Exception:
            pass

    # Palabras comunes en facturas argentinas que confirman texto legible
    _PALABRAS_FACTURA = {
        "factura", "remito", "total", "subtotal", "cantidad", "precio", "importe",
        "fecha", "cuit", "iva", "neto", "codigo", "descripcion", "proveedor",
        "cliente", "srl", "s.a", "s.r.l", "responsable", "inscripto", "pesos",
        "unitario", "descuento", "original", "duplicado", "numero", "venta",
        "compra", "stock", "articulo", "producto", "unidad", "bulto", "caja",
        "the", "and", "invoice", "total", "date", "item", "unit", "price",
        "emapi", "orange", "blue", "import", "export", "ferreteria", "metalurgica",
    }

    def _texto_es_legible(self, palabras_raw):
        """
        Detecta si el texto extraido por PyMuPDF es legible.
        Estrategia robusta: busca palabras reales del vocabulario de facturas.
        Si no encuentra ninguna entre las primeras 120 palabras, es encoding roto.
        """
        if not palabras_raw:
            self._log("_texto_es_legible: sin palabras -> False")
            return False

        muestra = [str(w[4]).strip().lower() for w in palabras_raw[:120] if str(w[4]).strip()]
        if not muestra:
            self._log("_texto_es_legible: muestra vacia -> False")
            return False

        self._log(f"_texto_es_legible: {len(muestra)} palabras. Primeras 10: {[w[:20] for w in muestra[:10]]}")

        # Criterio 1: patron (cid: explicito
        con_cid = sum(1 for t in muestra if "(cid:" in t)
        self._log(f"  con_cid={con_cid}/{len(muestra)}")
        if con_cid / len(muestra) > 0.15:
            self._log("  -> ILEGIBLE por cid")
            return False

        # Criterio 2: caracteres no-ASCII excesivos
        con_no_ascii = sum(
            1 for t in muestra
            if len(t) > 0 and sum(1 for c in t if ord(c) > 127) / len(t) > 0.5
        )
        self._log(f"  con_no_ascii={con_no_ascii}/{len(muestra)}")
        if con_no_ascii / len(muestra) > 0.25:
            self._log("  -> ILEGIBLE por no-ascii")
            return False

        # Criterio 3 (el mas robusto): buscar palabras reales del vocabulario de facturas
        # Normalizar quitando puntuacion para comparar
        encontradas = set()
        for t in muestra:
            t_limpio = re.sub(r"[^a-z.]", "", t)
            if t_limpio in self._PALABRAS_FACTURA:
                encontradas.add(t_limpio)

        self._log(f"  palabras_factura_encontradas={encontradas}")
        if len(encontradas) == 0:
            self._log("  -> ILEGIBLE: ninguna palabra de factura reconocida")
            return False

        self._log(f"  -> LEGIBLE ok ({len(encontradas)} palabras reconocidas)")
        return True

    def leer_pdf(self, ruta):
        self._log(f"leer_pdf: {ruta}")
        doc = fitz.open(ruta)
        paginas = []
        for page_index, page in enumerate(doc, start=1):
            # Renderizar siempre la imagen (necesaria para debug y preview)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(tempfile.gettempdir(), f"pagina_factura_{page_index}.png")
            pix.save(img_path)
            img = Image.open(img_path)

            # Intentar extraer texto nativo del PDF
            palabras_raw = page.get_text("words")
            self._log(f"  pagina {page_index}: {len(palabras_raw)} palabras raw")
            if self._texto_es_legible(palabras_raw):
                self._log(f"  -> usando texto nativo")
                palabras = []
                for w in palabras_raw:
                    x0, y0, x1, y1, text = w[:5]
                    text = str(text).strip()
                    if text:
                        palabras.append(WordBox(x0 * 2, y0 * 2, x1 * 2, y1 * 2, text, page_index))
            else:
                self._log(f"  -> usando OCR sobre imagen renderizada")
                palabras = self._ocr_sobre_imagen(img, page_index)
                self._log(f"  -> OCR produjo {len(palabras)} palabras")

            paginas.append((img, palabras))
        return paginas

    def _ocr_sobre_imagen(self, img, page_num):
        """Hace OCR sobre una imagen PIL ya renderizada (para PDFs con encoding roto)."""
        img_ocr = self.preparar_imagen(img)
        try:
            data = pytesseract.image_to_data(
                img_ocr, lang="spa", output_type=pytesseract.Output.DICT, config="--psm 6"
            )
        except Exception:
            data = pytesseract.image_to_data(
                img_ocr, output_type=pytesseract.Output.DICT, config="--psm 6"
            )
        palabras = []
        for i, text in enumerate(data["text"]):
            text = str(text).strip()
            if not text:
                continue
            try:
                conf = int(float(data["conf"][i]))
            except Exception:
                conf = -1
            if conf < 25:
                continue
            x, y, w, h = data["left"][i], data["top"][i], data["width"][i], data["height"][i]
            palabras.append(WordBox(x, y, x + w, y + h, text, page_num))
        return palabras

    def leer_imagen(self, ruta):
        img = Image.open(ruta)
        img_ocr = self.preparar_imagen(img)
        data = pytesseract.image_to_data(
            img_ocr, lang="spa", output_type=pytesseract.Output.DICT, config="--psm 6"
        )
        palabras = []
        for i, text in enumerate(data["text"]):
            text = str(text).strip()
            if not text:
                continue
            try:
                conf = int(float(data["conf"][i]))
            except Exception:
                conf = -1
            if conf < 25:
                continue
            x, y, w, h = data["left"][i], data["top"][i], data["width"][i], data["height"][i]
            palabras.append(WordBox(x, y, x + w, y + h, text, 1))
        return [(img, palabras)]

    def preparar_imagen(self, img):
        """
        Preprocesamiento avanzado de imagen para OCR.
        Usa OpenCV si esta disponible (deskew, denoising, binarizacion adaptiva).
        Si no, usa Pillow con mejoras basicas.
        """
        img = img.convert("RGB")
        if cv2 is not None:
            return self._preparar_imagen_cv2(img)
        else:
            return self._preparar_imagen_pillow(img)

    def _preparar_imagen_pillow(self, img):
        """Preprocesamiento basico con Pillow."""
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Sharpness(img).enhance(1.5)
        img = ImageEnhance.Contrast(img).enhance(2.2)
        return img

    def _preparar_imagen_cv2(self, img_pil):
        """Preprocesamiento completo con OpenCV: denoising, deskew y binarizacion adaptiva."""
        arr = np.array(img_pil)
        arr = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
        gray = cv2.cvtColor(arr, cv2.COLOR_BGR2GRAY)

        # Escalar si la imagen es muy chica
        h, w = gray.shape
        if w < 1000:
            factor = 1800 / w
            gray = cv2.resize(gray, None, fx=factor, fy=factor, interpolation=cv2.INTER_CUBIC)

        # Denoising
        gray = cv2.fastNlMeansDenoising(gray, h=10, templateWindowSize=7, searchWindowSize=21)

        # Deskew
        gray = self._deskew_cv2(gray)

        # Binarizacion adaptiva (maneja iluminacion no uniforme - fotos con sombras)
        binary = cv2.adaptiveThreshold(
            gray, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            blockSize=31, C=10
        )

        # Limpiar puntos aislados
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 1))
        binary = cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)

        return Image.fromarray(binary)

    def _deskew_cv2(self, gray):
        """Detecta y corrige la inclinacion de la imagen (hasta ~15 grados)."""
        try:
            edges = cv2.Canny(gray, 50, 150, apertureSize=3)
            lines = cv2.HoughLines(edges, 1, np.pi / 180, threshold=200)
            if lines is None or len(lines) < 5:
                return gray
            angles = []
            for line in lines[:40]:
                theta = line[0][1]
                angle = np.degrees(theta) - 90
                if -15 < angle < 15:
                    angles.append(angle)
            if not angles:
                return gray
            median_angle = float(np.median(angles))
            if abs(median_angle) < 0.3:
                return gray
            h, w = gray.shape
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, median_angle, 1.0)
            return cv2.warpAffine(gray, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
        except Exception:
            return gray

    def agrupar_lineas(self, palabras):
        palabras = sorted(palabras, key=lambda w: (w.y0, w.x0))
        lineas, actual = [], []
        for p in palabras:
            if not actual:
                actual = [p]
                continue
            y_prom = sum(w.y0 for w in actual) / len(actual)
            if abs(p.y0 - y_prom) <= 8:
                actual.append(p)
            else:
                lineas.append(self.crear_linea(actual))
                actual = [p]
        if actual:
            lineas.append(self.crear_linea(actual))
        return lineas

    def crear_linea(self, palabras):
        palabras = sorted(palabras, key=lambda w: w.x0)
        return LineBox(
            min(w.x0 for w in palabras), min(w.y0 for w in palabras),
            max(w.x1 for w in palabras), max(w.y1 for w in palabras),
            " ".join(w.text for w in palabras), palabras
        )

    def parse_decimal(self, t):
        t = str(t).replace("$", "").replace(" ", "").strip()
        if not t:
            return None
        if re.match(r"^-?\d{1,3}(\.\d{3})*,\d+$", t):
            t = t.replace(".", "").replace(",", ".")
        elif re.match(r"^-?\d+,\d+$", t):
            t = t.replace(",", ".")
        try:
            return float(t)
        except Exception:
            return None

    def es_numero(self, t):
        t = str(t).replace("$", "").strip()
        if self.is_cuit(t):
            return False
        return bool(re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2,4}$", t)) or bool(re.match(r"^-?\d+([,.]\d+)?$", t))

    def cantidad_columnas(self, linea):
        if len(linea.words) < 3:
            return 0
        gaps = 0
        for a, b in zip(linea.words, linea.words[1:]):
            if b.x0 - a.x1 > 25:
                gaps += 1
        return gaps + 1

    def puntuar_linea(self, linea):
        text = linea.text.lower()
        if any(self.is_cuit(w.text) for w in linea.words):
            linea.score = -10
            return linea.score
        if any(p in text for p in self.stop_words):
            linea.score = -5
            return linea.score
        nums = sum(1 for w in linea.words if self.es_numero(w.text))
        cols = self.cantidad_columnas(linea)
        score = 0
        if cols >= 3: score += 2
        if cols >= 5: score += 2
        if nums >= 2: score += 2
        if nums >= 4: score += 2
        if len(linea.words) >= 5: score += 1
        linea.score = score
        return score

    def detectar_zonas(self, lineas):
        for l in lineas:
            self.puntuar_linea(l)
        candidatas = [l for l in lineas if l.score >= 4]
        if not candidatas:
            return []
        grupos, grupo = [], [candidatas[0]]
        for l in candidatas[1:]:
            if l.y0 - grupo[-1].y1 < 75:
                grupo.append(l)
            else:
                grupos.append(grupo)
                grupo = [l]
        grupos.append(grupo)
        zonas = []
        for g in grupos:
            if len(g) < 2:
                continue
            zonas.append((min(l.x0 for l in g), min(l.y0 for l in g), max(l.x1 for l in g), max(l.y1 for l in g), g))
        return zonas

    def detectar_items(self, lineas):
        zonas = self.detectar_zonas(lineas)
        items = []
        for x0, y0, x1, y1, grupo in zonas:
            for linea in grupo:
                if linea.score < 4 or any(self.is_cuit(w.text) for w in linea.words):
                    continue
                item = self.linea_a_item(linea)
                if item:
                    items.append(item)
        return items, zonas

    def linea_a_item(self, linea):
        codigo_proveedor = ""
        descripcion_parts = []
        numeros_detectados = []
        for w in linea.words:
            t = w.text.strip()
            if self.is_cuit(t):
                return None
            if self.es_numero(t):
                numeros_detectados.append(t)
                continue
            if not codigo_proveedor and re.match(r"^[A-Z0-9][A-Z0-9.\-/]{2,}$", t, re.IGNORECASE):
                codigo_proveedor = t[:15]
                continue
            if t.upper() in ["U", "UN", "KG", "M", "L", "%"]:
                continue
            if re.match(r"^-?\d+([,.]\d+)?%?$", t):
                continue
            descripcion_parts.append(t)

        vals = [(n, self.parse_decimal(n)) for n in numeros_detectados]
        vals = [(n, v) for n, v in vals if v is not None]

        iva = ""
        if len(vals) >= 4:
            posibles = [(n, v) for n, v in vals if abs(v - 21.0) < 0.01 or abs(v - 10.5) < 0.01 or abs(v - 27.0) < 0.01 or abs(v - 0.0) < 0.01]
            if posibles:
                no_cero = [x for x in posibles if abs(x[1]) > 0.01]
                elegido = no_cero[0] if no_cero else posibles[0]
                iva = elegido[0]
                vals = [x for x in vals if x[0] != elegido[0]]

        cantidad = vals[0][0] if vals else ""
        precio = vals[-2][0] if len(vals) >= 2 else ""
        importe = vals[-1][0] if vals else ""
        descripcion = re.sub(r"\s+", " ", " ".join(descripcion_parts)).replace(" /", "").replace("/", "").strip()

        if not descripcion and not codigo_proveedor:
            return None

        advertencia = self.validar_calculo(cantidad, precio, importe)
        return {
            "usar": True, "existe_tango": "No", "accion": "Alta artículo", "codigo_tango": "",
            "descripcion": descripcion[:50], "desc_adic": "", "sinonimo": codigo_proveedor[:15],
            "codigo_barras": "", "cantidad": cantidad, "precio": precio, "importe": importe,
            "iva": iva, "advertencia": advertencia, "tipo": "Simple", "escala": "No usa",
            "linea_detectada": linea.text, "score": linea.score, "coincidencia": "", "match_score": 0,
        }

    def validar_calculo(self, cantidad, precio, importe):
        c, p, i = self.parse_decimal(cantidad), self.parse_decimal(precio), self.parse_decimal(importe)
        if c is None or p is None or i is None:
            return ""
        tolerancia = max(1.0, abs(i) * 0.01)
        if abs((c * p) - i) > tolerancia:
            return "Revisar cantidad/precio/importe"
        return "OK"

    def extraer_texto_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()
        if ext == ".pdf":
            try:
                doc = fitz.open(ruta)
                palabras_raw = doc[0].get_text("words") if len(doc) > 0 else []
                self._log(f"extraer_texto_documento: {len(palabras_raw)} palabras raw en pag 1")
                if self._texto_es_legible(palabras_raw):
                    self._log("extraer_texto_documento: -> texto nativo")
                    return "\n".join(page.get_text("text") for page in doc)
                else:
                    self._log("extraer_texto_documento: -> OCR")
                    textos = []
                    for page in doc:
                        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
                        img_path = os.path.join(tempfile.gettempdir(), "_ocr_texto_tmp.png")
                        pix.save(img_path)
                        img = self.preparar_imagen(Image.open(img_path))
                        try:
                            t = pytesseract.image_to_string(img, lang="spa")
                        except Exception:
                            t = pytesseract.image_to_string(img)
                        self._log(f"extraer_texto_documento OCR primeros 200 chars: {repr(t[:200])}")
                        textos.append(t)
                    return "\n".join(textos)
            except Exception as e:
                self._log(f"extraer_texto_documento ERROR: {e}")
                return ""
        try:
            img = self.preparar_imagen(Image.open(ruta))
            try:
                return pytesseract.image_to_string(img, lang="spa")
            except Exception:
                return pytesseract.image_to_string(img)
        except Exception:
            return ""

    def detectar_proveedor(self, texto):
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]
        for l in lineas[:20]:
            may = l.upper()
            if "S.R.L" in may or "SRL" in may or "S.A" in may or " SA" in may:
                return l
        for l in lineas[:10]:
            if len(l) > 4 and not re.search(r"\d{2}/\d{2}/\d{2,4}", l):
                return l
        return ""

    def detectar_cuit_proveedor(self, texto):
        for patron in [
            r"CUIT\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
            r"C\.?U\.?I\.?T\.?\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
        ]:
            for m in re.findall(patron, texto, re.IGNORECASE):
                cuit = self.normalizar_cuit(m)
                if len(cuit) == 11 and cuit[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
                    return cuit
        m = re.search(r"\b((?:20|23|24|27|30|33|34)\d{9})\b", texto)
        return m.group(1) if m else ""

    def detectar_tipo_comprobante(self, texto):
        may = texto.upper()
        if "FACTURA" in may:
            if re.search(r"FACTURA\s*A", may) or re.search(r"\bCOD\.?\s*0?1\b", may): return "FACTURA A"
            if re.search(r"FACTURA\s*B", may) or re.search(r"\bCOD\.?\s*0?6\b", may): return "FACTURA B"
            if re.search(r"FACTURA\s*C", may) or re.search(r"\bCOD\.?\s*0?11\b", may): return "FACTURA C"
            return "FACTURA"
        if "REMITO" in may:
            return "REMITO"
        return ""

    def detectar_numero_comprobante(self, texto):
        patrones = [
            r"Punto\s+de\s+Venta\s*[:\-]?\s*(\d{4,5}).{0,80}?Comp\.?\s*Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"Pto\.?\s*Vta\.?\s*[:\-]?\s*(\d{4,5}).{0,80}?Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"(?:FACTURA|REMITO)\s*(?:N[°º]?)?\s*[:\-]?\s*(\d{4,5})[-\s]?(\d{6,8})",
            r"\b(\d{4,5})[-\s](\d{6,8})\b",
        ]
        for p in patrones:
            m = re.search(p, texto, re.IGNORECASE | re.DOTALL)
            if m:
                pv, nro = m.group(1).zfill(4), m.group(2).zfill(8)
                return pv, nro, f"{pv}-{nro}"
        return "", "", ""

    def detectar_fecha(self, texto):
        m = re.search(r"(?:Fecha\s+de\s+Emisi[oó]n|Fecha)\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})", texto, re.IGNORECASE)
        if m: return m.group(1)
        m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", texto)
        return m.group(1) if m else ""

    def detectar_total(self, texto):
        posibles = re.findall(r"(?:TOTAL|Total)\s*[:$ ]+\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})", texto)
        return posibles[-1] if posibles else ""

    def clave_factura(self, datos):
        cuit = self.normalizar_cuit(datos.get("cuit_proveedor", ""))
        proveedor = str(datos.get("proveedor", "")).strip().upper()
        tipo = str(datos.get("tipo", "")).strip().upper()
        punto = str(datos.get("punto_venta", "")).strip().zfill(4) if datos.get("punto_venta") else ""
        numero = str(datos.get("numero", "")).strip().zfill(8) if datos.get("numero") else ""
        if cuit and punto and numero:
            return f"CUIT:{cuit}|TIPO:{tipo}|PV:{punto}|N:{numero}"
        if proveedor and punto and numero:
            return f"PROV:{re.sub(r'\\s+', ' ', proveedor)}|TIPO:{tipo}|PV:{punto}|N:{numero}"
        return f"ARCHIVO:{datos.get('archivo','')}"

    def leer_qr_afip(self, ruta):
        """
        Busca y decodifica el QR de AFIP en la factura.
        Retorna dict con cuit, tipo, punto_venta, numero, fecha, importe_total o {} si no encuentra.
        """
        try:
            # Obtener imagen de la página
            ext = os.path.splitext(ruta)[1].lower()
            if ext == ".pdf":
                doc = fitz.open(ruta)
                pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
                img_path = os.path.join(tempfile.gettempdir(), "_qr_scan.png")
                pix.save(img_path)
                img = Image.open(img_path).convert("RGB")
            else:
                img = Image.open(ruta).convert("RGB")

            datos = self._decodificar_qr(img)
            if datos:
                self._log(f"QR decodificado: {datos}")
            else:
                self._log("QR: no encontrado o no decodificable")
            return datos
        except Exception as e:
            self._log(f"leer_qr_afip ERROR: {e}")
            return {}

    def _decodificar_qr(self, img):
        """Intenta decodificar QR con pyzbar o cv2, y parsea la URL de AFIP."""
        url = None

        # Intento 1: pyzbar (mejor opcion, no requiere cv2)
        try:
            from pyzbar import pyzbar
            img_arr = np.array(img)
            codigos = pyzbar.decode(img_arr)
            for c in codigos:
                if c.type == "QRCODE":
                    url = c.data.decode("utf-8", errors="ignore")
                    break
            if not url:
                # Intentar con la imagen más grande
                img_big = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)
                codigos = pyzbar.decode(np.array(img_big))
                for c in codigos:
                    if c.type == "QRCODE":
                        url = c.data.decode("utf-8", errors="ignore")
                        break
        except ImportError:
            pass
        except Exception as e:
            self._log(f"_decodificar_qr pyzbar error: {e}")

        # Intento 2: cv2 QRCodeDetector
        if not url and cv2 is not None:
            try:
                img_gray = np.array(img.convert("L"))
                detector = cv2.QRCodeDetector()
                val, _, _ = detector.detectAndDecode(img_gray)
                if val:
                    url = val
            except Exception as e:
                self._log(f"_decodificar_qr cv2 error: {e}")

        if not url:
            return {}

        return self._parsear_url_afip(url)

    def _parsear_url_afip(self, url):
        """
        Parsea la URL del QR de AFIP.
        Formato: https://www.afip.gob.ar/fe/qr/?p=BASE64JSON
        El JSON contiene: ver, fecha, cuit, ptoVta, tipoCmp, nroCmp, importe, moneda, ctz, tipoDocRec, nroDocRec, tipoCodAut, codAut
        """
        try:
            import base64
            import urllib.parse

            self._log(f"URL QR: {url[:120]}")

            # Extraer el parámetro p=
            if "?p=" in url:
                b64 = url.split("?p=")[1].split("&")[0]
            elif url.startswith("{"):
                # A veces el QR contiene el JSON directamente
                datos_json = json.loads(url)
                return self._normalizar_datos_qr(datos_json)
            else:
                return {}

            # Decodificar base64 (puede necesitar padding)
            b64 = b64 + "=" * (-len(b64) % 4)
            datos_json = json.loads(base64.b64decode(b64).decode("utf-8"))
            self._log(f"QR JSON: {datos_json}")
            return self._normalizar_datos_qr(datos_json)
        except Exception as e:
            self._log(f"_parsear_url_afip error: {e} | url={url[:80]}")
            return {}

    def _normalizar_datos_qr(self, d):
        """Convierte el dict del QR de AFIP al formato interno."""
        TIPOS_COMPROBANTE = {
            1: "FACTURA A", 2: "NOTA DEBITO A", 3: "NOTA CREDITO A",
            6: "FACTURA B", 7: "NOTA DEBITO B", 8: "NOTA CREDITO B",
            11: "FACTURA C", 12: "NOTA DEBITO C", 13: "NOTA CREDITO C",
            51: "FACTURA M", 201: "FACTURA DE EXPORTACION",
        }
        try:
            cuit = str(d.get("cuit", "")).replace("-", "").strip()
            pv = str(d.get("ptoVta", "")).zfill(4)
            nro = str(d.get("nroCmp", "")).zfill(8)
            tipo_cod = int(d.get("tipoCmp", 0))
            tipo = TIPOS_COMPROBANTE.get(tipo_cod, f"COMPROBANTE {tipo_cod}")
            fecha = str(d.get("fecha", ""))
            importe = str(d.get("importe", ""))
            return {
                "cuit": cuit,
                "punto_venta": pv,
                "numero": nro,
                "numero_completo": f"{pv}-{nro}",
                "tipo": tipo,
                "fecha": fecha,
                "importe_total": importe,
                "fuente": "QR_AFIP",
            }
        except Exception as e:
            self._log(f"_normalizar_datos_qr error: {e}")
            return {}

    def detectar_datos_factura(self, ruta):
        """
        Detecta datos del encabezado de la factura.
        Primero intenta leer el QR (datos exactos).
        Si no hay QR, usa OCR/texto nativo.
        """
        archivo = os.path.basename(ruta)

        # --- PASO 1: QR AFIP (rápido, exacto, sin OCR) ---
        datos_qr = self.leer_qr_afip(ruta)
        cuit_qr = datos_qr.get("cuit", "")

        # --- PASO 2: Layout conocido para este proveedor ---
        layout = None
        if cuit_qr and hasattr(self, '_base'):
            layout = self._base.obtener_layout(cuit_qr)
            if layout:
                self._log(f"Layout conocido para CUIT {cuit_qr}: procesado {layout['veces_procesado']} veces")

        # --- PASO 3: Texto para datos de encabezado (proveedor, etc.) ---
        texto = self.extraer_texto_documento(ruta)

        proveedor = self.detectar_proveedor(texto)
        cuit = cuit_qr or self.detectar_cuit_proveedor(texto)
        tipo = datos_qr.get("tipo") or self.detectar_tipo_comprobante(texto)
        pv = datos_qr.get("punto_venta") or ""
        nro = datos_qr.get("numero") or ""
        completo = datos_qr.get("numero_completo") or ""
        fecha = datos_qr.get("fecha") or self.detectar_fecha(texto)
        total = datos_qr.get("importe_total") or self.detectar_total(texto)

        if not pv or not nro:
            pv2, nro2, completo2 = self.detectar_numero_comprobante(texto)
            pv = pv or pv2
            nro = nro or nro2
            completo = completo or completo2

        # Si el QR dio el CUIT pero el proveedor no se detectó, buscar en cache
        if cuit and (not proveedor) and hasattr(self, '_base'):
            cache = self._base.obtener_qr_cache(cuit)
            if cache:
                proveedor = cache["proveedor"]
                self._log(f"Proveedor desde QR cache: {proveedor}")

        partes = []
        if proveedor: partes.append(proveedor[:30])
        if completo: partes.append(completo)
        if fecha: partes.append(fecha)
        etiqueta = " | ".join(partes) if partes else archivo

        datos = {
            "archivo": archivo, "ruta": ruta,
            "proveedor": proveedor, "cuit_proveedor": cuit,
            "tipo": tipo, "punto_venta": pv, "numero": nro,
            "numero_completo": completo, "fecha": fecha,
            "total": total, "etiqueta": etiqueta,
            "_layout": layout,
            "_datos_qr": datos_qr,
        }
        datos["clave"] = self.clave_factura(datos)
        return datos

    def analizar(self, ruta):
        """
        Analiza una factura completa y extrae los renglones de artículos.
        Usa el layout aprendido del proveedor si está disponible.
        """
        datos = self.detectar_datos_factura(ruta)
        layout = datos.get("_layout")
        cuit = datos.get("cuit_proveedor", "")
        fuente_qr = bool(datos.get("_datos_qr"))

        paginas = self.leer_pdf(ruta) if ruta.lower().endswith(".pdf") else self.leer_imagen(ruta)
        debug = []
        items = []

        for img, palabras in paginas:
            page_w, page_h = img.size

            if layout and layout["veces_procesado"] >= 2:
                # --- MODO DIRIGIDO: sabemos dónde están los artículos ---
                self._log(f"Usando layout conocido para zona de items")
                palabras_zona = self._filtrar_zona(
                    palabras, page_w, page_h,
                    layout["y0_pct"], layout["y1_pct"],
                    layout["x0_pct"], layout["x1_pct"]
                )
                lineas = self.agrupar_lineas(palabras_zona)
                items_pag, zonas = self.detectar_items(lineas)
            else:
                # --- MODO COMPLETO: escanear toda la página ---
                lineas = self.agrupar_lineas(palabras)
                items_pag, zonas = self.detectar_items(lineas)

                # Aprender el layout si encontramos zonas y tenemos CUIT
                if zonas and cuit and hasattr(self, '_base'):
                    self._aprender_layout(cuit, datos.get("proveedor", ""),
                                         page_w, page_h, zonas,
                                         not layout or not layout["necesita_ocr"])

            # Enriquecer items con datos de encabezado
            for item in items_pag:
                item.update({
                    "proveedor": datos["proveedor"],
                    "cuit_proveedor": cuit,
                    "tipo": datos["tipo"],
                    "punto_venta": datos["punto_venta"],
                    "numero": datos["numero"],
                    "numero_completo": datos["numero_completo"],
                    "fecha": datos["fecha"],
                    "ruta": ruta,
                    "etiqueta": datos["etiqueta"],
                })
                items.append(item)

            # Debug
            img_debug = img.copy().convert("RGB")
            draw = ImageDraw.Draw(img_debug)
            for x0, y0, x1, y1, grupo in zonas:
                draw.rectangle([x0, y0, x1, y1], outline="blue", width=2)
            for item in items_pag:
                pass
            ruta_debug = os.path.join(tempfile.gettempdir(), f"debug_{os.path.basename(ruta)}.png")
            img_debug.save(ruta_debug)
            debug.append({"ruta": ruta_debug, "img": img_debug, "zonas": zonas})

        # Guardar QR cache si obtuvimos datos del QR
        if fuente_qr and cuit and datos.get("proveedor") and hasattr(self, '_base'):
            self._base.guardar_qr_cache(cuit, datos["proveedor"], datos["_datos_qr"])

        return items, debug, datos

    def _filtrar_zona(self, palabras, page_w, page_h, y0_pct, y1_pct, x0_pct, x1_pct):
        """Filtra palabras que caen dentro de la zona aprendida (en coordenadas absolutas)."""
        y0 = page_h * y0_pct
        y1 = page_h * y1_pct
        x0 = page_w * x0_pct
        x1 = page_w * x1_pct
        # Ampliar un poco la zona para no perder bordes
        margen_y = page_h * 0.03
        margen_x = page_w * 0.02
        return [
            p for p in palabras
            if (y0 - margen_y) <= p.y0 <= (y1 + margen_y)
            and (x0 - margen_x) <= p.x0 <= (x1 + margen_x)
        ]

    def _aprender_layout(self, cuit, proveedor, page_w, page_h, zonas, es_pdf_nativo):
        """Guarda la zona de items detectada como % del tamaño de página."""
        if not zonas:
            return
        # Usar la zona con más líneas (la tabla principal)
        zona_principal = max(zonas, key=lambda z: len(z[4]))
        x0, y0, x1, y1, _ = zona_principal
        # Convertir a porcentajes con algo de margen
        y0_pct = max(0.0, (y0 / page_h) - 0.02)
        y1_pct = min(1.0, (y1 / page_h) + 0.04)
        x0_pct = max(0.0, (x0 / page_w) - 0.01)
        x1_pct = min(1.0, (x1 / page_w) + 0.01)
        self._log(f"Aprendiendo layout CUIT {cuit}: y={y0_pct:.2f}-{y1_pct:.2f} x={x0_pct:.2f}-{x1_pct:.2f}")
        self._base.guardar_layout(
            cuit, proveedor, page_w, page_h,
            y0_pct, y1_pct, x0_pct, x1_pct,
            int(es_pdf_nativo), int(not es_pdf_nativo)
        )
        datos["clave"] = self.clave_factura(datos)
        return datos

class App:
    def __init__(self, root):
        self.root = root
        self.root.title(f"Facturas multiples → Tango {APP_VERSION}")

        # Adaptar tamaño al monitor actual
        sw = root.winfo_screenwidth()
        sh = root.winfo_screenheight()
        ancho = min(1800, int(sw * 0.97))
        alto  = min(960,  int(sh * 0.93))
        self.root.geometry(f"{ancho}x{alto}+0+0")
        self.root.minsize(900, 600)
        self._pantalla_angosta = sw < 1500   # 1366x768 y similares

        self.base = BaseLocal()
        self.reader = FacturaTableReader()
        self.reader._base = self.base
        self.catalogo = CatalogoTango()

        self.ruta_modelo_tango = None
        self.ruta_catalogo_tango = None
        self.resultados_debug = []
        self.facturas_cargadas = []
        self.claves_facturas = set()
        self.filtro_actual = tk.StringVar(value="Ver todos")
        self._todos_items = []
        self.umbral_confianza = tk.IntVar(value=UMBRAL_CONFIANZA_DEFAULT)

        # Para el hilo de procesamiento
        self._queue = queue.Queue()
        self._procesando = False

        # Vista previa
        self._preview_img = None
        self._preview_ruta_actual = None

        self.crear_ui()
        self.root.after(200, self._procesar_queue)

    def configurar_estilo(self):
        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        pad = 3 if self._pantalla_angosta else 5
        style.configure("TButton", padding=pad)
        fs = 8 if self._pantalla_angosta else 9
        style.configure("Treeview", rowheight=22 if self._pantalla_angosta else 24,
                        font=("Segoe UI", fs))
        style.configure("Treeview.Heading", font=("Segoe UI", fs, "bold"))
        style.configure("TLabel", font=("Segoe UI", fs))
        style.configure("Accent.TButton", foreground="white", background="#0b5ed7")

    def crear_ui(self):
        self.configurar_estilo()
        main = ttk.Frame(self.root, padding=5 if self._pantalla_angosta else 8)
        main.pack(fill="both", expand=True)

        # ---- Barra de botones: 1 fila en pantalla ancha, 2 filas en angosta ----
        def agregar_botones_fila1(parent):
            ttk.Button(parent, text="1. Articulos Tango", command=self.abrir_catalogo).pack(side="left", padx=2)
            ttk.Button(parent, text="2. Modelo Tango", command=self.abrir_modelo).pack(side="left", padx=2)
            ttk.Button(parent, text="3. Agregar facturas", command=self.agregar_facturas).pack(side="left", padx=2)
            ttk.Button(parent, text="Foto", command=self.sacar_foto).pack(side="left", padx=2)
            ttk.Separator(parent, orient="vertical").pack(side="left", fill="y", padx=4)
            ttk.Button(parent, text="Debug", command=self.ver_debug).pack(side="left", padx=2)
            ttk.Button(parent, text="Abrir factura", command=self.abrir_factura_original).pack(side="left", padx=2)
            ttk.Separator(parent, orient="vertical").pack(side="left", fill="y", padx=4)
            ttk.Button(parent, text="Guardar sesion", command=self.guardar_sesion).pack(side="left", padx=2)
            ttk.Button(parent, text="Cargar sesion", command=self.cargar_sesion).pack(side="left", padx=2)

        def agregar_botones_fila2(parent):
            ttk.Button(parent, text="Equivalencia", command=self.aprender_equivalencia_seleccionada).pack(side="left", padx=2)
            ttk.Button(parent, text="Candidatos", command=self.ver_candidatos_alternativos).pack(side="left", padx=2)
            ttk.Button(parent, text="Layouts", command=self.ver_layouts_aprendidos).pack(side="left", padx=2)
            ttk.Separator(parent, orient="vertical").pack(side="left", fill="y", padx=4)
            ttk.Label(parent, text="Umbral:").pack(side="left")
            self._spin_umbral = ttk.Spinbox(parent, from_=50, to=99, textvariable=self.umbral_confianza,
                                            width=4, command=self._on_umbral_change)
            self._spin_umbral.pack(side="left", padx=2)
            ttk.Label(parent, text="%").pack(side="left")

        if self._pantalla_angosta:
            # Dos filas de botones más compactas
            fila1 = ttk.Frame(main)
            fila1.pack(fill="x", pady=(0, 2))
            agregar_botones_fila1(fila1)
            fila2 = ttk.Frame(main)
            fila2.pack(fill="x", pady=(0, 3))
            agregar_botones_fila2(fila2)
        else:
            top = ttk.Frame(main)
            top.pack(fill="x", pady=(0, 4))
            agregar_botones_fila1(top)
            ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=4)
            agregar_botones_fila2(top)

        # ---- Filtro y resumen ----
        filtro_frame = ttk.Frame(main)
        filtro_frame.pack(fill="x", pady=(0, 3))
        ttk.Label(filtro_frame, text="Filtro:").pack(side="left")
        combo = ttk.Combobox(
            filtro_frame, textvariable=self.filtro_actual, state="readonly", width=20,
            values=["Ver todos", "Solo seleccionados", "Solo nuevos", "Solo existentes",
                    "Sin codigo Tango", "Baja confianza", "Destildados", "Con advertencias"]
        )
        combo.pack(side="left", padx=4)
        combo.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtro())
        self.lbl_resumen = ttk.Label(filtro_frame, text="", foreground="#555")
        self.lbl_resumen.pack(side="left", padx=10)

        self.lbl_catalogo = ttk.Label(main, text="Articulos Tango: no seleccionado")
        self.lbl_catalogo.pack(fill="x")
        self.lbl_modelo = ttk.Label(main, text="Modelo Tango: no seleccionado")
        self.lbl_modelo.pack(fill="x")
        self.lbl_facturas = ttk.Label(main, text="Facturas cargadas: 0")
        self.lbl_facturas.pack(fill="x", pady=(0, 2))

        # ---- Barra de progreso ----
        self._frame_progreso = ttk.Frame(main)
        self._frame_progreso.pack(fill="x", pady=(0, 2))
        self._lbl_progreso = ttk.Label(self._frame_progreso, text="")
        self._lbl_progreso.pack(side="left")
        self._barra = ttk.Progressbar(self._frame_progreso, mode="determinate", length=350)
        self._barra.pack(side="left", padx=6)
        self._frame_progreso.pack_forget()

        # ---- Panel central: tabla + preview ----
        panel_central = ttk.Frame(main)
        panel_central.pack(fill="both", expand=True)

        self.columns = (
            "usar", "factura", "proveedor", "cuit", "tipo", "numero", "fecha",
            "existe", "accion", "codigo_tango", "descripcion", "sinonimo",
            "codigo_barras", "cantidad", "precio", "importe", "iva", "advertencia",
            "match", "match_score", "tipo_art", "escala", "pagina", "ruta", "linea"
        )
        headers = {
            "usar": "✓", "factura": "Factura", "proveedor": "Proveedor", "cuit": "CUIT",
            "tipo": "Tipo", "numero": "N°", "fecha": "Fecha", "existe": "Existe",
            "accion": "Accion", "codigo_tango": "Cod.Tango", "descripcion": "Descripcion",
            "sinonimo": "Cod.prov.", "codigo_barras": "Cod.barras", "cantidad": "Cant.",
            "precio": "Precio", "importe": "Importe", "iva": "IVA%", "advertencia": "Advertencia",
            "match": "Coincidencia", "match_score": "Conf.", "tipo_art": "Tipo",
            "escala": "Escala", "pagina": "Pag.", "ruta": "Ruta", "linea": "Linea detectada",
        }
        # Anchos reducidos en pantalla angosta
        f = 0.82 if self._pantalla_angosta else 1.0
        widths = {
            "usar": 38, "factura": int(150*f), "proveedor": int(180*f), "cuit": int(100*f),
            "tipo": int(85*f), "numero": int(100*f), "fecha": int(80*f), "existe": int(55*f),
            "accion": int(130*f), "codigo_tango": int(115*f), "descripcion": int(280*f),
            "sinonimo": int(110*f), "codigo_barras": int(110*f), "cantidad": int(75*f),
            "precio": int(80*f), "importe": int(85*f), "iva": int(60*f),
            "advertencia": int(170*f), "match": int(190*f), "match_score": 50,
            "tipo_art": int(70*f), "escala": int(70*f), "pagina": 45,
            "ruta": int(220*f), "linea": int(350*f),
        }

        frame_tabla = ttk.Frame(panel_central)
        frame_tabla.pack(side="left", fill="both", expand=True)
        scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        # Altura de tabla dinámica: se ajusta al espacio disponible via expand=True
        self.tabla = ttk.Treeview(
            frame_tabla, columns=self.columns, show="headings",
            yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set
        )
        scroll_y.config(command=self.tabla.yview)
        scroll_x.config(command=self.tabla.xview)

        for c in self.columns:
            self.tabla.heading(c, text=headers[c],
                               command=lambda col=c: self._ordenar_columna(col))
            self.tabla.column(c, width=widths[c], stretch=False)

        self.tabla.tag_configure("verde", background="#d9f5d6")
        self.tabla.tag_configure("amarillo", background="#fff3bf")
        self.tabla.tag_configure("rojo", background="#ffd6d6")
        self.tabla.tag_configure("gris", background="#e9ecef")
        self.tabla.tag_configure("azul", background="#d6eaff")
        self.tabla.tag_configure("advertencia", background="#ffd8a8")
        self.tabla.pack(side="left", fill="both", expand=True)
        self.tabla.bind("<Double-1>", self.doble_click)
        self.tabla.bind("<<TreeviewSelect>>", self._on_seleccion)

        # Panel de preview: oculto en pantalla angosta (se puede abrir con "Abrir factura")
        if not self._pantalla_angosta:
            self._frame_preview = ttk.LabelFrame(panel_central, text="Vista previa", width=260, padding=4)
            self._frame_preview.pack(side="right", fill="y", padx=(5, 0))
            self._frame_preview.pack_propagate(False)
            self._lbl_preview = tk.Label(self._frame_preview, text="Selecciona una fila\npara ver la factura",
                                         bg="#f0f0f0", fg="#888", justify="center")
            self._lbl_preview.pack(fill="both", expand=True)
            self._lbl_preview_info = ttk.Label(self._frame_preview, text="", wraplength=240, justify="left")
            self._lbl_preview_info.pack(fill="x")
        else:
            self._frame_preview = None
            self._lbl_preview = None
            self._lbl_preview_info = None

        # ---- Fila 1: edicion ----
        bottom = ttk.Frame(main)
        bottom.pack(fill="x", pady=(4, 2))
        for txt, cmd in [
            ("Selec. todo", self.seleccionar_todo),
            ("Borrar sel.", self.borrar_seleccion),
            ("Invertir", self.invertir_seleccion),
            ("Agregar manual", self.agregar_manual),
            ("Eliminar fila", self.eliminar_fila),
            ("Limpiar todo", self.limpiar_todo),
            ("Recomparar", self.recomparar),
        ]:
            ttk.Button(bottom, text=txt, command=cmd).pack(side="left", padx=2)

        # ---- Fila 2: exportacion ----
        export_frame = ttk.LabelFrame(main, text="Exportar", padding=(4, 3))
        export_frame.pack(fill="x", pady=(2, 0))
        ttk.Button(export_frame, text="📋 Reporte revision",
                   command=self.generar_reporte_revision).pack(side="left", padx=5)
        ttk.Separator(export_frame, orient="vertical").pack(side="left", fill="y", padx=3)
        ttk.Button(export_frame, text="⬆ Alta articulos nuevos (Tango)",
                   command=self.generar_excel_alta_articulos).pack(side="left", padx=5)
        ttk.Button(export_frame, text="📦 Resumen compra / stock",
                   command=self.generar_resumen_compra_stock).pack(side="left", padx=5)

        self.lbl_estado = ttk.Label(
            main,
            text=f"{APP_VERSION}: QR AFIP | Layout por proveedor | OCR adaptativo | Preview | Candidatos",
            foreground="#0b5ed7"
        )
        self.lbl_estado.pack(fill="x", pady=(3, 0))

    # ------------------------------------------------------------------ #
    # Vista previa lateral                                                 #
    # ------------------------------------------------------------------ #

    def _on_seleccion(self, event):
        """Al seleccionar una fila, actualizar la vista previa si existe."""
        sel = self.tabla.selection()
        if not sel:
            return
        v = self.tabla.item(sel[0], "values")
        if not v:
            return
        if self._lbl_preview_info:
            info = f"Proveedor: {v[2]}\nFactura: {v[5]}\nFecha: {v[6]}\nCodigo: {v[9]}\nDescripcion: {v[10]}\nConfianza: {v[19]}%"
            self._lbl_preview_info.config(text=info)
        if self._lbl_preview is None:
            return
        ruta = v[23] if len(v) > 23 else ""
        pagina_str = v[22] if len(v) > 22 else "1"
        if ruta and os.path.exists(ruta) and ruta != self._preview_ruta_actual:
            self._preview_ruta_actual = ruta
            try:
                try:
                    pagina = max(0, int(pagina_str) - 1)
                except Exception:
                    pagina = 0
                img = self._cargar_miniatura_factura(ruta, pagina)
                if img:
                    self._preview_img = img  # evitar GC
                    self._lbl_preview.config(image=img, text="")
                else:
                    self._lbl_preview.config(image="", text="No se pudo\ncargar la imagen")
            except Exception:
                self._lbl_preview.config(image="", text="Error al cargar\nla imagen")

    def _cargar_miniatura_factura(self, ruta, pagina=0):
        """Carga la primera pagina de la factura como miniatura para el panel lateral."""
        try:
            from PIL import ImageTk
            ext = os.path.splitext(ruta)[1].lower()
            if ext == ".pdf":
                doc = fitz.open(ruta)
                if pagina >= len(doc):
                    pagina = 0
                page = doc[pagina]
                pix = page.get_pixmap(matrix=fitz.Matrix(0.4, 0.4))
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            else:
                img = Image.open(ruta)
                img = img.convert("RGB")

            # Ajustar al panel (260px ancho aprox)
            max_w, max_h = 260, 500
            img.thumbnail((max_w, max_h), Image.LANCZOS)
            return ImageTk.PhotoImage(img)
        except Exception:
            return None

    # ------------------------------------------------------------------ #
    # Candidatos alternativos                                              #
    # ------------------------------------------------------------------ #

    def ver_candidatos_alternativos(self):
        """Muestra los top candidatos de Tango para la fila seleccionada."""
        if not self.catalogo.articulos:
            messagebox.showinfo("Sin catalogo", "Primero carga el catalogo de articulos Tango.")
            return
        sel = self.tabla.selection()
        if not sel:
            messagebox.showerror("Error", "Selecciona una fila primero.")
            return
        v = list(self.tabla.item(sel[0], "values"))
        codigo_proveedor = v[11]
        codigo_barras = v[12]
        descripcion = v[10]

        candidatos = self.catalogo.buscar_candidatos(codigo_proveedor, codigo_barras, descripcion, top=5)
        if not candidatos:
            messagebox.showinfo("Sin candidatos", "No se encontraron candidatos en el catalogo para este articulo.")
            return

        # Ventana con lista de candidatos
        win = tk.Toplevel(self.root)
        win.title("Candidatos alternativos")
        win.geometry("800x340")
        win.grab_set()

        ttk.Label(win, text=f"Articulo de factura: {descripcion}  |  Cod.proveedor: {codigo_proveedor}",
                  font=("Segoe UI", 9, "bold")).pack(pady=(10, 4), padx=10, anchor="w")
        ttk.Label(win, text="Hacé doble clic para asignar el articulo a la fila seleccionada:").pack(padx=10, anchor="w")

        cols = ("conf", "codigo", "descripcion", "sinonimo", "barras", "metodo")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=6)
        tree.heading("conf", text="Conf.")
        tree.heading("codigo", text="Codigo Tango")
        tree.heading("descripcion", text="Descripcion")
        tree.heading("sinonimo", text="Sinonimo")
        tree.heading("barras", text="Cod. barras")
        tree.heading("metodo", text="Metodo")
        tree.column("conf", width=55); tree.column("codigo", width=130); tree.column("descripcion", width=280)
        tree.column("sinonimo", width=110); tree.column("barras", width=100); tree.column("metodo", width=180)
        tree.pack(fill="both", expand=True, padx=10, pady=6)

        for art, metodo, score in candidatos:
            tree.insert("", "end", values=(
                f"{score}%", art["codigo"], art["descripcion"],
                art["sinonimo"], art["codigo_barras"], metodo
            ))

        item_id_original = sel[0]

        def aplicar(event=None):
            sel2 = tree.selection()
            if not sel2:
                return
            vals_cand = tree.item(sel2[0], "values")
            # vals_cand: conf, codigo, descripcion, sinonimo, barras, metodo
            v[7] = "Si"
            v[8] = "Cargar stock / compra"
            v[9] = vals_cand[1]   # codigo_tango
            v[10] = vals_cand[2]  # descripcion
            v[12] = vals_cand[4]  # codigo_barras
            v[18] = vals_cand[5]  # metodo
            v[19] = vals_cand[0].replace("%", "")  # score
            self.tabla.item(item_id_original, values=v)
            self.actualizar_tags()
            self.aplicar_filtro()
            win.destroy()

        tree.bind("<Double-1>", aplicar)
        ttk.Button(win, text="Asignar seleccionado", command=aplicar).pack(pady=6)

    # ------------------------------------------------------------------ #
    # Ordenamiento de columnas                                             #
    # ------------------------------------------------------------------ #

    def _ordenar_columna(self, col):
        """Ordena la tabla por la columna clickeada. Segundo click invierte."""
        col_idx = self.columns.index(col)
        self.refrescar_indice_items()
        datos = []
        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
                datos.append((item_id, v))
            except Exception:
                pass

        # Detectar si ya estaba ordenado por esta columna (invertir)
        orden_actual = getattr(self, "_orden_columna", (None, False))
        invertir = (orden_actual[0] == col) and (not orden_actual[1])
        self._orden_columna = (col, invertir)

        def clave(x):
            val = x[1][col_idx] if x[1] else ""
            try:
                return (0, float(str(val).replace(",", ".")))
            except Exception:
                return (1, str(val).lower())

        datos.sort(key=clave, reverse=invertir)
        self._todos_items = [item_id for item_id, _ in datos]
        for item_id in self._todos_items:
            try:
                self.tabla.move(item_id, "", "end")
            except Exception:
                pass
        self.aplicar_filtro()

    # ------------------------------------------------------------------ #
    # Resumen de estado                                                    #
    # ------------------------------------------------------------------ #

    def actualizar_resumen(self):
        """Actualiza el label de resumen con conteos rapidos."""
        self.refrescar_indice_items()
        total = len(self._todos_items)
        sel, nuevos, existentes, adv = 0, 0, 0, 0
        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
                if not v:
                    continue
                if v[0] == CHECK_ON:
                    sel += 1
                if v[7] == "Si":
                    existentes += 1
                elif v[7] == "No":
                    nuevos += 1
                if v[17] and v[17] not in ("", "OK"):
                    adv += 1
            except Exception:
                pass
        self.lbl_resumen.config(
            text=f"Total: {total}  |  Seleccionados: {sel}  |  Existentes: {existentes}  |  Nuevos: {nuevos}  |  Advertencias: {adv}"
        )

    def _on_umbral_change(self):
        """Recomparar automaticamente cuando cambia el umbral."""
        if self.catalogo.articulos and self._todos_items:
            self.recomparar(silencioso=True)

    def ver_layouts_aprendidos(self):
        """Muestra los layouts y datos QR aprendidos por proveedor."""
        con = self.base.conectar()
        cur = con.cursor()
        cur.execute("""
            SELECT cuit, proveedor, veces_procesado, necesita_ocr,
                   zona_items_y0_pct, zona_items_y1_pct, ultima_fecha
            FROM proveedor_layout ORDER BY veces_procesado DESC
        """)
        rows = cur.fetchall()
        con.close()

        win = tk.Toplevel(self.root)
        win.title("Proveedores con layout aprendido")
        win.geometry("750x380")

        cols = ("cuit", "proveedor", "veces", "modo", "zona_y", "ultima")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=12)
        tree.heading("cuit", text="CUIT")
        tree.heading("proveedor", text="Proveedor")
        tree.heading("veces", text="Veces")
        tree.heading("modo", text="Modo")
        tree.heading("zona_y", text="Zona items (% pagina)")
        tree.heading("ultima", text="Ultima vez")
        tree.column("cuit", width=120); tree.column("proveedor", width=200)
        tree.column("veces", width=55); tree.column("modo", width=80)
        tree.column("zona_y", width=160); tree.column("ultima", width=120)
        tree.pack(fill="both", expand=True, padx=10, pady=8)

        for r in rows:
            cuit, prov, veces, necesita_ocr, y0, y1, fecha = r
            modo = "OCR" if necesita_ocr else "Nativo"
            zona = f"{int(y0*100)}% - {int(y1*100)}%"
            tree.insert("", "end", values=(cuit, prov, veces, modo, zona, fecha or ""))

        if not rows:
            ttk.Label(win, text="Todavia no hay layouts aprendidos.\nProcesa al menos una factura de cada proveedor.",
                      justify="center").pack(pady=40)

        ttk.Button(win, text="Cerrar", command=win.destroy).pack(pady=6)

    def refrescar_indice_items(self):
        visibles = list(self.tabla.get_children(""))
        existentes = []
        for item_id in self._todos_items:
            try:
                self.tabla.item(item_id)
                existentes.append(item_id)
            except Exception:
                pass
        for item_id in visibles:
            if item_id not in existentes:
                existentes.append(item_id)
        self._todos_items = existentes

    def abrir_catalogo(self):
        ruta = filedialog.askopenfilename(title="Seleccionar Artículos.xlsx exportado desde Tango", filetypes=[("Excel", "*.xlsx")])
        if not ruta: return
        try:
            self.catalogo.cargar(ruta)
            self.ruta_catalogo_tango = ruta
            self.lbl_catalogo.config(text=f"Artículos Tango: {ruta} ({len(self.catalogo.articulos)} artículos)")
            self.recomparar(silencioso=True)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def abrir_modelo(self):
        ruta = filedialog.askopenfilename(title="Seleccionar Excel modelo exportado desde Tango", filetypes=[("Excel", "*.xlsx")])
        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def factura_duplicada_en_sesion(self, datos):
        return datos.get("clave", "") in self.claves_facturas

    def confirmar_factura_duplicada(self, datos, fuente):
        msg = (
            f"Esta factura ya fue cargada ({fuente}).\n\n"
            f"Proveedor: {datos.get('proveedor','')}\nCUIT: {datos.get('cuit_proveedor','')}\n"
            f"Tipo: {datos.get('tipo','')}\nNúmero: {datos.get('numero_completo','')}\n"
            f"Fecha: {datos.get('fecha','')}\nArchivo: {datos.get('archivo','')}\n\n"
            "¿Querés cargarla igualmente?"
        )
        return messagebox.askyesno("Factura duplicada", msg)

    # ------------------------------------------------------------------ #
    # Procesamiento en hilo + queue                                        #
    # ------------------------------------------------------------------ #

    def _procesar_queue(self):
        """Ejecutado en el hilo principal cada 200ms para leer resultados del hilo worker."""
        try:
            while True:
                msg = self._queue.get_nowait()
                tipo = msg.get("tipo")

                if tipo == "progreso":
                    self._lbl_progreso.config(text=msg["texto"])
                    self._barra["value"] = msg["valor"]
                    self._barra["maximum"] = msg["maximo"]

                elif tipo == "item":
                    self.insertar_item(msg["item"])

                elif tipo == "debug":
                    self.resultados_debug.extend(msg["debug"])

                elif tipo == "factura":
                    self.facturas_cargadas.append(msg["datos"])
                    self.claves_facturas.add(msg["datos"]["clave"])
                    self.actualizar_label_facturas()

                elif tipo == "fin":
                    self._frame_progreso.pack_forget()
                    self._procesando = False
                    if self.catalogo.articulos:
                        self.recomparar(silencioso=True)
                    else:
                        self.actualizar_tags()
                        self.aplicar_filtro()
                    self.actualizar_label_facturas()
                    self.actualizar_resumen()
                    messagebox.showinfo(
                        "Lectura terminada",
                        f"Facturas agregadas: {msg['cargadas']}\n"
                        f"Facturas salteadas (duplicadas): {msg['salteadas']}\n"
                        f"Lineas candidatas agregadas: {msg['total_items']}"
                    )

                elif tipo == "error":
                    messagebox.showerror("Error", msg["texto"])

        except queue.Empty:
            pass
        self.root.after(200, self._procesar_queue)

    def _worker_agregar_facturas(self, rutas):
        total_items, cargadas, salteadas = 0, 0, 0
        total = len(rutas)
        for idx, ruta in enumerate(rutas, 1):
            self._queue.put({"tipo": "progreso", "texto": f"Procesando {idx}/{total}: {os.path.basename(ruta)}", "valor": idx - 1, "maximo": total})
            try:
                datos_previos = self.reader.detectar_datos_factura(ruta)
                clave = datos_previos.get("clave", "")

                # Duplicado en sesion: skip silencioso en hilo (se decidio antes de lanzar el hilo)
                if clave in self.claves_facturas:
                    salteadas += 1
                    continue

                historial = self.base.factura_existe(clave)
                if historial:
                    # No podemos mostrar dialogo desde el hilo — marcar como advertencia
                    # El usuario ya fue consultado antes de lanzar el hilo
                    pass

                items, debug, datos = self.reader.analizar(ruta)
                self._queue.put({"tipo": "debug", "debug": debug})
                self._queue.put({"tipo": "factura", "datos": datos})
                for item in items:
                    self._queue.put({"tipo": "item", "item": item})
                total_items += len(items)
                cargadas += 1
            except Exception as e:
                self._queue.put({"tipo": "error", "texto": f"No se pudo procesar:\n{ruta}\n\n{e}"})

        self._queue.put({"tipo": "progreso", "texto": "Listo.", "valor": total, "maximo": total})
        self._queue.put({"tipo": "fin", "cargadas": cargadas, "salteadas": salteadas, "total_items": total_items})

    def agregar_facturas(self):
        if self._procesando:
            messagebox.showwarning("Procesando", "Ya hay facturas siendo procesadas. Espera que terminen.")
            return
        rutas = filedialog.askopenfilenames(
            title="Seleccionar una o varias facturas",
            filetypes=[("Facturas", "*.pdf *.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")]
        )
        if not rutas:
            return

        # Verificar duplicados ANTES de lanzar el hilo (necesitamos dialogo)
        rutas_a_procesar = []
        salteadas_previo = 0
        for ruta in rutas:
            try:
                datos_previos = self.reader.detectar_datos_factura(ruta)
            except Exception:
                rutas_a_procesar.append(ruta)
                continue
            if self.factura_duplicada_en_sesion(datos_previos):
                if not self.confirmar_factura_duplicada(datos_previos, "en esta sesion"):
                    salteadas_previo += 1
                    continue
            historial = self.base.factura_existe(datos_previos.get("clave", ""))
            if historial:
                if not self.confirmar_factura_duplicada(datos_previos, "historial permanente"):
                    salteadas_previo += 1
                    continue
            rutas_a_procesar.append(ruta)

        if not rutas_a_procesar:
            return

        # Mostrar barra de progreso y lanzar hilo
        self._procesando = True
        self._barra["value"] = 0
        self._barra["maximum"] = len(rutas_a_procesar)
        self._lbl_progreso.config(text="Iniciando...")
        self._frame_progreso.pack(fill="x", pady=(0, 4))

        hilo = threading.Thread(target=self._worker_agregar_facturas, args=(rutas_a_procesar,), daemon=True)
        hilo.start()

    def sacar_foto(self):
        if cv2 is None:
            messagebox.showerror("Error", "Falta instalar opencv-python")
            return
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "No se pudo abrir la cámara.")
            return
        messagebox.showinfo("Cámara", "ESPACIO captura / ESC cancela")
        ruta_foto = None
        while True:
            ret, frame = cap.read()
            if not ret: break
            cv2.imshow("Capturar factura", frame)
            key = cv2.waitKey(1)
            if key == 27: break
            if key == 32:
                ruta_foto = os.path.join(tempfile.gettempdir(), f"factura_capturada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
                cv2.imwrite(ruta_foto, frame)
                break
        cap.release()
        cv2.destroyAllWindows()
        if ruta_foto:
            self.procesar_ruta_unica(ruta_foto)

    def procesar_ruta_unica(self, ruta):
        try:
            datos_previos = self.reader.detectar_datos_factura(ruta)
            if self.factura_duplicada_en_sesion(datos_previos):
                if not self.confirmar_factura_duplicada(datos_previos, "en esta sesion"): return
            historial = self.base.factura_existe(datos_previos.get("clave", ""))
            if historial:
                if not self.confirmar_factura_duplicada(datos_previos, "historial permanente"): return
        except Exception as e:
            messagebox.showerror("Error", str(e))
            return

        self._procesando = True
        self._barra["value"] = 0
        self._barra["maximum"] = 1
        self._lbl_progreso.config(text=f"Procesando foto...")
        self._frame_progreso.pack(fill="x", pady=(0, 4))
        hilo = threading.Thread(target=self._worker_agregar_facturas, args=([ruta],), daemon=True)
        hilo.start()

    def insertar_item(self, item):
        valores = (
            CHECK_ON if item["usar"] else CHECK_OFF,
            item.get("factura_etiqueta", ""), item.get("proveedor_factura", ""), item.get("cuit_proveedor", ""),
            item.get("tipo_factura", ""), item.get("numero_completo", ""), item.get("fecha_factura", ""),
            item["existe_tango"], item["accion"], item["codigo_tango"], item["descripcion"], item["sinonimo"],
            item["codigo_barras"], item["cantidad"], item["precio"], item["importe"], item.get("iva", ""),
            item.get("advertencia", ""), item.get("coincidencia", ""), item.get("match_score", ""),
            item["tipo"], item["escala"], item.get("pagina", ""), item.get("ruta_factura", ""), item["linea_detectada"],
        )
        tag = self.tag_para_valores(valores)
        item_id = self.tabla.insert("", "end", values=valores, tags=(tag,))
        if item_id not in self._todos_items:
            self._todos_items.append(item_id)

    def tag_para_valores(self, v):
        usar, existe, codigo, advertencia, match, conf = v[0], v[7], v[9], v[17], v[18], v[19]
        if usar == CHECK_OFF: return "gris"
        if advertencia and advertencia != "OK": return "advertencia"
        if match == "Equivalencia aprendida": return "azul"
        try: conf = int(conf)
        except Exception: conf = 0
        umbral = self.umbral_confianza.get()
        if existe in ("Si", "Sí") and conf >= umbral: return "verde"
        if existe in ("Si", "Sí") and conf < umbral: return "amarillo"
        if existe == "No" or not codigo: return "rojo"
        return ""

    def actualizar_tags(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
                if v:
                    self.tabla.item(item_id, tags=(self.tag_para_valores(v),))
            except Exception:
                pass

    def recomparar(self, silencioso=False):
        if not self.catalogo.articulos:
            if not silencioso:
                messagebox.showinfo("Catalogo no cargado", "Primero carga el Excel Articulos.xlsx exportado desde Tango.")
            return
        umbral = self.umbral_confianza.get()
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try: v = list(self.tabla.item(item_id, "values"))
            except Exception: continue
            if not v: continue
            cuit, codigo_proveedor, codigo_barras, descripcion = v[3], v[11], v[12], v[10]
            eq = self.base.buscar_equivalencia(cuit, codigo_proveedor)
            if eq:
                v[7], v[8], v[9], v[10], v[12], v[18], v[19] = "Si", "Cargar stock / compra", eq["codigo_tango"], eq["descripcion_tango"] or descripcion, eq["codigo_barras"], "Equivalencia aprendida", "100"
            else:
                art, metodo, score = self.catalogo.buscar(codigo_proveedor, codigo_barras, descripcion, umbral)
                if art:
                    v[7], v[8], v[9], v[10], v[12], v[18], v[19] = "Si", "Cargar stock / compra", art["codigo"], art["descripcion"], art["codigo_barras"], metodo, str(score)
                else:
                    v[7], v[8], v[18], v[19] = "No", "Alta articulo", "", ""
            self.tabla.item(item_id, values=v)
        self.actualizar_tags()
        self.aplicar_filtro()
        self.actualizar_resumen()
        if not silencioso:
            messagebox.showinfo("Listo", "Comparacion terminada.")

    def doble_click(self, event):
        item_id = self.tabla.focus()
        if not item_id: return
        col = self.tabla.identify_column(event.x)
        idx = int(col.replace("#", "")) - 1
        valores = list(self.tabla.item(item_id, "values"))
        if idx == 0:
            valores[0] = CHECK_OFF if valores[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=valores)
            self.actualizar_tags()
            self.aplicar_filtro()
            return
        nombres = [
            "Usar", "Factura", "Proveedor", "CUIT", "Tipo", "Número", "Fecha", "Existe", "Acción",
            "Código Tango", "Descripción", "Cod. proveedor", "Cód. barras", "Cantidad", "Precio",
            "Importe", "IVA %", "Advertencia", "Coincidencia", "Confianza", "Tipo art.", "Escala",
            "Página", "Ruta factura", "Línea detectada"
        ]
        nuevo = simpledialog.askstring("Editar", nombres[idx], initialvalue=valores[idx])
        if nuevo is not None:
            valores[idx] = nuevo
            self.tabla.item(item_id, values=valores)
            self.actualizar_tags()
            self.aplicar_filtro()

    def seleccionar_todo(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = list(self.tabla.item(item_id, "values"))
                if v:
                    v[0] = CHECK_ON
                    self.tabla.item(item_id, values=v)
            except Exception:
                pass
        self.actualizar_tags(); self.aplicar_filtro()

    def borrar_seleccion(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = list(self.tabla.item(item_id, "values"))
                if v:
                    v[0] = CHECK_OFF
                    self.tabla.item(item_id, values=v)
            except Exception:
                pass
        self.actualizar_tags(); self.aplicar_filtro()

    def invertir_seleccion(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = list(self.tabla.item(item_id, "values"))
                if v:
                    v[0] = CHECK_OFF if v[0] == CHECK_ON else CHECK_ON
                    self.tabla.item(item_id, values=v)
            except Exception:
                pass
        self.actualizar_tags(); self.aplicar_filtro()

    def agregar_manual(self):
        self.insertar_item({
            "usar": True, "factura_etiqueta": "Manual", "proveedor_factura": "", "cuit_proveedor": "",
            "tipo_factura": "", "numero_completo": "", "fecha_factura": "", "existe_tango": "No",
            "accion": "Alta artículo", "codigo_tango": "", "descripcion": "", "desc_adic": "",
            "sinonimo": "", "codigo_barras": "", "cantidad": "", "precio": "", "importe": "", "iva": "",
            "advertencia": "", "tipo": "Simple", "escala": "No usa", "pagina": "", "ruta_factura": "",
            "linea_detectada": "Manual", "coincidencia": "", "match_score": "",
        })
        self.aplicar_filtro()

    def eliminar_fila(self):
        for item_id in self.tabla.selection():
            try: self.tabla.delete(item_id)
            except Exception: pass
            if item_id in self._todos_items:
                self._todos_items.remove(item_id)

    def limpiar_todo(self):
        if not messagebox.askyesno("Confirmar", "¿Querés borrar todas las facturas y filas cargadas?"):
            return
        self.limpiar_todo_sin_preguntar()

    def limpiar_todo_sin_preguntar(self):
        self.refrescar_indice_items()
        for item_id in list(self._todos_items):
            try: self.tabla.delete(item_id)
            except Exception: pass
        self.facturas_cargadas, self.resultados_debug, self.claves_facturas, self._todos_items = [], [], set(), []
        self.actualizar_label_facturas()

    def aplicar_filtro(self):
        filtro = self.filtro_actual.get()
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
                if not v: continue
                if self.pasa_filtro(v, filtro):
                    self.tabla.move(item_id, "", "end")
                else:
                    self.tabla.detach(item_id)
            except Exception:
                pass
        self.actualizar_resumen()

    def pasa_filtro(self, v, filtro):
        usar, existe, codigo, advertencia, conf = v[0], v[7], v[9], v[17], v[19]
        try: conf_int = int(conf)
        except Exception: conf_int = 0
        if filtro == "Ver todos": return True
        if filtro == "Solo seleccionados": return usar == CHECK_ON
        if filtro == "Solo nuevos": return existe == "No"
        if filtro == "Solo existentes": return existe in ("Si", "Sí")
        if filtro == "Sin codigo Tango": return not codigo
        if filtro == "Baja confianza": return existe in ("Si", "Sí") and conf_int < self.umbral_confianza.get()
        if filtro == "Destildados": return usar == CHECK_OFF
        if filtro == "Con advertencias": return bool(advertencia and advertencia != "OK")
        return True

    def actualizar_label_facturas(self):
        nombres = []
        for f in self.facturas_cargadas:
            etiqueta = f.get("numero_completo") or f.get("archivo", "")
            proveedor = f.get("proveedor", "")
            if proveedor:
                etiqueta = f"{proveedor[:25]} {etiqueta}"
            nombres.append(etiqueta)
        detalle = " | ".join(nombres[:4]) + (f" | ... (+{len(nombres)-4})" if len(nombres) > 4 else "")
        self.lbl_facturas.config(text=f"Facturas cargadas: {len(self.facturas_cargadas)}   {detalle}")

    def ver_debug(self):
        if not self.resultados_debug:
            messagebox.showerror("Error", "Primero agregá una factura.")
            return
        for archivo, pagina, debug_path, lineas, zonas in self.resultados_debug:
            try: os.startfile(debug_path)
            except Exception as e:
                messagebox.showerror("Error", str(e)); return

    def abrir_factura_original(self):
        item_id = self.tabla.focus()
        if not item_id:
            messagebox.showerror("Error", "Seleccioná una fila.")
            return
        v = self.tabla.item(item_id, "values")
        ruta = v[23]
        if not ruta or not os.path.exists(ruta):
            messagebox.showerror("Error", "No se encontró la ruta de la factura original.")
            return
        os.startfile(ruta)

    def leer_filas(self, solo_visibles=False):
        ids = self.tabla.get_children("") if solo_visibles else self._todos_items
        filas = []
        for item_id in ids:
            try: v = self.tabla.item(item_id, "values")
            except Exception: continue
            if not v: continue
            filas.append({
                "usar": v[0], "factura": v[1], "proveedor": v[2], "cuit": v[3],
                "tipo_comprobante": v[4], "numero_comprobante": v[5], "fecha": v[6],
                "existe": v[7], "accion": v[8], "codigo_tango": v[9].strip(),
                "descripcion": v[10].strip(), "sinonimo": v[11].strip(), "codigo_barras": v[12].strip(),
                "cantidad": v[13].strip(), "precio": v[14].strip(), "importe": v[15].strip(),
                "iva": v[16].strip(), "advertencia": v[17].strip(), "coincidencia": v[18].strip(),
                "match_score": v[19].strip(), "tipo": v[20].strip() or "Simple", "escala": v[21].strip() or "No usa",
                "pagina": v[22], "ruta": v[23], "linea": v[24],
            })
        return filas

    def deduplicar_nuevos(self, nuevos):
        normalizador = CatalogoTango()
        vistos, resultado, duplicados = {}, [], []
        for f in nuevos:
            clave = (normalizador.normalizar(f["sinonimo"]), normalizador.normalizar(f["codigo_barras"]), normalizador.normalizar(f["descripcion"]))
            if clave in vistos:
                duplicados.append((f, vistos[clave])); continue
            vistos[clave] = f; resultado.append(f)
        return resultado, duplicados

    def aprender_equivalencia_seleccionada(self):
        sel = self.tabla.selection()
        if not sel:
            messagebox.showerror("Error", "Seleccioná una fila para aprender la equivalencia.")
            return
        aprendidas = 0
        for item_id in sel:
            v = self.tabla.item(item_id, "values")
            cuit, proveedor, cod_prov, linea, cod_tango, desc, barras = v[3], v[2], v[11], v[24], v[9], v[10], v[12]
            if not cuit or not cod_prov or not cod_tango:
                continue
            self.base.guardar_equivalencia(cuit, proveedor, cod_prov, linea, cod_tango, desc, barras)
            aprendidas += 1
        messagebox.showinfo("Listo", f"Equivalencias aprendidas: {aprendidas}")
        self.recomparar(silencioso=True)

    def aprender_equivalencias_automaticas(self):
        for f in self.leer_filas():
            if f["usar"] != CHECK_ON: continue
            if not f["cuit"] or not f["sinonimo"] or not f["codigo_tango"]: continue
            self.base.guardar_equivalencia(
                f["cuit"], f["proveedor"], f["sinonimo"], f["linea"],
                f["codigo_tango"], f["descripcion"], f["codigo_barras"]
            )

    def generar_excel_alta_articulos(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo Tango.")
            return
        filas = self.leer_filas()
        nuevos = [f for f in filas if f["usar"] == CHECK_ON and f["existe"] == "No"]
        if not nuevos:
            messagebox.showinfo("Sin artículos nuevos", "No hay artículos nuevos para dar de alta.")
            return
        if [f for f in nuevos if not f["codigo_tango"]]:
            messagebox.showerror("Faltan códigos Tango", "Hay artículos nuevos sin Código Tango.\n\nCompletalos antes de generar el Excel.")
            return
        nuevos_unicos, duplicados = self.deduplicar_nuevos(nuevos)
        if duplicados:
            messagebox.showinfo("Duplicados agrupados", f"Se encontraron {len(duplicados)} artículos nuevos repetidos.\nSe exportará solo una vez cada artículo nuevo.")
        salida = filedialog.asksaveasfilename(title="Guardar Excel Alta Artículos Tango", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Alta_Articulos_Tango.xlsx")
        if not salida: return
        try:
            wb = load_workbook(self.ruta_modelo_tango)
            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja 'Artículos'."); return
            ws = wb["Artículos"]
            if ws.max_row > 1: ws.delete_rows(2, ws.max_row - 1)
            for fila, item in enumerate(nuevos_unicos, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = ""
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]
            wb.save(salida)
            self.aprender_equivalencias_automaticas()
            messagebox.showinfo("Listo", f"Excel de alta de artículos generado:\n\n{salida}\n\nArtículos nuevos únicos: {len(nuevos_unicos)}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def generar_resumen_compra_stock(self):
        filas = [f for f in self.leer_filas() if f["usar"] == CHECK_ON]
        if not filas:
            messagebox.showerror("Error", "No hay filas seleccionadas."); return
        salida = filedialog.asksaveasfilename(title="Guardar resumen compra/stock", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Resumen_Compra_Stock.xlsx")
        if not salida: return
        self.crear_excel_resumen(salida, filas)
        for f in self.facturas_cargadas:
            self.base.guardar_factura_procesada(f, archivo_resumen=salida)
        self.aprender_equivalencias_automaticas()
        messagebox.showinfo("Listo", f"Resumen de compra/stock generado:\n\n{salida}\n\nFacturas registradas en historial permanente: {len(self.facturas_cargadas)}")

    def crear_excel_resumen(self, salida, filas):
        wb = Workbook()
        bold = Font(bold=True)
        ws = wb.active
        ws.title = "Compra_Stock"
        headers = [
            "Factura", "Proveedor", "CUIT", "Tipo comprobante", "Número comprobante", "Fecha",
            "Existe en Tango", "Acción", "Código Tango", "Descripción Tango/OCR",
            "Código proveedor", "Código barras", "Cantidad", "Precio", "Importe", "IVA %",
            "Advertencia", "Coincidencia", "Confianza", "Línea detectada"
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col).value = h
            ws.cell(1, col).font = bold
        for row, f in enumerate(filas, start=2):
            values = [
                f["factura"], f["proveedor"], f["cuit"], f["tipo_comprobante"], f["numero_comprobante"],
                f["fecha"], f["existe"], f["accion"], f["codigo_tango"], f["descripcion"], f["sinonimo"],
                f["codigo_barras"], f["cantidad"], f["precio"], f["importe"], f["iva"], f["advertencia"],
                f["coincidencia"], f["match_score"], f["linea"]
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row, col).value = value
        widths = [25, 35, 15, 18, 18, 12, 16, 22, 18, 50, 18, 18, 12, 12, 12, 10, 30, 30, 12, 80]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width

        ws2 = wb.create_sheet("Facturas_Cargadas")
        h2 = ["Clave", "Archivo", "Proveedor detectado", "CUIT proveedor", "Tipo", "Punto de venta", "Número", "Número completo", "Fecha", "Total", "Ruta"]
        for col, h in enumerate(h2, start=1):
            ws2.cell(1, col).value = h; ws2.cell(1, col).font = bold
        for row, f in enumerate(self.facturas_cargadas, start=2):
            ws2.cell(row, 1).value = f["clave"]; ws2.cell(row, 2).value = f["archivo"]
            ws2.cell(row, 3).value = f["proveedor"]; ws2.cell(row, 4).value = f["cuit_proveedor"]
            ws2.cell(row, 5).value = f["tipo"]; ws2.cell(row, 6).value = f["punto_venta"]
            ws2.cell(row, 7).value = f["numero"]; ws2.cell(row, 8).value = f["numero_completo"]
            ws2.cell(row, 9).value = f["fecha"]; ws2.cell(row, 10).value = f.get("total", "")
            ws2.cell(row, 11).value = f["ruta"]
        for col in range(1, 12):
            ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 25
        wb.save(salida)

    def generar_reporte_revision(self):
        filas = self.leer_filas()
        if not filas:
            messagebox.showerror("Error", "No hay filas para reportar."); return
        salida = filedialog.asksaveasfilename(title="Guardar reporte de revisión", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Reporte_Revision.xlsx")
        if not salida: return
        wb = Workbook()
        bold = Font(bold=True)
        ws = wb.active; ws.title = "Resumen"
        resumen = [
            ("Filas totales", len(filas)),
            ("Filas seleccionadas", len([f for f in filas if f["usar"] == CHECK_ON])),
            ("Articulos existentes", len([f for f in filas if f["existe"] in ("Si", "Sí")])),
            ("Articulos nuevos", len([f for f in filas if f["existe"] == "No"])),
            ("Sin código Tango", len([f for f in filas if not f["codigo_tango"]])),
            ("Advertencias cantidad/precio/importe", len([f for f in filas if f["advertencia"] and f["advertencia"] != "OK"])),
            ("Facturas cargadas", len(self.facturas_cargadas)),
        ]
        for row, (k, v) in enumerate(resumen, start=1):
            ws.cell(row, 1).value = k; ws.cell(row, 1).font = bold; ws.cell(row, 2).value = v
        ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 20
        ws2 = wb.create_sheet("Detalle")
        headers = ["Usar", "Factura", "Proveedor", "CUIT", "Número", "Existe", "Acción", "Código Tango", "Descripción", "Código proveedor", "Cantidad", "Precio", "Importe", "IVA", "Advertencia", "Coincidencia", "Confianza"]
        for col, h in enumerate(headers, start=1):
            ws2.cell(1, col).value = h; ws2.cell(1, col).font = bold
        for row, f in enumerate(filas, start=2):
            vals = [f["usar"], f["factura"], f["proveedor"], f["cuit"], f["numero_comprobante"], f["existe"], f["accion"], f["codigo_tango"], f["descripcion"], f["sinonimo"], f["cantidad"], f["precio"], f["importe"], f["iva"], f["advertencia"], f["coincidencia"], f["match_score"]]
            for col, val in enumerate(vals, start=1):
                ws2.cell(row, col).value = val
        for col in range(1, len(headers)+1):
            ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 22
        wb.save(salida)
        messagebox.showinfo("Listo", f"Reporte de revisión generado:\n\n{salida}")

    def guardar_sesion(self):
        salida = filedialog.asksaveasfilename(title="Guardar sesión", defaultextension=".json", filetypes=[("JSON", "*.json")], initialfile=f"Sesion_Revision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        if not salida: return
        data = {
            "version": "V8.1", "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ruta_modelo_tango": self.ruta_modelo_tango, "ruta_catalogo_tango": self.ruta_catalogo_tango,
            "facturas_cargadas": self.facturas_cargadas, "filas": self.leer_filas()
        }
        with open(salida, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("Listo", f"Sesión guardada:\n\n{salida}")

    def cargar_sesion(self):
        ruta = filedialog.askopenfilename(title="Cargar sesión", filetypes=[("JSON", "*.json")])
        if not ruta: return
        with open(ruta, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.limpiar_todo_sin_preguntar()
        self.ruta_modelo_tango = data.get("ruta_modelo_tango")
        self.ruta_catalogo_tango = data.get("ruta_catalogo_tango")
        if self.ruta_modelo_tango:
            self.lbl_modelo.config(text=f"Modelo Tango: {self.ruta_modelo_tango}")
        if self.ruta_catalogo_tango and os.path.exists(self.ruta_catalogo_tango):
            try:
                self.catalogo.cargar(self.ruta_catalogo_tango)
                self.lbl_catalogo.config(text=f"Artículos Tango: {self.ruta_catalogo_tango} ({len(self.catalogo.articulos)} artículos)")
            except Exception:
                pass
        self.facturas_cargadas = data.get("facturas_cargadas", [])
        self.claves_facturas = set(f.get("clave", "") for f in self.facturas_cargadas)
        for f in data.get("filas", []):
            item = {
                "usar": f.get("usar") == CHECK_ON, "factura_etiqueta": f.get("factura", ""),
                "proveedor_factura": f.get("proveedor", ""), "cuit_proveedor": f.get("cuit", ""),
                "tipo_factura": f.get("tipo_comprobante", ""), "numero_completo": f.get("numero_comprobante", ""),
                "fecha_factura": f.get("fecha", ""), "existe_tango": f.get("existe", "No"),
                "accion": f.get("accion", ""), "codigo_tango": f.get("codigo_tango", ""),
                "descripcion": f.get("descripcion", ""), "sinonimo": f.get("sinonimo", ""),
                "codigo_barras": f.get("codigo_barras", ""), "cantidad": f.get("cantidad", ""),
                "precio": f.get("precio", ""), "importe": f.get("importe", ""), "iva": f.get("iva", ""),
                "advertencia": f.get("advertencia", ""), "tipo": f.get("tipo", "Simple"),
                "escala": f.get("escala", "No usa"), "pagina": f.get("pagina", ""),
                "ruta_factura": f.get("ruta", ""), "linea_detectada": f.get("linea", ""),
                "coincidencia": f.get("coincidencia", ""), "match_score": f.get("match_score", ""),
            }
            self.insertar_item(item)
        self.actualizar_label_facturas(); self.actualizar_tags(); self.aplicar_filtro()
        messagebox.showinfo("Listo", "Sesión cargada correctamente.")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()