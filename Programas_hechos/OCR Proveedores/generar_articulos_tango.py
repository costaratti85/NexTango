import os
import re
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from PIL import Image, ImageOps, ImageEnhance, ImageFilter
from openpyxl import load_workbook
import pytesseract
import fitz

try:
    import cv2
except ImportError:
    cv2 = None


pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("OCR Factura → Excel Artículos Tango")
        self.root.geometry("1200x640")

        self.ruta_factura = None
        self.ruta_modelo_tango = None
        self.texto_ocr = ""

        self.crear_interfaz()

    def crear_interfaz(self):
        frame = tk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=10)

        tk.Button(frame, text="1. Seleccionar factura imagen/PDF", command=self.seleccionar_factura, width=32).pack(side="left", padx=4)
        tk.Button(frame, text="Sacar foto", command=self.sacar_foto, width=14).pack(side="left", padx=4)
        tk.Button(frame, text="2. Seleccionar modelo Tango", command=self.seleccionar_modelo, width=28).pack(side="left", padx=4)
        tk.Button(frame, text="3. Generar Excel Tango", command=self.generar_excel_tango, width=25).pack(side="left", padx=4)

        self.lbl_factura = tk.Label(self.root, text="Factura: no seleccionada", anchor="w")
        self.lbl_factura.pack(fill="x", padx=15)

        self.lbl_modelo = tk.Label(self.root, text="Modelo Tango: no seleccionado", anchor="w")
        self.lbl_modelo.pack(fill="x", padx=15)

        columnas = (
            "codigo_tango",
            "descripcion",
            "desc_adic",
            "sinonimo",
            "codigo_barras",
            "tipo",
            "escala",
            "linea_ocr",
        )

        self.tabla = ttk.Treeview(self.root, columns=columnas, show="headings", height=20)

        headers = {
            "codigo_tango": "Código Tango",
            "descripcion": "Descripción",
            "desc_adic": "Desc. adicional",
            "sinonimo": "Sinónimo / Cod. proveedor",
            "codigo_barras": "Código barras",
            "tipo": "Tipo",
            "escala": "Escala",
            "linea_ocr": "Línea OCR original",
        }

        widths = {
            "codigo_tango": 120,
            "descripcion": 330,
            "desc_adic": 120,
            "sinonimo": 150,
            "codigo_barras": 150,
            "tipo": 90,
            "escala": 90,
            "linea_ocr": 450,
        }

        for col in columnas:
            self.tabla.heading(col, text=headers[col])
            self.tabla.column(col, width=widths[col])

        self.tabla.pack(fill="both", expand=True, padx=10, pady=10)
        self.tabla.bind("<Double-1>", self.editar_celda)

        frame_bot = tk.Frame(self.root)
        frame_bot.pack(fill="x", padx=10, pady=5)

        tk.Button(frame_bot, text="Agregar artículo manual", command=self.agregar_manual, width=24).pack(side="left", padx=5)
        tk.Button(frame_bot, text="Eliminar seleccionado", command=self.eliminar_seleccionado, width=22).pack(side="left", padx=5)
        tk.Button(frame_bot, text="Ver texto leído", command=self.ver_texto_ocr, width=18).pack(side="left", padx=5)

        tk.Label(
            self.root,
            text="Doble click para editar. El programa ahora intenta detectar solo renglones reales de artículos.",
            fg="blue"
        ).pack(pady=4)

    def seleccionar_factura(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar factura",
            filetypes=[
                ("Facturas", "*.jpg *.jpeg *.png *.bmp *.tif *.tiff *.pdf"),
                ("Todos", "*.*"),
            ]
        )
        if not ruta:
            return

        self.ruta_factura = ruta
        self.lbl_factura.config(text=f"Factura: {ruta}")
        self.procesar_factura(ruta)

    def sacar_foto(self):
        if cv2 is None:
            messagebox.showerror("Error", "Falta instalar:\n\npip install opencv-python")
            return

        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "No se pudo abrir la cámara.")
            return

        messagebox.showinfo("Cámara", "ESPACIO captura / ESC cancela")

        ruta_foto = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            cv2.imshow("Capturar factura", frame)
            tecla = cv2.waitKey(1)

            if tecla == 27:
                break
            if tecla == 32:
                ruta_foto = os.path.join(tempfile.gettempdir(), "factura_capturada.jpg")
                cv2.imwrite(ruta_foto, frame)
                break

        cap.release()
        cv2.destroyAllWindows()

        if ruta_foto:
            self.ruta_factura = ruta_foto
            self.lbl_factura.config(text=f"Factura: {ruta_foto}")
            self.procesar_factura(ruta_foto)

    def seleccionar_modelo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel modelo exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )
        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def preparar_imagen(self, img):
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.0)
        img = ImageEnhance.Sharpness(img).enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)
        return img

    def leer_pdf_texto_real(self, ruta_pdf):
        texto_total = ""
        doc = fitz.open(ruta_pdf)

        for page in doc:
            texto_total += page.get_text("text") + "\n"

        return texto_total.strip()

    def pdf_a_imagenes(self, ruta_pdf):
        doc = fitz.open(ruta_pdf)
        imagenes = []

        for pagina in doc:
            pix = pagina.get_pixmap(matrix=fitz.Matrix(2, 2))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            imagenes.append(img)

        return imagenes

    def hacer_ocr(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()

        if ext == ".pdf":
            texto_pdf = self.leer_pdf_texto_real(ruta)

            if len(texto_pdf) > 100:
                return texto_pdf

            imagenes = self.pdf_a_imagenes(ruta)
        else:
            imagenes = [Image.open(ruta)]

        textos = []

        for img in imagenes:
            img = self.preparar_imagen(img)

            try:
                texto = pytesseract.image_to_string(img, lang="spa")
            except Exception:
                texto = pytesseract.image_to_string(img)

            textos.append(texto)

        return "\n".join(textos)

    def procesar_factura(self, ruta):
        try:
            self.texto_ocr = self.hacer_ocr(ruta)
            items = self.detectar_items_refinado(self.texto_ocr)
            self.cargar_items_tabla(items)

            messagebox.showinfo(
                "Lectura terminada",
                f"Se detectaron {len(items)} artículos probables.\n\nRevisá y completá Código Tango."
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def detectar_items_refinado(self, texto):
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]
        items = []

        patron_inicio_item = re.compile(
            r"^([A-Z]{2,}[A-Z0-9]{3,})\s+(.+?)\s+([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})\s+([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})$"
        )

        patron_cantidad = re.compile(
            r"^([0-9]+,[0-9]{2})\s+U\s+"
        )

        palabras_corte = [
            "IMPORTE NETO",
            "TOTAL:",
            "IVA 21",
            "CAE",
            "FECHA DE VTO",
            "SON PESOS",
            "OBSERVACIONES",
            "LA MERCADERIA",
            "BONIFICACION",
            "SUBTOTAL",
        ]

        i = 0

        while i < len(lineas):
            linea = lineas[i]
            m = patron_inicio_item.match(linea)

            if not m:
                i += 1
                continue

            codigo_proveedor = m.group(1).strip()
            descripcion_partes = [m.group(2).strip()]
            linea_original = [linea]

            i += 1

            while i < len(lineas):
                sig = lineas[i].strip()
                sig_upper = sig.upper()

                if patron_inicio_item.match(sig):
                    break

                if any(p in sig_upper for p in palabras_corte):
                    break

                if patron_cantidad.match(sig):
                    linea_original.append(sig)
                    i += 1
                    continue

                if re.match(r"^\([0-9A-Z]+\)", sig):
                    linea_original.append(sig)
                    i += 1
                    continue

                if len(sig) > 3 and not self.es_linea_basura(sig):
                    descripcion_partes.append(sig)
                    linea_original.append(sig)

                i += 1

            descripcion = " ".join(descripcion_partes)
            descripcion = re.sub(r"\s+", " ", descripcion).strip()
            descripcion = descripcion.replace(" /", "").replace("/", "")

            items.append({
                "codigo_tango": "",
                "descripcion": descripcion[:50],
                "desc_adic": "",
                "sinonimo": codigo_proveedor[:15],
                "codigo_barras": "",
                "tipo": "Simple",
                "escala": "No usa",
                "linea_ocr": " | ".join(linea_original),
            })

        return items

    def es_linea_basura(self, linea):
        mayus = linea.upper()

        excluir = [
            "FACTURA",
            "ORIGINAL",
            "CUIT",
            "DOMICILIO",
            "RESPONSABLE",
            "FECHA",
            "PUNTO DE VENTA",
            "COMP. NRO",
            "SEÑOR",
            "CONDICIÓN",
            "CONDICION",
            "VENDEDOR",
            "REMITO",
            "CÓDIGO:",
            "CODIGO:",
            "PRECIO LISTA",
            "IVA%",
            "DESC %",
            "CAPITAL FEDERAL",
            "ARGENTINA",
            "TELEFONO",
            "PÁG.",
            "PAG.",
        ]

        return any(x in mayus for x in excluir)

    def cargar_items_tabla(self, items):
        for item in self.tabla.get_children():
            self.tabla.delete(item)

        for item in items:
            self.tabla.insert(
                "",
                "end",
                values=(
                    item["codigo_tango"],
                    item["descripcion"],
                    item["desc_adic"],
                    item["sinonimo"],
                    item["codigo_barras"],
                    item["tipo"],
                    item["escala"],
                    item["linea_ocr"],
                )
            )

    def editar_celda(self, event):
        item_id = self.tabla.focus()
        if not item_id:
            return

        columna = self.tabla.identify_column(event.x)
        col_index = int(columna.replace("#", "")) - 1

        campos = [
            "Código Tango",
            "Descripción",
            "Desc. adicional",
            "Sinónimo",
            "Código barras",
            "Tipo",
            "Escala",
            "Línea OCR",
        ]

        valores = list(self.tabla.item(item_id, "values"))
        actual = valores[col_index]

        nuevo = simpledialog.askstring("Editar", campos[col_index], initialvalue=actual)

        if nuevo is not None:
            valores[col_index] = nuevo
            self.tabla.item(item_id, values=valores)

    def leer_items_tabla(self):
        items = []

        for item_id in self.tabla.get_children():
            v = self.tabla.item(item_id, "values")
            items.append({
                "codigo_tango": v[0].strip(),
                "descripcion": v[1].strip(),
                "desc_adic": v[2].strip(),
                "sinonimo": v[3].strip(),
                "codigo_barras": v[4].strip(),
                "tipo": v[5].strip() or "Simple",
                "escala": v[6].strip() or "No usa",
            })

        return items

    def agregar_manual(self):
        self.tabla.insert("", "end", values=("", "", "", "", "", "Simple", "No usa", "Manual"))

    def eliminar_seleccionado(self):
        for item in self.tabla.selection():
            self.tabla.delete(item)

    def ver_texto_ocr(self):
        w = tk.Toplevel(self.root)
        w.title("Texto leído")
        w.geometry("900x600")

        txt = tk.Text(w, wrap="word")
        txt.pack(fill="both", expand=True)
        txt.insert("1.0", self.texto_ocr)

    def generar_excel_tango(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo de Tango.")
            return

        items = self.leer_items_tabla()

        if not items:
            messagebox.showerror("Error", "No hay artículos.")
            return

        if any(not i["codigo_tango"] for i in items):
            messagebox.showerror("Faltan códigos", "Hay artículos sin Código Tango.")
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar Excel Tango",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Articulos_Para_Importar_Tango.xlsx"
        )

        if not salida:
            return

        try:
            wb = load_workbook(self.ruta_modelo_tango)

            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja Artículos.")
                return

            ws = wb["Artículos"]

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for fila, item in enumerate(items, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = item["desc_adic"][:20]
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]

            wb.save(salida)

            messagebox.showinfo("Listo", f"Excel generado:\n\n{salida}")

        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()