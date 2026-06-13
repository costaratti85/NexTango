import os
import re
import tempfile
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from dataclasses import dataclass
from typing import List
from difflib import SequenceMatcher
from PIL import Image, ImageDraw, ImageOps, ImageEnhance
from openpyxl import load_workbook, Workbook
import fitz
import pytesseract

try:
    import cv2
except ImportError:
    cv2 = None


pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

CHECK_ON = "☑"
CHECK_OFF = "☐"


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    page: int


@dataclass
class LineBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    words: List[WordBox]
    score: int = 0


class CatalogoTango:
    def __init__(self):
        self.articulos = []

    def quitar_acentos(self, texto):
        texto = unicodedata.normalize("NFD", texto)
        texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
        return texto

    def normalizar(self, texto):
        if texto is None:
            return ""

        texto = str(texto).upper().strip()
        texto = self.quitar_acentos(texto)

        reemplazos = {
            "P/ACERO": "ACERO",
            "P ACERO": "ACERO",
            "PACERO": "ACERO",
            "P/ ACERO": "ACERO",
            "P/INOX": "INOX",
            "P INOX": "INOX",
            "P/ INOX": "INOX",
            "INOXIDABLE": "INOX",
            "C/": "CON ",
            "S/": "SIN ",
            "MM.": "MM",
            "M.M.": "MM",
            "  ": " ",
        }

        for viejo, nuevo in reemplazos.items():
            texto = texto.replace(viejo, nuevo)

        texto = texto.replace("/", " ")
        texto = texto.replace("-", " ")
        texto = texto.replace(",", " ")
        texto = texto.replace(".", " ")
        texto = texto.replace("(", " ")
        texto = texto.replace(")", " ")
        texto = texto.replace("[", " ")
        texto = texto.replace("]", " ")

        texto = re.sub(r"\s+", " ", texto).strip()
        return texto

    def tokens(self, texto):
        texto = self.normalizar(texto)
        return [t for t in texto.split() if len(t) >= 2]

    def cargar(self, ruta_excel):
        wb = load_workbook(ruta_excel, data_only=True)

        if "Artículos" not in wb.sheetnames:
            raise Exception("El Excel no tiene una hoja llamada 'Artículos'.")

        ws = wb["Artículos"]
        self.articulos = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = row[0] if len(row) > 0 else ""
            descripcion = row[1] if len(row) > 1 else ""
            desc_adic = row[2] if len(row) > 2 else ""
            sinonimo = row[3] if len(row) > 3 else ""
            codigo_barras = row[4] if len(row) > 4 else ""

            if not codigo:
                continue

            art = {
                "codigo": str(codigo).strip(),
                "descripcion": str(descripcion or "").strip(),
                "desc_adic": str(desc_adic or "").strip(),
                "sinonimo": str(sinonimo or "").strip(),
                "codigo_barras": str(codigo_barras or "").strip(),
            }

            art["_codigo_n"] = self.normalizar(art["codigo"])
            art["_descripcion_n"] = self.normalizar(art["descripcion"])
            art["_desc_adic_n"] = self.normalizar(art["desc_adic"])
            art["_sinonimo_n"] = self.normalizar(art["sinonimo"])
            art["_codigo_barras_n"] = self.normalizar(art["codigo_barras"])
            art["_texto_completo_n"] = self.normalizar(
                f"{art['codigo']} {art['descripcion']} {art['desc_adic']} {art['sinonimo']} {art['codigo_barras']}"
            )

            self.articulos.append(art)

    def similitud(self, a, b):
        a = self.normalizar(a)
        b = self.normalizar(b)

        if not a or not b:
            return 0

        return int(SequenceMatcher(None, a, b).ratio() * 100)

    def score_tokens(self, texto_ocr, texto_tango):
        tokens_ocr = set(self.tokens(texto_ocr))
        tokens_tango = set(self.tokens(texto_tango))

        if not tokens_ocr or not tokens_tango:
            return 0

        comunes = tokens_ocr.intersection(tokens_tango)
        return int((len(comunes) / len(tokens_ocr)) * 100)

    def buscar(self, codigo_proveedor="", codigo_barras="", descripcion=""):
        codigo_proveedor_n = self.normalizar(codigo_proveedor)
        codigo_barras_n = self.normalizar(codigo_barras)
        descripcion_n = self.normalizar(descripcion)

        if codigo_barras_n:
            for art in self.articulos:
                if art["_codigo_barras_n"] and art["_codigo_barras_n"] == codigo_barras_n:
                    return art, "Código de barras exacto", 100

        if codigo_proveedor_n:
            for art in self.articulos:
                if art["_codigo_n"] == codigo_proveedor_n:
                    return art, "Código Tango exacto", 100

            for art in self.articulos:
                if art["_sinonimo_n"] and art["_sinonimo_n"] == codigo_proveedor_n:
                    return art, "Sinónimo exacto", 100

            for art in self.articulos:
                if codigo_proveedor_n in art["_descripcion_n"]:
                    return art, "Código proveedor dentro de descripción", 96

            for art in self.articulos:
                if codigo_proveedor_n in art["_texto_completo_n"]:
                    return art, "Código proveedor contenido", 94

        mejor_art = None
        mejor_metodo = ""
        mejor_score = 0

        if descripcion_n:
            for art in self.articulos:
                score_desc = self.similitud(descripcion_n, art["_descripcion_n"])
                score_tokens = self.score_tokens(descripcion_n, art["_descripcion_n"])

                score_final = max(score_desc, score_tokens)

                if score_final > mejor_score:
                    mejor_score = score_final
                    mejor_art = art
                    mejor_metodo = "Similitud descripción"

            if mejor_art and mejor_score >= 82:
                return mejor_art, mejor_metodo, mejor_score

        return None, "", 0


class FacturaTableReader:
    def __init__(self):
        self.stop_words = [
            "total", "subtotal", "cae", "iva 21", "iva 10", "iva 27",
            "son pesos", "observaciones", "importe neto", "fecha de vto"
        ]

    def is_cuit(self, text):
        clean = text.strip()

        if re.fullmatch(r"\d{2}-\d{8}-\d", clean):
            return True

        only_digits = re.sub(r"\D", "", clean)

        if len(only_digits) == 11 and only_digits[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
            return True

        return False

    def abrir_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()

        if ext == ".pdf":
            return self.leer_pdf(ruta)

        return self.leer_imagen(ruta)

    def leer_pdf(self, ruta):
        doc = fitz.open(ruta)
        paginas = []

        for page_index, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(tempfile.gettempdir(), f"pagina_factura_{page_index}.png")
            pix.save(img_path)
            img = Image.open(img_path)

            palabras = []
            escala = 2

            for w in page.get_text("words"):
                x0, y0, x1, y1, text = w[:5]
                text = str(text).strip()

                if text:
                    palabras.append(
                        WordBox(
                            x0 * escala,
                            y0 * escala,
                            x1 * escala,
                            y1 * escala,
                            text,
                            page_index
                        )
                    )

            paginas.append((img, palabras))

        return paginas

    def leer_imagen(self, ruta):
        img = Image.open(ruta)
        img_ocr = self.preparar_imagen(img)

        data = pytesseract.image_to_data(
            img_ocr,
            lang="spa",
            output_type=pytesseract.Output.DICT,
            config="--psm 6"
        )

        palabras = []

        for i, text in enumerate(data["text"]):
            text = str(text).strip()

            if not text:
                continue

            try:
                conf = int(float(data["conf"][i]))
            except Exception:
                conf = -1

            if conf < 25:
                continue

            x = data["left"][i]
            y = data["top"][i]
            w = data["width"][i]
            h = data["height"][i]

            palabras.append(WordBox(x, y, x + w, y + h, text, 1))

        return [(img, palabras)]

    def preparar_imagen(self, img):
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.2)
        return img

    def agrupar_lineas(self, palabras):
        palabras = sorted(palabras, key=lambda w: (w.y0, w.x0))
        lineas = []
        actual = []

        for palabra in palabras:
            if not actual:
                actual = [palabra]
                continue

            y_prom = sum(w.y0 for w in actual) / len(actual)

            if abs(palabra.y0 - y_prom) <= 8:
                actual.append(palabra)
            else:
                lineas.append(self.crear_linea(actual))
                actual = [palabra]

        if actual:
            lineas.append(self.crear_linea(actual))

        return lineas

    def crear_linea(self, palabras):
        palabras = sorted(palabras, key=lambda w: w.x0)

        return LineBox(
            x0=min(w.x0 for w in palabras),
            y0=min(w.y0 for w in palabras),
            x1=max(w.x1 for w in palabras),
            y1=max(w.y1 for w in palabras),
            text=" ".join(w.text for w in palabras),
            words=palabras
        )

    def es_numero(self, t):
        t = t.replace("$", "").strip()

        if self.is_cuit(t):
            return False

        return (
            bool(re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2,4}$", t))
            or bool(re.match(r"^-?\d+([,.]\d+)?$", t))
        )

    def cantidad_columnas(self, linea):
        if len(linea.words) < 3:
            return 0

        gaps = 0

        for a, b in zip(linea.words, linea.words[1:]):
            if b.x0 - a.x1 > 25:
                gaps += 1

        return gaps + 1

    def puntuar_linea(self, linea):
        texto = linea.text.lower()

        if any(self.is_cuit(w.text) for w in linea.words):
            linea.score = -10
            return linea.score

        if any(p in texto for p in self.stop_words):
            linea.score = -5
            return linea.score

        numeros = sum(1 for w in linea.words if self.es_numero(w.text))
        columnas = self.cantidad_columnas(linea)

        score = 0

        if columnas >= 3:
            score += 2
        if columnas >= 5:
            score += 2
        if numeros >= 2:
            score += 2
        if numeros >= 4:
            score += 2
        if len(linea.words) >= 5:
            score += 1

        linea.score = score
        return score

    def detectar_zonas(self, lineas):
        for linea in lineas:
            self.puntuar_linea(linea)

        candidatas = [l for l in lineas if l.score >= 4]

        if not candidatas:
            return []

        grupos = []
        grupo = [candidatas[0]]

        for linea in candidatas[1:]:
            if linea.y0 - grupo[-1].y1 < 75:
                grupo.append(linea)
            else:
                grupos.append(grupo)
                grupo = [linea]

        grupos.append(grupo)

        zonas = []

        for grupo in grupos:
            if len(grupo) < 2:
                continue

            zonas.append((
                min(l.x0 for l in grupo),
                min(l.y0 for l in grupo),
                max(l.x1 for l in grupo),
                max(l.y1 for l in grupo),
                grupo
            ))

        return zonas

    def detectar_items(self, lineas):
        zonas = self.detectar_zonas(lineas)
        items = []

        for x0, y0, x1, y1, grupo in zonas:
            for linea in grupo:
                if linea.score < 4:
                    continue

                if any(self.is_cuit(w.text) for w in linea.words):
                    continue

                item = self.linea_a_item(linea)

                if item:
                    items.append(item)

        return items, zonas

    def linea_a_item(self, linea):
        words = linea.words

        codigo_proveedor = ""
        descripcion_parts = []
        cantidad = ""
        precio = ""
        importe = ""
        numeros_detectados = []

        for w in words:
            t = w.text.strip()

            if self.is_cuit(t):
                return None

            if self.es_numero(t):
                numeros_detectados.append(t)
                continue

            if not codigo_proveedor and re.match(r"^[A-Z0-9][A-Z0-9.\-/]{2,}$", t, re.IGNORECASE):
                codigo_proveedor = t[:15]
                continue

            if t.upper() in ["U", "UN", "KG", "M", "L", "%"]:
                continue

            if re.match(r"^-?\d+([,.]\d+)?%?$", t):
                continue

            descripcion_parts.append(t)

        if numeros_detectados:
            cantidad = numeros_detectados[0]

        if len(numeros_detectados) >= 2:
            precio = numeros_detectados[-2]

        if len(numeros_detectados) >= 1:
            importe = numeros_detectados[-1]

        descripcion = " ".join(descripcion_parts)
        descripcion = re.sub(r"\s+", " ", descripcion).strip()
        descripcion = descripcion.replace(" /", "").replace("/", "")

        if not descripcion and not codigo_proveedor:
            return None

        return {
            "usar": True,
            "existe_tango": "No",
            "accion": "Alta artículo",
            "codigo_tango": "",
            "descripcion": descripcion[:50],
            "desc_adic": "",
            "sinonimo": codigo_proveedor[:15],
            "codigo_barras": "",
            "cantidad": cantidad,
            "precio": precio,
            "importe": importe,
            "tipo": "Simple",
            "escala": "No usa",
            "linea_detectada": linea.text,
            "score": linea.score,
            "coincidencia": "",
            "match_score": 0,
        }

    def analizar(self, ruta):
        paginas = self.abrir_documento(ruta)
        resultados = []
        todos_items = []

        for page_num, (img, palabras) in enumerate(paginas, start=1):
            lineas = self.agrupar_lineas(palabras)
            items, zonas = self.detectar_items(lineas)
            debug_path = self.dibujar_debug(img, lineas, zonas, page_num)

            for item in items:
                item["pagina"] = page_num
                todos_items.append(item)

            resultados.append((page_num, debug_path, lineas, zonas))

        return todos_items, resultados

    def dibujar_debug(self, img, lineas, zonas, page_num):
        debug = img.convert("RGB")
        draw = ImageDraw.Draw(debug)

        for linea in lineas:
            if linea.score >= 4:
                draw.rectangle([linea.x0, linea.y0, linea.x1, linea.y1], outline="blue", width=2)
                draw.text((linea.x0, max(0, linea.y0 - 14)), f"S{linea.score}", fill="blue")

        for x0, y0, x1, y1, grupo in zonas:
            draw.rectangle([x0, y0, x1, y1], outline="red", width=4)

        out = os.path.join(tempfile.gettempdir(), f"debug_factura_tango_pagina_{page_num}.png")
        debug.save(out)
        return out


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Factura → Tango V4 Matching Inteligente")
        self.root.geometry("1600x760")

        self.reader = FacturaTableReader()
        self.catalogo = CatalogoTango()

        self.ruta_factura = None
        self.ruta_modelo_tango = None
        self.ruta_catalogo_tango = None
        self.resultados_debug = []

        self.crear_ui()

    def crear_ui(self):
        top = tk.Frame(self.root)
        top.pack(fill="x", padx=10, pady=10)

        tk.Button(top, text="1. Abrir factura", command=self.abrir_factura, width=18).pack(side="left", padx=3)
        tk.Button(top, text="2. Artículos Tango", command=self.abrir_catalogo, width=18).pack(side="left", padx=3)
        tk.Button(top, text="3. Modelo Tango", command=self.abrir_modelo, width=18).pack(side="left", padx=3)
        tk.Button(top, text="Ver debug", command=self.ver_debug, width=12).pack(side="left", padx=3)

        self.lbl_factura = tk.Label(self.root, text="Factura: no seleccionada", anchor="w")
        self.lbl_factura.pack(fill="x", padx=15)

        self.lbl_catalogo = tk.Label(self.root, text="Artículos Tango: no seleccionado", anchor="w")
        self.lbl_catalogo.pack(fill="x", padx=15)

        self.lbl_modelo = tk.Label(self.root, text="Modelo Tango: no seleccionado", anchor="w")
        self.lbl_modelo.pack(fill="x", padx=15)

        columns = (
            "usar",
            "existe",
            "accion",
            "codigo_tango",
            "descripcion",
            "sinonimo",
            "codigo_barras",
            "cantidad",
            "precio",
            "importe",
            "match",
            "match_score",
            "tipo",
            "escala",
            "pagina",
            "linea",
        )

        self.tabla = ttk.Treeview(self.root, columns=columns, show="headings", height=23)

        headers = {
            "usar": "✓",
            "existe": "Existe",
            "accion": "Acción",
            "codigo_tango": "Código Tango",
            "descripcion": "Descripción",
            "sinonimo": "Cod. proveedor",
            "codigo_barras": "Cód. barras",
            "cantidad": "Cantidad",
            "precio": "Precio",
            "importe": "Importe",
            "match": "Coincidencia",
            "match_score": "Conf.",
            "tipo": "Tipo",
            "escala": "Escala",
            "pagina": "Pág.",
            "linea": "Línea detectada",
        }

        widths = {
            "usar": 45,
            "existe": 65,
            "accion": 140,
            "codigo_tango": 120,
            "descripcion": 310,
            "sinonimo": 130,
            "codigo_barras": 125,
            "cantidad": 80,
            "precio": 85,
            "importe": 90,
            "match": 210,
            "match_score": 60,
            "tipo": 80,
            "escala": 80,
            "pagina": 50,
            "linea": 450,
        }

        for c in columns:
            self.tabla.heading(c, text=headers[c])
            self.tabla.column(c, width=widths[c])

        self.tabla.pack(fill="both", expand=True, padx=10, pady=10)
        self.tabla.bind("<Double-1>", self.doble_click)

        bottom = tk.Frame(self.root)
        bottom.pack(fill="x", padx=10, pady=8)

        tk.Button(bottom, text="Seleccionar todo", command=self.seleccionar_todo, width=18).pack(side="left", padx=4)
        tk.Button(bottom, text="Borrar selección", command=self.borrar_seleccion, width=18).pack(side="left", padx=4)
        tk.Button(bottom, text="Invertir selección", command=self.invertir_seleccion, width=18).pack(side="left", padx=4)
        tk.Button(bottom, text="Agregar manual", command=self.agregar_manual, width=16).pack(side="left", padx=4)
        tk.Button(bottom, text="Eliminar fila", command=self.eliminar_fila, width=14).pack(side="left", padx=4)
        tk.Button(bottom, text="Recomparar", command=self.recomparar, width=14).pack(side="left", padx=4)
        tk.Button(bottom, text="Generar alta artículos", command=self.generar_excel_alta_articulos, width=24).pack(side="right", padx=4)
        tk.Button(bottom, text="Generar resumen compra/stock", command=self.generar_resumen_compra_stock, width=28).pack(side="right", padx=4)

        tk.Label(
            self.root,
            text="☑ = se usa. Doble click en la primera columna activa/desactiva. Los artículos existentes no van al alta, quedan para compra/stock.",
            fg="blue"
        ).pack(pady=4)

    def abrir_factura(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar factura",
            filetypes=[
                ("Facturas", "*.pdf *.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                ("Todos", "*.*")
            ]
        )

        if not ruta:
            return

        self.ruta_factura = ruta
        self.lbl_factura.config(text=f"Factura: {ruta}")
        self.procesar_factura(ruta)

    def abrir_catalogo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Artículos.xlsx exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not ruta:
            return

        try:
            self.catalogo.cargar(ruta)
            self.ruta_catalogo_tango = ruta
            self.lbl_catalogo.config(
                text=f"Artículos Tango: {ruta} ({len(self.catalogo.articulos)} artículos)"
            )
            self.recomparar()
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def abrir_modelo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel modelo exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def procesar_factura(self, ruta):
        try:
            items, debug = self.reader.analizar(ruta)
            self.resultados_debug = debug
            self.cargar_items(items)

            if self.catalogo.articulos:
                self.recomparar()

            messagebox.showinfo(
                "Lectura terminada",
                f"Se detectaron {len(items)} líneas candidatas."
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def cargar_items(self, items):
        for i in self.tabla.get_children():
            self.tabla.delete(i)

        for item in items:
            self.insertar_item(item)

    def insertar_item(self, item):
        self.tabla.insert(
            "",
            "end",
            values=(
                CHECK_ON if item["usar"] else CHECK_OFF,
                item["existe_tango"],
                item["accion"],
                item["codigo_tango"],
                item["descripcion"],
                item["sinonimo"],
                item["codigo_barras"],
                item["cantidad"],
                item["precio"],
                item["importe"],
                item.get("coincidencia", ""),
                item.get("match_score", ""),
                item["tipo"],
                item["escala"],
                item.get("pagina", ""),
                item["linea_detectada"],
            )
        )

    def recomparar(self):
        if not self.catalogo.articulos:
            messagebox.showinfo("Catálogo no cargado", "Primero cargá el Excel Artículos.xlsx exportado desde Tango.")
            return

        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))

            codigo_proveedor = v[5]
            codigo_barras = v[6]
            descripcion = v[4]

            art, metodo, score = self.catalogo.buscar(
                codigo_proveedor=codigo_proveedor,
                codigo_barras=codigo_barras,
                descripcion=descripcion
            )

            if art:
                v[1] = "Sí"
                v[2] = "Cargar stock / compra"
                v[3] = art["codigo"]
                v[4] = art["descripcion"]
                v[6] = art["codigo_barras"]
                v[10] = metodo
                v[11] = str(score)
            else:
                v[1] = "No"
                v[2] = "Alta artículo"
                v[10] = ""
                v[11] = ""

            self.tabla.item(item_id, values=v)

    def doble_click(self, event):
        item_id = self.tabla.focus()

        if not item_id:
            return

        col = self.tabla.identify_column(event.x)
        idx = int(col.replace("#", "")) - 1

        valores = list(self.tabla.item(item_id, "values"))

        if idx == 0:
            valores[0] = CHECK_OFF if valores[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=valores)
            return

        nombres = [
            "Usar",
            "Existe",
            "Acción",
            "Código Tango",
            "Descripción",
            "Cod. proveedor",
            "Cód. barras",
            "Cantidad",
            "Precio",
            "Importe",
            "Coincidencia",
            "Confianza",
            "Tipo",
            "Escala",
            "Página",
            "Línea detectada",
        ]

        actual = valores[idx]
        nuevo = simpledialog.askstring("Editar", nombres[idx], initialvalue=actual)

        if nuevo is not None:
            valores[idx] = nuevo
            self.tabla.item(item_id, values=valores)

    def seleccionar_todo(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_ON
            self.tabla.item(item_id, values=v)

    def borrar_seleccion(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_OFF
            self.tabla.item(item_id, values=v)

    def invertir_seleccion(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_OFF if v[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=v)

    def agregar_manual(self):
        item = {
            "usar": True,
            "existe_tango": "No",
            "accion": "Alta artículo",
            "codigo_tango": "",
            "descripcion": "",
            "desc_adic": "",
            "sinonimo": "",
            "codigo_barras": "",
            "cantidad": "",
            "precio": "",
            "importe": "",
            "tipo": "Simple",
            "escala": "No usa",
            "pagina": "",
            "linea_detectada": "Manual",
            "coincidencia": "",
            "match_score": "",
        }
        self.insertar_item(item)

    def eliminar_fila(self):
        for item_id in self.tabla.selection():
            self.tabla.delete(item_id)

    def ver_debug(self):
        if not self.resultados_debug:
            messagebox.showerror("Error", "Primero abrí una factura.")
            return

        for pagina, debug_path, lineas, zonas in self.resultados_debug:
            os.startfile(debug_path)

    def leer_filas(self):
        filas = []

        for item_id in self.tabla.get_children():
            v = self.tabla.item(item_id, "values")

            filas.append({
                "usar": v[0],
                "existe": v[1],
                "accion": v[2],
                "codigo_tango": v[3].strip(),
                "descripcion": v[4].strip(),
                "sinonimo": v[5].strip(),
                "codigo_barras": v[6].strip(),
                "cantidad": v[7].strip(),
                "precio": v[8].strip(),
                "importe": v[9].strip(),
                "coincidencia": v[10].strip(),
                "match_score": v[11].strip(),
                "tipo": v[12].strip() or "Simple",
                "escala": v[13].strip() or "No usa",
            })

        return filas

    def generar_excel_alta_articulos(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo Tango.")
            return

        filas = self.leer_filas()

        nuevos = [
            f for f in filas
            if f["usar"] == CHECK_ON and f["existe"] == "No"
        ]

        if not nuevos:
            messagebox.showinfo(
                "Sin artículos nuevos",
                "No hay artículos nuevos para dar de alta.\n\nLos existentes quedan para la futura carga de stock/comprobante."
            )
            return

        faltan = [f for f in nuevos if not f["codigo_tango"]]

        if faltan:
            messagebox.showerror(
                "Faltan códigos Tango",
                "Hay artículos nuevos sin Código Tango.\n\nCompletalos antes de generar el Excel."
            )
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar Excel Alta Artículos Tango",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Alta_Articulos_Tango.xlsx"
        )

        if not salida:
            return

        try:
            wb = load_workbook(self.ruta_modelo_tango)

            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja 'Artículos'.")
                return

            ws = wb["Artículos"]

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for fila, item in enumerate(nuevos, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = ""
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]

            wb.save(salida)

            messagebox.showinfo(
                "Listo",
                f"Excel de alta de artículos generado:\n\n{salida}\n\nArtículos nuevos: {len(nuevos)}"
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def generar_resumen_compra_stock(self):
        filas = [
            f for f in self.leer_filas()
            if f["usar"] == CHECK_ON
        ]

        if not filas:
            messagebox.showerror("Error", "No hay filas seleccionadas.")
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar resumen compra/stock",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Resumen_Compra_Stock.xlsx"
        )

        if not salida:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Compra_Stock"

        headers = [
            "Existe en Tango",
            "Acción",
            "Código Tango",
            "Descripción Tango/OCR",
            "Código proveedor",
            "Código barras",
            "Cantidad",
            "Precio",
            "Importe",
            "Coincidencia",
            "Confianza",
        ]

        for col, h in enumerate(headers, start=1):
            ws.cell(1, col).value = h

        for row, f in enumerate(filas, start=2):
            ws.cell(row, 1).value = f["existe"]
            ws.cell(row, 2).value = f["accion"]
            ws.cell(row, 3).value = f["codigo_tango"]
            ws.cell(row, 4).value = f["descripcion"]
            ws.cell(row, 5).value = f["sinonimo"]
            ws.cell(row, 6).value = f["codigo_barras"]
            ws.cell(row, 7).value = f["cantidad"]
            ws.cell(row, 8).value = f["precio"]
            ws.cell(row, 9).value = f["importe"]
            ws.cell(row, 10).value = f["coincidencia"]
            ws.cell(row, 11).value = f["match_score"]

        widths = [16, 22, 18, 50, 18, 18, 12, 12, 12, 30, 12]

        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + idx)].width = width

        wb.save(salida)

        messagebox.showinfo(
            "Listo",
            f"Resumen de compra/stock generado:\n\n{salida}\n\nEste archivo es de revisión; todavía no es la plantilla oficial de comprobante de compra."
        )


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()