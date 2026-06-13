import os
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageOps, ImageEnhance, ImageFilter
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font
import fitz  # PyMuPDF


# Ajustar si Tesseract está en otra ruta
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


def preparar_imagen(imagen):
    imagen = imagen.convert("L")
    imagen = ImageOps.autocontrast(imagen)
    imagen = ImageEnhance.Contrast(imagen).enhance(2.0)
    imagen = ImageEnhance.Sharpness(imagen).enhance(2.0)
    imagen = imagen.filter(ImageFilter.SHARPEN)
    return imagen


def pdf_a_imagenes(ruta_pdf):
    doc = fitz.open(ruta_pdf)
    imagenes = []

    for pagina in doc:
        pix = pagina.get_pixmap(matrix=fitz.Matrix(2, 2))
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        imagenes.append(img)

    return imagenes


def hacer_ocr(ruta):
    extension = os.path.splitext(ruta)[1].lower()

    textos = []

    if extension == ".pdf":
        imagenes = pdf_a_imagenes(ruta)
    else:
        imagenes = [Image.open(ruta)]

    for imagen in imagenes:
        imagen_preparada = preparar_imagen(imagen)

        try:
            texto = pytesseract.image_to_string(imagen_preparada, lang="spa")
        except:
            texto = pytesseract.image_to_string(imagen_preparada)

        textos.append(texto)

    return "\n\n".join(textos)


def limpiar_numero(valor):
    if not valor:
        return ""

    valor = valor.strip()
    valor = valor.replace("$", "")
    valor = valor.replace(" ", "")

    return valor


def buscar_patron(texto, patron, grupo=1):
    m = re.search(patron, texto, re.IGNORECASE | re.MULTILINE)
    if m:
        return m.group(grupo).strip()
    return ""


def extraer_datos(texto):
    datos = {}

    datos["Proveedor"] = detectar_proveedor(texto)

    datos["Tipo comprobante"] = buscar_patron(
        texto,
        r"(FACTURA\s+[ABC]|REMITO)"
    )

    datos["Número comprobante"] = buscar_patron(
        texto,
        r"(?:FACTURA|REMITO)\s*(?:N[°º*]?)?\s*[:\-]?\s*([0-9]{4,5}\s*[-]\s*[0-9]{6,8})"
    )

    datos["Fecha"] = buscar_patron(
        texto,
        r"(?:FECHA)\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})"
    )

    if not datos["Fecha"]:
        datos["Fecha"] = buscar_patron(
            texto,
            r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b"
        )

    datos["CUIT proveedor"] = buscar_patron(
        texto,
        r"CUIT\s*(?:N[°º*]?)?\s*[:\-]?\s*(\d{2}-?\d{8}-?\d)"
    )

    datos["CUIT cliente"] = buscar_patron(
        texto,
        r"C\.?U\.?I\.?T\.?\s*(?:N[°º*]?)?\s*[:\-]?\s*(\d{2}-?\d{8}-?\d)"
    )

    datos["Cliente"] = buscar_patron(
        texto,
        r"(?:SEÑOR\s*/\s*ES|SEÑOR|CLIENTE)\s*[:\-]?\s*(.+)"
    )

    datos["Remito"] = buscar_patron(
        texto,
        r"REMITO\s*(?:N[°º*]?)?\s*[:\-]?\s*([0-9]{4,5}\s*[-]\s*[0-9]{6,8})"
    )

    datos["CAE"] = buscar_patron(
        texto,
        r"CAE\s*(?:N[°º*]?)?\s*[:\-]?\s*(\d{10,20})"
    )

    datos["Vencimiento CAE"] = buscar_patron(
        texto,
        r"(?:Fecha\s*CAE|Vto\.?\s*CAE|Vencimiento\s*CAE)\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})"
    )

    datos["Subtotal"] = buscar_patron(
        texto,
        r"SUBTOTAL\s*\$?\s*([0-9\.,]+)"
    )

    datos["IVA"] = buscar_patron(
        texto,
        r"I\.?V\.?A\.?\s*(?:21\s*%)?\s*\$?\s*([0-9\.,]+)"
    )

    datos["Total"] = buscar_patron(
        texto,
        r"TOTAL\s*\$?\s*([0-9\.,]+)"
    )

    if not datos["Total"]:
        posibles_totales = re.findall(r"\$\s*([0-9]{1,3}(?:[.,][0-9]{3})*(?:[.,][0-9]{2}))", texto)
        if posibles_totales:
            datos["Total"] = posibles_totales[-1]

    return datos


def detectar_proveedor(texto):
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]

    for linea in lineas[:20]:
        l = linea.upper()
        if "S.R.L" in l or "SRL" in l or "S.A" in l or "SA " in l:
            return linea

    for linea in lineas[:10]:
        if len(linea) > 3 and not re.search(r"\d{2}/\d{2}/\d{2,4}", linea):
            return linea

    return ""


def extraer_items(texto):
    items = []
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]

    for linea in lineas:
        original = linea

        if len(linea) < 8:
            continue

        l = linea.upper()

        palabras_excluir = [
            "FACTURA", "CUIT", "FECHA", "CAE", "SUBTOTAL", "TOTAL",
            "IVA RESPONSABLE", "DOMICILIO", "CONDICION", "CONDICIONES",
            "LOCALIDAD", "PROVINCIA", "REMITO", "PAGINA", "ORIGINAL",
            "INGRESOS BRUTOS", "INICIO DE ACTIVIDADES"
        ]

        if any(p in l for p in palabras_excluir):
            continue

        # Caso típico:
        # 4 220187 - Electrodo, O2, 80 Amp 21,00 27,983.06 111,932.24
        patron = re.search(
            r"^\s*(\d+(?:[.,]\d+)?)\s+([A-Z0-9\.\-]+)\s*[-–]\s*(.+?)\s+(\d{1,2}[,.]\d{2})\s+([0-9\.,]+)\s+([0-9\.,]+)\s*$",
            linea,
            re.IGNORECASE
        )

        if patron:
            items.append({
                "Cantidad": patron.group(1),
                "Código": patron.group(2),
                "Descripción": patron.group(3),
                "IVA %": patron.group(4),
                "Precio unitario": limpiar_numero(patron.group(5)),
                "Total": limpiar_numero(patron.group(6)),
                "Línea OCR original": original
            })
            continue

        # Caso más flexible: cantidad + código + descripción
        patron_simple = re.search(
            r"^\s*(\d+(?:[.,]\d+)?)\s+([A-Z0-9\.\-]{3,})\s*[-–]?\s*(.+)$",
            linea,
            re.IGNORECASE
        )

        if patron_simple:
            descripcion = patron_simple.group(3)

            if len(descripcion) > 3:
                items.append({
                    "Cantidad": patron_simple.group(1),
                    "Código": patron_simple.group(2),
                    "Descripción": descripcion,
                    "IVA %": "",
                    "Precio unitario": "",
                    "Total": "",
                    "Línea OCR original": original
                })

    return items


def crear_excel(ruta_archivo, texto, datos, items):
    carpeta = os.path.dirname(ruta_archivo)
    nombre = os.path.splitext(os.path.basename(ruta_archivo))[0]
    salida = os.path.join(carpeta, f"{nombre}_OCR.xlsx")

    wb = Workbook()
    bold = Font(bold=True)

    ws = wb.active
    ws.title = "Datos detectados"

    ws["A1"] = "Campo"
    ws["B1"] = "Valor"
    ws["A1"].font = bold
    ws["B1"].font = bold

    fila = 2
    for campo, valor in datos.items():
        ws.cell(fila, 1, campo)
        ws.cell(fila, 2, valor)
        fila += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    ws_items = wb.create_sheet("Items detectados")

    encabezados = [
        "Cantidad",
        "Código",
        "Código de barras",
        "Descripción",
        "IVA %",
        "Precio unitario",
        "Total",
        "Código Tango",
        "Existe en Tango",
        "Acción sugerida",
        "Línea OCR original"
    ]

    for col, enc in enumerate(encabezados, start=1):
        ws_items.cell(1, col, enc).font = bold

    if items:
        for fila, item in enumerate(items, start=2):
            ws_items.cell(fila, 1, item.get("Cantidad", ""))
            ws_items.cell(fila, 2, item.get("Código", ""))
            ws_items.cell(fila, 3, "")
            ws_items.cell(fila, 4, item.get("Descripción", ""))
            ws_items.cell(fila, 5, item.get("IVA %", ""))
            ws_items.cell(fila, 6, item.get("Precio unitario", ""))
            ws_items.cell(fila, 7, item.get("Total", ""))
            ws_items.cell(fila, 8, "")
            ws_items.cell(fila, 9, "")
            ws_items.cell(fila, 10, "Revisar / Cargar")
            ws_items.cell(fila, 11, item.get("Línea OCR original", ""))
    else:
        ws_items.cell(2, 1, "No se detectaron ítems automáticamente")
        ws_items.cell(2, 11, "Revisar hoja Texto OCR")

    anchos = {
        "A": 12,
        "B": 18,
        "C": 22,
        "D": 55,
        "E": 10,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
        "J": 22,
        "K": 100,
    }

    for col, ancho in anchos.items():
        ws_items.column_dimensions[col].width = ancho

    ws_texto = wb.create_sheet("Texto OCR")
    ws_texto["A1"] = "Texto OCR completo"
    ws_texto["A1"].font = bold
    ws_texto["A2"] = texto
    ws_texto.column_dimensions["A"].width = 160

    ws_revision = wb.create_sheet("Revision manual")
    rev_headers = [
        "Campo",
        "Valor detectado",
        "Valor corregido",
        "Observaciones"
    ]

    for col, enc in enumerate(rev_headers, start=1):
        ws_revision.cell(1, col, enc).font = bold

    fila = 2
    for campo, valor in datos.items():
        ws_revision.cell(fila, 1, campo)
        ws_revision.cell(fila, 2, valor)
        ws_revision.cell(fila, 3, valor)
        fila += 1

    ws_revision.column_dimensions["A"].width = 30
    ws_revision.column_dimensions["B"].width = 60
    ws_revision.column_dimensions["C"].width = 60
    ws_revision.column_dimensions["D"].width = 40

    ws_tango = wb.create_sheet("Base futura Tango")
    tango_headers = [
        "Proveedor",
        "CUIT proveedor",
        "Tipo comprobante",
        "Número comprobante",
        "Fecha",
        "Código proveedor",
        "Código de barras",
        "Descripción",
        "Cantidad",
        "Precio unitario",
        "IVA %",
        "Importe",
        "Código Tango",
        "Acción"
    ]

    for col, enc in enumerate(tango_headers, start=1):
        ws_tango.cell(1, col, enc).font = bold

    for fila, item in enumerate(items, start=2):
        ws_tango.cell(fila, 1, datos.get("Proveedor", ""))
        ws_tango.cell(fila, 2, datos.get("CUIT proveedor", ""))
        ws_tango.cell(fila, 3, datos.get("Tipo comprobante", ""))
        ws_tango.cell(fila, 4, datos.get("Número comprobante", ""))
        ws_tango.cell(fila, 5, datos.get("Fecha", ""))
        ws_tango.cell(fila, 6, item.get("Código", ""))
        ws_tango.cell(fila, 7, "")
        ws_tango.cell(fila, 8, item.get("Descripción", ""))
        ws_tango.cell(fila, 9, item.get("Cantidad", ""))
        ws_tango.cell(fila, 10, item.get("Precio unitario", ""))
        ws_tango.cell(fila, 11, item.get("IVA %", ""))
        ws_tango.cell(fila, 12, item.get("Total", ""))
        ws_tango.cell(fila, 13, "")
        ws_tango.cell(fila, 14, "Revisar")

    for col in range(1, len(tango_headers) + 1):
        ws_tango.column_dimensions[chr(64 + col)].width = 22

    wb.save(salida)
    return salida


def procesar_archivo():
    ruta = filedialog.askopenfilename(
        title="Seleccionar factura/remito",
        filetypes=[
            ("Facturas", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff *.pdf"),
            ("Imágenes", "*.png *.jpg *.jpeg *.bmp *.tif *.tiff"),
            ("PDF", "*.pdf"),
            ("Todos los archivos", "*.*")
        ]
    )

    if not ruta:
        return

    try:
        texto = hacer_ocr(ruta)
        datos = extraer_datos(texto)
        items = extraer_items(texto)
        salida = crear_excel(ruta, texto, datos, items)

        messagebox.showinfo(
            "Proceso terminado",
            f"Excel generado correctamente:\n\n{salida}\n\n"
            f"Ítems detectados: {len(items)}"
        )

    except Exception as e:
        messagebox.showerror("Error", str(e))


def crear_interfaz():
    root = tk.Tk()
    root.title("OCR Facturas a Excel - V2")
    root.geometry("520x260")

    titulo = tk.Label(
        root,
        text="OCR de Facturas / Remitos a Excel",
        font=("Arial", 15, "bold")
    )
    titulo.pack(pady=20)

    boton = tk.Button(
        root,
        text="Seleccionar imagen/PDF y generar Excel",
        command=procesar_archivo,
        font=("Arial", 11),
        width=38,
        height=2
    )
    boton.pack(pady=18)

    nota = tk.Label(
        root,
        text="V2: mejora imagen, lee OCR, detecta datos, ítems y genera base futura Tango.",
        font=("Arial", 9)
    )
    nota.pack(pady=10)

    root.mainloop()


if __name__ == "__main__":
    crear_interfaz()