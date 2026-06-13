import os
import re
import tempfile
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from dataclasses import dataclass
from typing import List
from difflib import SequenceMatcher
from PIL import Image, ImageDraw, ImageOps, ImageEnhance
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
import fitz
import pytesseract

try:
    import cv2
except ImportError:
    cv2 = None


# Ajustar esta ruta si Tesseract está instalado en otro lugar
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

CHECK_ON = "☑"
CHECK_OFF = "☐"


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    page: int


@dataclass
class LineBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    words: List[WordBox]
    score: int = 0


class CatalogoTango:
    def __init__(self):
        self.articulos = []

    def quitar_acentos(self, texto):
        texto = unicodedata.normalize("NFD", texto)
        texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
        return texto

    def normalizar(self, texto):
        if texto is None:
            return ""

        texto = str(texto).upper().strip()
        texto = self.quitar_acentos(texto)

        reemplazos = {
            "P/ACERO": "ACERO",
            "P ACERO": "ACERO",
            "PACERO": "ACERO",
            "P/ ACERO": "ACERO",
            "P/INOX": "INOX",
            "P INOX": "INOX",
            "P/ INOX": "INOX",
            "INOXIDABLE": "INOX",
            "C/": "CON ",
            "S/": "SIN ",
            "MM.": "MM",
            "M.M.": "MM",
            "  ": " ",
        }

        for viejo, nuevo in reemplazos.items():
            texto = texto.replace(viejo, nuevo)

        texto = texto.replace("/", " ")
        texto = texto.replace("-", " ")
        texto = texto.replace(",", " ")
        texto = texto.replace(".", " ")
        texto = texto.replace("(", " ")
        texto = texto.replace(")", " ")
        texto = texto.replace("[", " ")
        texto = texto.replace("]", " ")

        texto = re.sub(r"\s+", " ", texto).strip()
        return texto

    def tokens(self, texto):
        texto = self.normalizar(texto)
        return [t for t in texto.split() if len(t) >= 2]

    def cargar(self, ruta_excel):
        wb = load_workbook(ruta_excel, data_only=True)

        if "Artículos" not in wb.sheetnames:
            raise Exception("El Excel no tiene una hoja llamada 'Artículos'.")

        ws = wb["Artículos"]
        self.articulos = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = row[0] if len(row) > 0 else ""
            descripcion = row[1] if len(row) > 1 else ""
            desc_adic = row[2] if len(row) > 2 else ""
            sinonimo = row[3] if len(row) > 3 else ""
            codigo_barras = row[4] if len(row) > 4 else ""

            if not codigo:
                continue

            art = {
                "codigo": str(codigo).strip(),
                "descripcion": str(descripcion or "").strip(),
                "desc_adic": str(desc_adic or "").strip(),
                "sinonimo": str(sinonimo or "").strip(),
                "codigo_barras": str(codigo_barras or "").strip(),
            }

            art["_codigo_n"] = self.normalizar(art["codigo"])
            art["_descripcion_n"] = self.normalizar(art["descripcion"])
            art["_desc_adic_n"] = self.normalizar(art["desc_adic"])
            art["_sinonimo_n"] = self.normalizar(art["sinonimo"])
            art["_codigo_barras_n"] = self.normalizar(art["codigo_barras"])
            art["_texto_completo_n"] = self.normalizar(
                f"{art['codigo']} {art['descripcion']} {art['desc_adic']} {art['sinonimo']} {art['codigo_barras']}"
            )

            self.articulos.append(art)

    def similitud(self, a, b):
        a = self.normalizar(a)
        b = self.normalizar(b)

        if not a or not b:
            return 0

        return int(SequenceMatcher(None, a, b).ratio() * 100)

    def score_tokens(self, texto_ocr, texto_tango):
        tokens_ocr = set(self.tokens(texto_ocr))
        tokens_tango = set(self.tokens(texto_tango))

        if not tokens_ocr or not tokens_tango:
            return 0

        comunes = tokens_ocr.intersection(tokens_tango)
        return int((len(comunes) / len(tokens_ocr)) * 100)

    def buscar(self, codigo_proveedor="", codigo_barras="", descripcion=""):
        codigo_proveedor_n = self.normalizar(codigo_proveedor)
        codigo_barras_n = self.normalizar(codigo_barras)
        descripcion_n = self.normalizar(descripcion)

        if codigo_barras_n:
            for art in self.articulos:
                if art["_codigo_barras_n"] and art["_codigo_barras_n"] == codigo_barras_n:
                    return art, "Código de barras exacto", 100

        if codigo_proveedor_n:
            for art in self.articulos:
                if art["_codigo_n"] == codigo_proveedor_n:
                    return art, "Código Tango exacto", 100

            for art in self.articulos:
                if art["_sinonimo_n"] and art["_sinonimo_n"] == codigo_proveedor_n:
                    return art, "Sinónimo exacto", 100

            for art in self.articulos:
                if codigo_proveedor_n in art["_descripcion_n"]:
                    return art, "Código proveedor en descripción", 96

            for art in self.articulos:
                if codigo_proveedor_n in art["_texto_completo_n"]:
                    return art, "Código proveedor contenido", 94

        mejor_art = None
        mejor_metodo = ""
        mejor_score = 0

        if descripcion_n:
            for art in self.articulos:
                score_desc = self.similitud(descripcion_n, art["_descripcion_n"])
                score_tokens = self.score_tokens(descripcion_n, art["_descripcion_n"])
                score_final = max(score_desc, score_tokens)

                if score_final > mejor_score:
                    mejor_score = score_final
                    mejor_art = art
                    mejor_metodo = "Similitud descripción"

            if mejor_art and mejor_score >= 82:
                return mejor_art, mejor_metodo, mejor_score

        return None, "", 0


class FacturaTableReader:
    def __init__(self):
        self.stop_words = [
            "total", "subtotal", "cae", "iva 21", "iva 10", "iva 27",
            "son pesos", "observaciones", "importe neto", "fecha de vto"
        ]

    def is_cuit(self, text):
        clean = text.strip()

        if re.fullmatch(r"\d{2}-\d{8}-\d", clean):
            return True

        only_digits = re.sub(r"\D", "", clean)

        if len(only_digits) == 11 and only_digits[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
            return True

        return False

    def normalizar_cuit(self, cuit):
        if not cuit:
            return ""
        return re.sub(r"\D", "", str(cuit))

    def abrir_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()

        if ext == ".pdf":
            return self.leer_pdf(ruta)

        return self.leer_imagen(ruta)

    def leer_pdf(self, ruta):
        doc = fitz.open(ruta)
        paginas = []

        for page_index, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(tempfile.gettempdir(), f"pagina_factura_{page_index}.png")
            pix.save(img_path)
            img = Image.open(img_path)

            palabras = []
            escala = 2

            for w in page.get_text("words"):
                x0, y0, x1, y1, text = w[:5]
                text = str(text).strip()

                if text:
                    palabras.append(
                        WordBox(
                            x0 * escala,
                            y0 * escala,
                            x1 * escala,
                            y1 * escala,
                            text,
                            page_index
                        )
                    )

            paginas.append((img, palabras))

        return paginas

    def leer_imagen(self, ruta):
        img = Image.open(ruta)
        img_ocr = self.preparar_imagen(img)

        data = pytesseract.image_to_data(
            img_ocr,
            lang="spa",
            output_type=pytesseract.Output.DICT,
            config="--psm 6"
        )

        palabras = []

        for i, text in enumerate(data["text"]):
            text = str(text).strip()

            if not text:
                continue

            try:
                conf = int(float(data["conf"][i]))
            except Exception:
                conf = -1

            if conf < 25:
                continue

            x = data["left"][i]
            y = data["top"][i]
            w = data["width"][i]
            h = data["height"][i]

            palabras.append(WordBox(x, y, x + w, y + h, text, 1))

        return [(img, palabras)]

    def preparar_imagen(self, img):
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.2)
        return img

    def agrupar_lineas(self, palabras):
        palabras = sorted(palabras, key=lambda w: (w.y0, w.x0))
        lineas = []
        actual = []

        for palabra in palabras:
            if not actual:
                actual = [palabra]
                continue

            y_prom = sum(w.y0 for w in actual) / len(actual)

            if abs(palabra.y0 - y_prom) <= 8:
                actual.append(palabra)
            else:
                lineas.append(self.crear_linea(actual))
                actual = [palabra]

        if actual:
            lineas.append(self.crear_linea(actual))

        return lineas

    def crear_linea(self, palabras):
        palabras = sorted(palabras, key=lambda w: w.x0)

        return LineBox(
            x0=min(w.x0 for w in palabras),
            y0=min(w.y0 for w in palabras),
            x1=max(w.x1 for w in palabras),
            y1=max(w.y1 for w in palabras),
            text=" ".join(w.text for w in palabras),
            words=palabras
        )

    def es_numero(self, t):
        t = t.replace("$", "").strip()

        if self.is_cuit(t):
            return False

        return (
            bool(re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2,4}$", t))
            or bool(re.match(r"^-?\d+([,.]\d+)?$", t))
        )

    def cantidad_columnas(self, linea):
        if len(linea.words) < 3:
            return 0

        gaps = 0

        for a, b in zip(linea.words, linea.words[1:]):
            if b.x0 - a.x1 > 25:
                gaps += 1

        return gaps + 1

    def puntuar_linea(self, linea):
        texto = linea.text.lower()

        if any(self.is_cuit(w.text) for w in linea.words):
            linea.score = -10
            return linea.score

        if any(p in texto for p in self.stop_words):
            linea.score = -5
            return linea.score

        numeros = sum(1 for w in linea.words if self.es_numero(w.text))
        columnas = self.cantidad_columnas(linea)

        score = 0

        if columnas >= 3:
            score += 2
        if columnas >= 5:
            score += 2
        if numeros >= 2:
            score += 2
        if numeros >= 4:
            score += 2
        if len(linea.words) >= 5:
            score += 1

        linea.score = score
        return score

    def detectar_zonas(self, lineas):
        for linea in lineas:
            self.puntuar_linea(linea)

        candidatas = [l for l in lineas if l.score >= 4]

        if not candidatas:
            return []

        grupos = []
        grupo = [candidatas[0]]

        for linea in candidatas[1:]:
            if linea.y0 - grupo[-1].y1 < 75:
                grupo.append(linea)
            else:
                grupos.append(grupo)
                grupo = [linea]

        grupos.append(grupo)

        zonas = []

        for grupo in grupos:
            if len(grupo) < 2:
                continue

            zonas.append((
                min(l.x0 for l in grupo),
                min(l.y0 for l in grupo),
                max(l.x1 for l in grupo),
                max(l.y1 for l in grupo),
                grupo
            ))

        return zonas

    def detectar_items(self, lineas):
        zonas = self.detectar_zonas(lineas)
        items = []

        for x0, y0, x1, y1, grupo in zonas:
            for linea in grupo:
                if linea.score < 4:
                    continue

                if any(self.is_cuit(w.text) for w in linea.words):
                    continue

                item = self.linea_a_item(linea)

                if item:
                    items.append(item)

        return items, zonas

    def linea_a_item(self, linea):
        words = linea.words

        codigo_proveedor = ""
        descripcion_parts = []
        cantidad = ""
        precio = ""
        importe = ""
        numeros_detectados = []

        for w in words:
            t = w.text.strip()

            if self.is_cuit(t):
                return None

            if self.es_numero(t):
                numeros_detectados.append(t)
                continue

            if not codigo_proveedor and re.match(r"^[A-Z0-9][A-Z0-9.\-/]{2,}$", t, re.IGNORECASE):
                codigo_proveedor = t[:15]
                continue

            if t.upper() in ["U", "UN", "KG", "M", "L", "%"]:
                continue

            if re.match(r"^-?\d+([,.]\d+)?%?$", t):
                continue

            descripcion_parts.append(t)

        if numeros_detectados:
            cantidad = numeros_detectados[0]

        if len(numeros_detectados) >= 2:
            precio = numeros_detectados[-2]

        if len(numeros_detectados) >= 1:
            importe = numeros_detectados[-1]

        descripcion = " ".join(descripcion_parts)
        descripcion = re.sub(r"\s+", " ", descripcion).strip()
        descripcion = descripcion.replace(" /", "").replace("/", "")

        if not descripcion and not codigo_proveedor:
            return None

        return {
            "usar": True,
            "existe_tango": "No",
            "accion": "Alta artículo",
            "codigo_tango": "",
            "descripcion": descripcion[:50],
            "desc_adic": "",
            "sinonimo": codigo_proveedor[:15],
            "codigo_barras": "",
            "cantidad": cantidad,
            "precio": precio,
            "importe": importe,
            "tipo": "Simple",
            "escala": "No usa",
            "linea_detectada": linea.text,
            "score": linea.score,
            "coincidencia": "",
            "match_score": 0,
        }

    def extraer_texto_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()
        texto = ""

        if ext == ".pdf":
            try:
                doc = fitz.open(ruta)
                for page in doc:
                    texto += page.get_text("text") + "\n"
                return texto
            except Exception:
                return ""

        try:
            img = Image.open(ruta)
            img = self.preparar_imagen(img)
            try:
                return pytesseract.image_to_string(img, lang="spa")
            except Exception:
                return pytesseract.image_to_string(img)
        except Exception:
            return ""

    def detectar_proveedor(self, texto):
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]

        for l in lineas[:20]:
            may = l.upper()
            if "S.R.L" in may or "SRL" in may or "S.A" in may or " SA" in may:
                return l

        for l in lineas[:10]:
            if len(l) > 4 and not re.search(r"\d{2}/\d{2}/\d{2,4}", l):
                return l

        return ""

    def detectar_cuit_proveedor(self, texto):
        patrones = [
            r"CUIT\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
            r"C\.?U\.?I\.?T\.?\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
        ]

        for patron in patrones:
            matches = re.findall(patron, texto, re.IGNORECASE)
            for m in matches:
                cuit = self.normalizar_cuit(m)
                if len(cuit) == 11 and cuit[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
                    return cuit

        m2 = re.search(r"\b((?:20|23|24|27|30|33|34)\d{9})\b", texto)
        if m2:
            return m2.group(1)

        return ""

    def detectar_tipo_comprobante(self, texto):
        may = texto.upper()

        if "FACTURA" in may:
            if re.search(r"FACTURA\s*A", may) or re.search(r"\bCOD\.?\s*0?1\b", may):
                return "FACTURA A"
            if re.search(r"FACTURA\s*B", may) or re.search(r"\bCOD\.?\s*0?6\b", may):
                return "FACTURA B"
            if re.search(r"FACTURA\s*C", may) or re.search(r"\bCOD\.?\s*0?11\b", may):
                return "FACTURA C"
            return "FACTURA"

        if "REMITO" in may:
            return "REMITO"

        return ""

    def detectar_numero_comprobante(self, texto):
        patrones = [
            r"Punto\s+de\s+Venta\s*[:\-]?\s*(\d{4,5}).{0,80}?Comp\.?\s*Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"Pto\.?\s*Vta\.?\s*[:\-]?\s*(\d{4,5}).{0,80}?Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"(?:FACTURA|REMITO)\s*(?:N[°º]?)?\s*[:\-]?\s*(\d{4,5})[-\s]?(\d{6,8})",
            r"\b(\d{4,5})[-\s](\d{6,8})\b",
        ]

        for patron in patrones:
            m = re.search(patron, texto, re.IGNORECASE | re.DOTALL)
            if m:
                punto = m.group(1).zfill(4)
                numero = m.group(2).zfill(8)
                return punto, numero, f"{punto}-{numero}"

        return "", "", ""

    def detectar_fecha(self, texto):
        m = re.search(
            r"(?:Fecha\s+de\s+Emisi[oó]n|Fecha)\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})",
            texto,
            re.IGNORECASE
        )

        if m:
            return m.group(1)

        m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", texto)
        if m:
            return m.group(1)

        return ""

    def clave_factura(self, datos):
        cuit = self.normalizar_cuit(datos.get("cuit_proveedor", ""))
        proveedor = str(datos.get("proveedor", "")).strip().upper()
        tipo = str(datos.get("tipo", "")).strip().upper()
        punto = str(datos.get("punto_venta", "")).strip().zfill(4) if datos.get("punto_venta") else ""
        numero = str(datos.get("numero", "")).strip().zfill(8) if datos.get("numero") else ""

        if cuit and punto and numero:
            return f"CUIT:{cuit}|TIPO:{tipo}|PV:{punto}|N:{numero}"

        if proveedor and punto and numero:
            proveedor_n = re.sub(r"\s+", " ", proveedor)
            return f"PROV:{proveedor_n}|TIPO:{tipo}|PV:{punto}|N:{numero}"

        archivo = datos.get("archivo", "")
        return f"ARCHIVO:{archivo}"

    def detectar_datos_factura(self, ruta):
        texto = self.extraer_texto_documento(ruta)

        archivo = os.path.basename(ruta)
        proveedor = self.detectar_proveedor(texto)
        cuit = self.detectar_cuit_proveedor(texto)
        tipo = self.detectar_tipo_comprobante(texto)
        punto, numero, numero_completo = self.detectar_numero_comprobante(texto)
        fecha = self.detectar_fecha(texto)

        etiqueta = archivo

        partes = []
        if proveedor:
            partes.append(proveedor[:30])
        if numero_completo:
            partes.append(numero_completo)
        if fecha:
            partes.append(fecha)

        if partes:
            etiqueta = " | ".join(partes)

        datos = {
            "archivo": archivo,
            "ruta": ruta,
            "proveedor": proveedor,
            "cuit_proveedor": cuit,
            "tipo": tipo,
            "punto_venta": punto,
            "numero": numero,
            "numero_completo": numero_completo,
            "fecha": fecha,
            "etiqueta": etiqueta,
        }

        datos["clave"] = self.clave_factura(datos)
        return datos

    def analizar(self, ruta):
        paginas = self.abrir_documento(ruta)
        resultados = []
        todos_items = []

        datos_factura = self.detectar_datos_factura(ruta)

        for page_num, (img, palabras) in enumerate(paginas, start=1):
            lineas = self.agrupar_lineas(palabras)
            items, zonas = self.detectar_items(lineas)
            debug_path = self.dibujar_debug(img, lineas, zonas, page_num, datos_factura["archivo"])

            for item in items:
                item["pagina"] = page_num
                item["factura_archivo"] = datos_factura["archivo"]
                item["factura_etiqueta"] = datos_factura["etiqueta"]
                item["proveedor_factura"] = datos_factura["proveedor"]
                item["cuit_proveedor"] = datos_factura["cuit_proveedor"]
                item["tipo_factura"] = datos_factura["tipo"]
                item["punto_venta"] = datos_factura["punto_venta"]
                item["nro_factura"] = datos_factura["numero"]
                item["numero_completo"] = datos_factura["numero_completo"]
                item["fecha_factura"] = datos_factura["fecha"]
                item["clave_factura"] = datos_factura["clave"]
                todos_items.append(item)

            resultados.append((datos_factura["archivo"], page_num, debug_path, lineas, zonas))

        return todos_items, resultados, datos_factura

    def dibujar_debug(self, img, lineas, zonas, page_num, nombre_archivo):
        debug = img.convert("RGB")
        draw = ImageDraw.Draw(debug)

        for linea in lineas:
            if linea.score >= 4:
                draw.rectangle([linea.x0, linea.y0, linea.x1, linea.y1], outline="blue", width=2)
                draw.text((linea.x0, max(0, linea.y0 - 14)), f"S{linea.score}", fill="blue")

        for x0, y0, x1, y1, grupo in zonas:
            draw.rectangle([x0, y0, x1, y1], outline="red", width=4)

        base = re.sub(r"[^A-Za-z0-9_-]+", "_", nombre_archivo)
        out = os.path.join(tempfile.gettempdir(), f"debug_{base}_pagina_{page_num}.png")
        debug.save(out)
        return out


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Facturas múltiples → Tango V7")
        self.root.geometry("1720x820")

        self.reader = FacturaTableReader()
        self.catalogo = CatalogoTango()

        self.ruta_modelo_tango = None
        self.ruta_catalogo_tango = None
        self.resultados_debug = []
        self.facturas_cargadas = []
        self.claves_facturas = set()

        self.crear_ui()

    def crear_ui(self):
        top = tk.Frame(self.root)
        top.pack(fill="x", padx=10, pady=10)

        tk.Button(top, text="1. Artículos Tango", command=self.abrir_catalogo, width=18).pack(side="left", padx=3)
        tk.Button(top, text="2. Modelo Tango", command=self.abrir_modelo, width=18).pack(side="left", padx=3)
        tk.Button(top, text="3. Agregar factura(s)", command=self.agregar_facturas, width=22).pack(side="left", padx=3)
        tk.Button(top, text="Sacar foto", command=self.sacar_foto, width=12).pack(side="left", padx=3)
        tk.Button(top, text="Ver debug", command=self.ver_debug, width=12).pack(side="left", padx=3)

        self.lbl_catalogo = tk.Label(self.root, text="Artículos Tango: no seleccionado", anchor="w")
        self.lbl_catalogo.pack(fill="x", padx=15)

        self.lbl_modelo = tk.Label(self.root, text="Modelo Tango: no seleccionado", anchor="w")
        self.lbl_modelo.pack(fill="x", padx=15)

        self.lbl_facturas = tk.Label(self.root, text="Facturas cargadas: 0", anchor="w")
        self.lbl_facturas.pack(fill="x", padx=15)

        columns = (
            "usar",
            "factura",
            "proveedor",
            "cuit",
            "tipo",
            "numero",
            "fecha",
            "existe",
            "accion",
            "codigo_tango",
            "descripcion",
            "sinonimo",
            "codigo_barras",
            "cantidad",
            "precio",
            "importe",
            "match",
            "match_score",
            "tipo_art",
            "escala",
            "pagina",
            "linea",
        )

        headers = {
            "usar": "✓",
            "factura": "Factura",
            "proveedor": "Proveedor",
            "cuit": "CUIT",
            "tipo": "Tipo",
            "numero": "N°",
            "fecha": "Fecha",
            "existe": "Existe",
            "accion": "Acción",
            "codigo_tango": "Código Tango",
            "descripcion": "Descripción",
            "sinonimo": "Cod. proveedor",
            "codigo_barras": "Cód. barras",
            "cantidad": "Cantidad",
            "precio": "Precio",
            "importe": "Importe",
            "match": "Coincidencia",
            "match_score": "Conf.",
            "tipo_art": "Tipo art.",
            "escala": "Escala",
            "pagina": "Pág.",
            "linea": "Línea detectada",
        }

        widths = {
            "usar": 45,
            "factura": 150,
            "proveedor": 190,
            "cuit": 100,
            "tipo": 90,
            "numero": 105,
            "fecha": 85,
            "existe": 65,
            "accion": 140,
            "codigo_tango": 120,
            "descripcion": 300,
            "sinonimo": 125,
            "codigo_barras": 120,
            "cantidad": 80,
            "precio": 85,
            "importe": 90,
            "match": 200,
            "match_score": 60,
            "tipo_art": 80,
            "escala": 80,
            "pagina": 50,
            "linea": 400,
        }

        # Frame contenedor para permitir barras horizontal y vertical
        frame_tabla = tk.Frame(self.root)
        frame_tabla.pack(fill="both", expand=True, padx=10, pady=10)

        scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")
        scroll_y.pack(side="right", fill="y")

        scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        self.tabla = ttk.Treeview(
            frame_tabla,
            columns=columns,
            show="headings",
            height=25,
            yscrollcommand=scroll_y.set,
            xscrollcommand=scroll_x.set
        )

        scroll_y.config(command=self.tabla.yview)
        scroll_x.config(command=self.tabla.xview)

        for c in columns:
            self.tabla.heading(c, text=headers[c])
            self.tabla.column(c, width=widths[c], stretch=False)

        self.tabla.pack(side="left", fill="both", expand=True)
        self.tabla.bind("<Double-1>", self.doble_click)

        bottom = tk.Frame(self.root)
        bottom.pack(fill="x", padx=10, pady=8)

        tk.Button(bottom, text="Seleccionar todo", command=self.seleccionar_todo, width=18).pack(side="left", padx=4)
        tk.Button(bottom, text="Borrar selección", command=self.borrar_seleccion, width=18).pack(side="left", padx=4)
        tk.Button(bottom, text="Invertir selección", command=self.invertir_seleccion, width=18).pack(side="left", padx=4)
        tk.Button(bottom, text="Agregar manual", command=self.agregar_manual, width=16).pack(side="left", padx=4)
        tk.Button(bottom, text="Eliminar fila", command=self.eliminar_fila, width=14).pack(side="left", padx=4)
        tk.Button(bottom, text="Limpiar todo", command=self.limpiar_todo, width=14).pack(side="left", padx=4)
        tk.Button(bottom, text="Recomparar", command=self.recomparar, width=14).pack(side="left", padx=4)

        tk.Button(bottom, text="Generar alta artículos", command=self.generar_excel_alta_articulos, width=24).pack(side="right", padx=4)
        tk.Button(bottom, text="Generar resumen compra/stock", command=self.generar_resumen_compra_stock, width=30).pack(side="right", padx=4)

        tk.Label(
            self.root,
            text="Barra horizontal abajo para ver todas las columnas. Control duplicados: CUIT + tipo + punto de venta + número.",
            fg="blue"
        ).pack(pady=4)

    def abrir_catalogo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Artículos.xlsx exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if not ruta:
            return

        try:
            self.catalogo.cargar(ruta)
            self.ruta_catalogo_tango = ruta
            self.lbl_catalogo.config(
                text=f"Artículos Tango: {ruta} ({len(self.catalogo.articulos)} artículos)"
            )
            self.recomparar(silencioso=True)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def abrir_modelo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel modelo exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def factura_duplicada(self, datos_factura):
        clave = datos_factura.get("clave", "")
        return clave in self.claves_facturas

    def confirmar_factura_duplicada(self, datos):
        mensaje = (
            "Esta factura ya fue cargada en esta sesión.\n\n"
            f"Proveedor: {datos.get('proveedor', '')}\n"
            f"CUIT: {datos.get('cuit_proveedor', '')}\n"
            f"Tipo: {datos.get('tipo', '')}\n"
            f"Número: {datos.get('numero_completo', '')}\n"
            f"Fecha: {datos.get('fecha', '')}\n"
            f"Archivo: {datos.get('archivo', '')}\n\n"
            "¿Querés cargarla igualmente?"
        )

        return messagebox.askyesno("Factura duplicada", mensaje)

    def agregar_facturas(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar una o varias facturas",
            filetypes=[
                ("Facturas", "*.pdf *.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                ("Todos", "*.*")
            ]
        )

        if not rutas:
            return

        total_items = 0
        cargadas = 0
        salteadas = 0

        for ruta in rutas:
            try:
                datos_previos = self.reader.detectar_datos_factura(ruta)

                if self.factura_duplicada(datos_previos):
                    if not self.confirmar_factura_duplicada(datos_previos):
                        salteadas += 1
                        continue

                items, debug, datos_factura = self.reader.analizar(ruta)

                self.resultados_debug.extend(debug)
                self.facturas_cargadas.append(datos_factura)
                self.claves_facturas.add(datos_factura["clave"])

                for item in items:
                    self.insertar_item(item)

                total_items += len(items)
                cargadas += 1

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo procesar:\n{ruta}\n\n{e}")

        if self.catalogo.articulos:
            self.recomparar(silencioso=True)

        self.actualizar_label_facturas()

        messagebox.showinfo(
            "Lectura terminada",
            f"Facturas agregadas: {cargadas}\n"
            f"Facturas salteadas por duplicadas: {salteadas}\n"
            f"Líneas candidatas agregadas: {total_items}"
        )

    def sacar_foto(self):
        if cv2 is None:
            messagebox.showerror("Error", "Falta instalar opencv-python")
            return

        cap = cv2.VideoCapture(0)

        if not cap.isOpened():
            messagebox.showerror("Error", "No se pudo abrir la cámara.")
            return

        messagebox.showinfo("Cámara", "ESPACIO captura / ESC cancela")
        ruta_foto = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            cv2.imshow("Capturar factura", frame)
            key = cv2.waitKey(1)

            if key == 27:
                break

            if key == 32:
                ruta_foto = os.path.join(tempfile.gettempdir(), "factura_capturada.jpg")
                cv2.imwrite(ruta_foto, frame)
                break

        cap.release()
        cv2.destroyAllWindows()

        if ruta_foto:
            try:
                datos_previos = self.reader.detectar_datos_factura(ruta_foto)

                if self.factura_duplicada(datos_previos):
                    if not self.confirmar_factura_duplicada(datos_previos):
                        return

                items, debug, datos_factura = self.reader.analizar(ruta_foto)

                self.resultados_debug.extend(debug)
                self.facturas_cargadas.append(datos_factura)
                self.claves_facturas.add(datos_factura["clave"])

                for item in items:
                    self.insertar_item(item)

                if self.catalogo.articulos:
                    self.recomparar(silencioso=True)

                self.actualizar_label_facturas()

                messagebox.showinfo(
                    "Foto procesada",
                    f"Líneas candidatas agregadas: {len(items)}"
                )

            except Exception as e:
                messagebox.showerror("Error", str(e))

    def actualizar_label_facturas(self):
        nombres = []

        for f in self.facturas_cargadas:
            etiqueta = f.get("numero_completo") or f.get("archivo", "")
            proveedor = f.get("proveedor", "")
            if proveedor:
                etiqueta = f"{proveedor[:25]} {etiqueta}"
            nombres.append(etiqueta)

        if len(nombres) <= 4:
            detalle = " | ".join(nombres)
        else:
            detalle = " | ".join(nombres[:4]) + f" | ... (+{len(nombres) - 4})"

        self.lbl_facturas.config(text=f"Facturas cargadas: {len(self.facturas_cargadas)}   {detalle}")

    def insertar_item(self, item):
        self.tabla.insert(
            "",
            "end",
            values=(
                CHECK_ON if item["usar"] else CHECK_OFF,
                item.get("factura_etiqueta", ""),
                item.get("proveedor_factura", ""),
                item.get("cuit_proveedor", ""),
                item.get("tipo_factura", ""),
                item.get("numero_completo", ""),
                item.get("fecha_factura", ""),
                item["existe_tango"],
                item["accion"],
                item["codigo_tango"],
                item["descripcion"],
                item["sinonimo"],
                item["codigo_barras"],
                item["cantidad"],
                item["precio"],
                item["importe"],
                item.get("coincidencia", ""),
                item.get("match_score", ""),
                item["tipo"],
                item["escala"],
                item.get("pagina", ""),
                item["linea_detectada"],
            )
        )

    def recomparar(self, silencioso=False):
        if not self.catalogo.articulos:
            if not silencioso:
                messagebox.showinfo("Catálogo no cargado", "Primero cargá el Excel Artículos.xlsx exportado desde Tango.")
            return

        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))

            codigo_proveedor = v[11]
            codigo_barras = v[12]
            descripcion = v[10]

            art, metodo, score = self.catalogo.buscar(
                codigo_proveedor=codigo_proveedor,
                codigo_barras=codigo_barras,
                descripcion=descripcion
            )

            if art:
                v[7] = "Sí"
                v[8] = "Cargar stock / compra"
                v[9] = art["codigo"]
                v[10] = art["descripcion"]
                v[12] = art["codigo_barras"]
                v[16] = metodo
                v[17] = str(score)
            else:
                v[7] = "No"
                v[8] = "Alta artículo"
                v[16] = ""
                v[17] = ""

            self.tabla.item(item_id, values=v)

        if not silencioso:
            messagebox.showinfo("Listo", "Comparación terminada.")

    def doble_click(self, event):
        item_id = self.tabla.focus()

        if not item_id:
            return

        col = self.tabla.identify_column(event.x)
        idx = int(col.replace("#", "")) - 1

        valores = list(self.tabla.item(item_id, "values"))

        if idx == 0:
            valores[0] = CHECK_OFF if valores[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=valores)
            return

        nombres = [
            "Usar",
            "Factura",
            "Proveedor",
            "CUIT",
            "Tipo",
            "Número",
            "Fecha",
            "Existe",
            "Acción",
            "Código Tango",
            "Descripción",
            "Cod. proveedor",
            "Cód. barras",
            "Cantidad",
            "Precio",
            "Importe",
            "Coincidencia",
            "Confianza",
            "Tipo art.",
            "Escala",
            "Página",
            "Línea detectada",
        ]

        actual = valores[idx]
        nuevo = simpledialog.askstring("Editar", nombres[idx], initialvalue=actual)

        if nuevo is not None:
            valores[idx] = nuevo
            self.tabla.item(item_id, values=valores)

    def seleccionar_todo(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_ON
            self.tabla.item(item_id, values=v)

    def borrar_seleccion(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_OFF
            self.tabla.item(item_id, values=v)

    def invertir_seleccion(self):
        for item_id in self.tabla.get_children():
            v = list(self.tabla.item(item_id, "values"))
            v[0] = CHECK_OFF if v[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=v)

    def agregar_manual(self):
        item = {
            "usar": True,
            "factura_etiqueta": "Manual",
            "proveedor_factura": "",
            "cuit_proveedor": "",
            "tipo_factura": "",
            "numero_completo": "",
            "fecha_factura": "",
            "existe_tango": "No",
            "accion": "Alta artículo",
            "codigo_tango": "",
            "descripcion": "",
            "desc_adic": "",
            "sinonimo": "",
            "codigo_barras": "",
            "cantidad": "",
            "precio": "",
            "importe": "",
            "tipo": "Simple",
            "escala": "No usa",
            "pagina": "",
            "linea_detectada": "Manual",
            "coincidencia": "",
            "match_score": "",
        }
        self.insertar_item(item)

    def eliminar_fila(self):
        for item_id in self.tabla.selection():
            self.tabla.delete(item_id)

    def limpiar_todo(self):
        if not messagebox.askyesno("Confirmar", "¿Querés borrar todas las facturas y filas cargadas?"):
            return

        for item_id in self.tabla.get_children():
            self.tabla.delete(item_id)

        self.facturas_cargadas = []
        self.resultados_debug = []
        self.claves_facturas = set()
        self.actualizar_label_facturas()

    def ver_debug(self):
        if not self.resultados_debug:
            messagebox.showerror("Error", "Primero agregá una factura.")
            return

        for archivo, pagina, debug_path, lineas, zonas in self.resultados_debug:
            try:
                os.startfile(debug_path)
            except Exception as e:
                messagebox.showerror("Error", str(e))
                return

    def leer_filas(self):
        filas = []

        for item_id in self.tabla.get_children():
            v = self.tabla.item(item_id, "values")

            filas.append({
                "usar": v[0],
                "factura": v[1],
                "proveedor": v[2],
                "cuit": v[3],
                "tipo_comprobante": v[4],
                "numero_comprobante": v[5],
                "fecha": v[6],
                "existe": v[7],
                "accion": v[8],
                "codigo_tango": v[9].strip(),
                "descripcion": v[10].strip(),
                "sinonimo": v[11].strip(),
                "codigo_barras": v[12].strip(),
                "cantidad": v[13].strip(),
                "precio": v[14].strip(),
                "importe": v[15].strip(),
                "coincidencia": v[16].strip(),
                "match_score": v[17].strip(),
                "tipo": v[18].strip() or "Simple",
                "escala": v[19].strip() or "No usa",
                "pagina": v[20],
                "linea": v[21],
            })

        return filas

    def deduplicar_nuevos(self, nuevos):
        normalizador = CatalogoTango()
        vistos = {}
        resultado = []
        duplicados = []

        for f in nuevos:
            clave = (
                normalizador.normalizar(f["sinonimo"]),
                normalizador.normalizar(f["codigo_barras"]),
                normalizador.normalizar(f["descripcion"]),
            )

            if clave in vistos:
                duplicados.append((f, vistos[clave]))
                continue

            vistos[clave] = f
            resultado.append(f)

        return resultado, duplicados

    def generar_excel_alta_articulos(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo Tango.")
            return

        filas = self.leer_filas()

        nuevos = [
            f for f in filas
            if f["usar"] == CHECK_ON and f["existe"] == "No"
        ]

        if not nuevos:
            messagebox.showinfo(
                "Sin artículos nuevos",
                "No hay artículos nuevos para dar de alta.\n\nLos existentes quedan para la carga de compra/stock."
            )
            return

        faltan = [f for f in nuevos if not f["codigo_tango"]]

        if faltan:
            messagebox.showerror(
                "Faltan códigos Tango",
                "Hay artículos nuevos sin Código Tango.\n\nCompletalos antes de generar el Excel."
            )
            return

        nuevos_unicos, duplicados = self.deduplicar_nuevos(nuevos)

        if duplicados:
            messagebox.showinfo(
                "Duplicados agrupados",
                f"Se encontraron {len(duplicados)} artículos nuevos repetidos.\n"
                f"Se exportará solo una vez cada artículo nuevo."
            )

        salida = filedialog.asksaveasfilename(
            title="Guardar Excel Alta Artículos Tango",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Alta_Articulos_Tango.xlsx"
        )

        if not salida:
            return

        try:
            wb = load_workbook(self.ruta_modelo_tango)

            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja 'Artículos'.")
                return

            ws = wb["Artículos"]

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for fila, item in enumerate(nuevos_unicos, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = ""
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]

            wb.save(salida)

            messagebox.showinfo(
                "Listo",
                f"Excel de alta de artículos generado:\n\n{salida}\n\n"
                f"Artículos nuevos únicos: {len(nuevos_unicos)}"
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def generar_resumen_compra_stock(self):
        filas = [
            f for f in self.leer_filas()
            if f["usar"] == CHECK_ON
        ]

        if not filas:
            messagebox.showerror("Error", "No hay filas seleccionadas.")
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar resumen compra/stock",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Resumen_Compra_Stock.xlsx"
        )

        if not salida:
            return

        wb = Workbook()
        bold = Font(bold=True)

        ws = wb.active
        ws.title = "Compra_Stock"

        headers = [
            "Factura",
            "Proveedor",
            "CUIT",
            "Tipo comprobante",
            "Número comprobante",
            "Fecha",
            "Existe en Tango",
            "Acción",
            "Código Tango",
            "Descripción Tango/OCR",
            "Código proveedor",
            "Código barras",
            "Cantidad",
            "Precio",
            "Importe",
            "Coincidencia",
            "Confianza",
            "Línea detectada",
        ]

        for col, h in enumerate(headers, start=1):
            ws.cell(1, col).value = h
            ws.cell(1, col).font = bold

        for row, f in enumerate(filas, start=2):
            ws.cell(row, 1).value = f["factura"]
            ws.cell(row, 2).value = f["proveedor"]
            ws.cell(row, 3).value = f["cuit"]
            ws.cell(row, 4).value = f["tipo_comprobante"]
            ws.cell(row, 5).value = f["numero_comprobante"]
            ws.cell(row, 6).value = f["fecha"]
            ws.cell(row, 7).value = f["existe"]
            ws.cell(row, 8).value = f["accion"]
            ws.cell(row, 9).value = f["codigo_tango"]
            ws.cell(row, 10).value = f["descripcion"]
            ws.cell(row, 11).value = f["sinonimo"]
            ws.cell(row, 12).value = f["codigo_barras"]
            ws.cell(row, 13).value = f["cantidad"]
            ws.cell(row, 14).value = f["precio"]
            ws.cell(row, 15).value = f["importe"]
            ws.cell(row, 16).value = f["coincidencia"]
            ws.cell(row, 17).value = f["match_score"]
            ws.cell(row, 18).value = f["linea"]

        widths = [25, 35, 15, 18, 18, 12, 16, 22, 18, 50, 18, 18, 12, 12, 12, 30, 12, 80]

        for idx, width in enumerate(widths, start=1):
            col_letter = ws.cell(1, idx).column_letter
            ws.column_dimensions[col_letter].width = width

        ws2 = wb.create_sheet("Facturas_Cargadas")

        fact_headers = [
            "Clave",
            "Archivo",
            "Proveedor detectado",
            "CUIT proveedor",
            "Tipo",
            "Punto de venta",
            "Número",
            "Número completo",
            "Fecha",
            "Ruta",
        ]

        for col, h in enumerate(fact_headers, start=1):
            ws2.cell(1, col).value = h
            ws2.cell(1, col).font = bold

        for row, f in enumerate(self.facturas_cargadas, start=2):
            ws2.cell(row, 1).value = f["clave"]
            ws2.cell(row, 2).value = f["archivo"]
            ws2.cell(row, 3).value = f["proveedor"]
            ws2.cell(row, 4).value = f["cuit_proveedor"]
            ws2.cell(row, 5).value = f["tipo"]
            ws2.cell(row, 6).value = f["punto_venta"]
            ws2.cell(row, 7).value = f["numero"]
            ws2.cell(row, 8).value = f["numero_completo"]
            ws2.cell(row, 9).value = f["fecha"]
            ws2.cell(row, 10).value = f["ruta"]

        for col in range(1, 11):
            col_letter = ws2.cell(1, col).column_letter
            ws2.column_dimensions[col_letter].width = 25

        wb.save(salida)

        messagebox.showinfo(
            "Listo",
            f"Resumen de compra/stock generado:\n\n{salida}\n\n"
            f"Este archivo es de revisión; todavía no es la plantilla oficial de comprobante de compra."
        )


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()