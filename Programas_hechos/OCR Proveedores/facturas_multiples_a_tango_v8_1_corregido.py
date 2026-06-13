
import os, re, json, sqlite3, tempfile, unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from dataclasses import dataclass
from difflib import SequenceMatcher
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from PIL import Image, ImageOps, ImageEnhance, ImageDraw
import fitz
import pytesseract

try:
    import cv2
except Exception:
    cv2 = None

# Ajustar si Tesseract está en otra ruta
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

APP_DB = "facturas_tango_v8.db"
CHECK_ON = "☑"
CHECK_OFF = "☐"


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    page: int


@dataclass
class LineBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    words: list
    score: int = 0


class BaseLocal:
    def __init__(self, db_path=APP_DB):
        self.db_path = db_path
        self.inicializar()

    def conectar(self):
        return sqlite3.connect(self.db_path)

    def inicializar(self):
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS facturas_procesadas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clave TEXT UNIQUE,
                proveedor TEXT,
                cuit TEXT,
                tipo TEXT,
                punto_venta TEXT,
                numero TEXT,
                numero_completo TEXT,
                fecha TEXT,
                archivo TEXT,
                ruta TEXT,
                fecha_proceso TEXT,
                archivo_resumen TEXT,
                archivo_alta TEXT
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS equivalencias (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cuit TEXT,
                proveedor TEXT,
                codigo_proveedor TEXT,
                descripcion_proveedor TEXT,
                codigo_tango TEXT,
                descripcion_tango TEXT,
                codigo_barras TEXT,
                veces_usado INTEGER DEFAULT 1,
                ultima_fecha TEXT,
                UNIQUE(cuit, codigo_proveedor)
            )
        """)
        con.commit()
        con.close()

    def factura_existe(self, clave):
        if not clave:
            return None
        con = self.conectar()
        cur = con.cursor()
        cur.execute("SELECT fecha_proceso, archivo_resumen FROM facturas_procesadas WHERE clave=?", (clave,))
        row = cur.fetchone()
        con.close()
        return row

    def guardar_factura_procesada(self, datos, archivo_resumen="", archivo_alta=""):
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            INSERT OR REPLACE INTO facturas_procesadas (
                clave, proveedor, cuit, tipo, punto_venta, numero, numero_completo,
                fecha, archivo, ruta, fecha_proceso, archivo_resumen, archivo_alta
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            datos.get("clave", ""), datos.get("proveedor", ""), datos.get("cuit_proveedor", ""),
            datos.get("tipo", ""), datos.get("punto_venta", ""), datos.get("numero", ""),
            datos.get("numero_completo", ""), datos.get("fecha", ""), datos.get("archivo", ""),
            datos.get("ruta", ""), datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            archivo_resumen, archivo_alta
        ))
        con.commit()
        con.close()

    def buscar_equivalencia(self, cuit, codigo_proveedor):
        if not cuit or not codigo_proveedor:
            return None
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            SELECT codigo_tango, descripcion_tango, codigo_barras
            FROM equivalencias
            WHERE cuit=? AND codigo_proveedor=?
        """, (str(cuit).strip(), str(codigo_proveedor).strip()))
        row = cur.fetchone()
        con.close()
        if not row:
            return None
        return {"codigo_tango": row[0], "descripcion_tango": row[1], "codigo_barras": row[2]}

    def guardar_equivalencia(self, cuit, proveedor, codigo_proveedor, descripcion_proveedor,
                             codigo_tango, descripcion_tango, codigo_barras):
        if not cuit or not codigo_proveedor or not codigo_tango:
            return
        con = self.conectar()
        cur = con.cursor()
        cur.execute("""
            INSERT INTO equivalencias (
                cuit, proveedor, codigo_proveedor, descripcion_proveedor,
                codigo_tango, descripcion_tango, codigo_barras, veces_usado, ultima_fecha
            ) VALUES (?, ?, ?, ?, ?, ?, ?, 1, ?)
            ON CONFLICT(cuit, codigo_proveedor) DO UPDATE SET
                proveedor=excluded.proveedor,
                descripcion_proveedor=excluded.descripcion_proveedor,
                codigo_tango=excluded.codigo_tango,
                descripcion_tango=excluded.descripcion_tango,
                codigo_barras=excluded.codigo_barras,
                veces_usado=veces_usado+1,
                ultima_fecha=excluded.ultima_fecha
        """, (
            str(cuit).strip(), str(proveedor).strip(), str(codigo_proveedor).strip(),
            str(descripcion_proveedor).strip(), str(codigo_tango).strip(),
            str(descripcion_tango).strip(), str(codigo_barras).strip(),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        con.commit()
        con.close()


class Normalizador:
    def quitar_acentos(self, texto):
        texto = unicodedata.normalize("NFD", str(texto))
        return "".join(c for c in texto if unicodedata.category(c) != "Mn")

    def normalizar(self, texto):
        if texto is None:
            return ""
        texto = self.quitar_acentos(str(texto).upper().strip())
        reemplazos = {
            "P/ACERO": "ACERO", "P ACERO": "ACERO", "PACERO": "ACERO",
            "P/ ACERO": "ACERO", "P/INOX": "INOX", "P INOX": "INOX",
            "INOXIDABLE": "INOX", "C/": "CON ", "S/": "SIN ",
            "MM.": "MM", "M.M.": "MM"
        }
        for a, b in reemplazos.items():
            texto = texto.replace(a, b)
        for ch in ["/", "-", ",", ".", "(", ")", "[", "]", "_", ":"]:
            texto = texto.replace(ch, " ")
        return re.sub(r"\s+", " ", texto).strip()

    def tokens(self, texto):
        return [t for t in self.normalizar(texto).split() if len(t) >= 2]


class CatalogoTango(Normalizador):
    def __init__(self):
        self.articulos = []

    def cargar(self, ruta_excel):
        wb = load_workbook(ruta_excel, data_only=True)
        if "Artículos" not in wb.sheetnames:
            raise Exception("El Excel no tiene una hoja llamada 'Artículos'.")
        ws = wb["Artículos"]
        self.articulos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            codigo = row[0] if len(row) > 0 else ""
            if not codigo:
                continue
            descripcion = row[1] if len(row) > 1 else ""
            desc_adic = row[2] if len(row) > 2 else ""
            sinonimo = row[3] if len(row) > 3 else ""
            codigo_barras = row[4] if len(row) > 4 else ""
            art = {
                "codigo": str(codigo).strip(),
                "descripcion": str(descripcion or "").strip(),
                "desc_adic": str(desc_adic or "").strip(),
                "sinonimo": str(sinonimo or "").strip(),
                "codigo_barras": str(codigo_barras or "").strip(),
            }
            art["_codigo_n"] = self.normalizar(art["codigo"])
            art["_descripcion_n"] = self.normalizar(art["descripcion"])
            art["_sinonimo_n"] = self.normalizar(art["sinonimo"])
            art["_codigo_barras_n"] = self.normalizar(art["codigo_barras"])
            art["_texto_completo_n"] = self.normalizar(
                f"{art['codigo']} {art['descripcion']} {art['desc_adic']} {art['sinonimo']} {art['codigo_barras']}"
            )
            self.articulos.append(art)

    def similitud(self, a, b):
        a = self.normalizar(a)
        b = self.normalizar(b)
        if not a or not b:
            return 0
        return int(SequenceMatcher(None, a, b).ratio() * 100)

    def score_tokens(self, a, b):
        ta, tb = set(self.tokens(a)), set(self.tokens(b))
        if not ta or not tb:
            return 0
        return int((len(ta.intersection(tb)) / len(ta)) * 100)

    def buscar(self, codigo_proveedor="", codigo_barras="", descripcion=""):
        cp = self.normalizar(codigo_proveedor)
        cb = self.normalizar(codigo_barras)
        desc = self.normalizar(descripcion)

        if cb:
            for art in self.articulos:
                if art["_codigo_barras_n"] and art["_codigo_barras_n"] == cb:
                    return art, "Código de barras exacto", 100

        if cp:
            for art in self.articulos:
                if art["_codigo_n"] == cp:
                    return art, "Código Tango exacto", 100
            for art in self.articulos:
                if art["_sinonimo_n"] and art["_sinonimo_n"] == cp:
                    return art, "Sinónimo exacto", 100
            for art in self.articulos:
                if cp in art["_descripcion_n"]:
                    return art, "Código proveedor en descripción", 96
            for art in self.articulos:
                if cp in art["_texto_completo_n"]:
                    return art, "Código proveedor contenido", 94

        mejor, mejor_score = None, 0
        if desc:
            for art in self.articulos:
                score = max(self.similitud(desc, art["_descripcion_n"]), self.score_tokens(desc, art["_descripcion_n"]))
                if score > mejor_score:
                    mejor, mejor_score = art, score
            if mejor and mejor_score >= 82:
                return mejor, "Similitud descripción", mejor_score

        return None, "", 0


class FacturaTableReader:
    def __init__(self):
        self.stop_words = [
            "total", "subtotal", "cae", "iva 21", "iva 10", "iva 27",
            "son pesos", "observaciones", "importe neto", "fecha de vto"
        ]

    def is_cuit(self, text):
        clean = str(text).strip()
        if re.fullmatch(r"\d{2}-\d{8}-\d", clean):
            return True
        digits = re.sub(r"\D", "", clean)
        return len(digits) == 11 and digits[:2] in ["20", "23", "24", "27", "30", "33", "34"]

    def normalizar_cuit(self, cuit):
        return re.sub(r"\D", "", str(cuit or ""))

    def abrir_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()
        return self.leer_pdf(ruta) if ext == ".pdf" else self.leer_imagen(ruta)

    def leer_pdf(self, ruta):
        doc = fitz.open(ruta)
        paginas = []
        for page_index, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(tempfile.gettempdir(), f"pagina_factura_{page_index}.png")
            pix.save(img_path)
            img = Image.open(img_path)
            palabras = []
            for w in page.get_text("words"):
                x0, y0, x1, y1, text = w[:5]
                text = str(text).strip()
                if text:
                    palabras.append(WordBox(x0 * 2, y0 * 2, x1 * 2, y1 * 2, text, page_index))
            paginas.append((img, palabras))
        return paginas

    def leer_imagen(self, ruta):
        img = Image.open(ruta)
        img_ocr = self.preparar_imagen(img)
        data = pytesseract.image_to_data(
            img_ocr, lang="spa", output_type=pytesseract.Output.DICT, config="--psm 6"
        )
        palabras = []
        for i, text in enumerate(data["text"]):
            text = str(text).strip()
            if not text:
                continue
            try:
                conf = int(float(data["conf"][i]))
            except Exception:
                conf = -1
            if conf < 25:
                continue
            x, y, w, h = data["left"][i], data["top"][i], data["width"][i], data["height"][i]
            palabras.append(WordBox(x, y, x + w, y + h, text, 1))
        return [(img, palabras)]

    def preparar_imagen(self, img):
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.2)
        return img

    def agrupar_lineas(self, palabras):
        palabras = sorted(palabras, key=lambda w: (w.y0, w.x0))
        lineas, actual = [], []
        for p in palabras:
            if not actual:
                actual = [p]
                continue
            y_prom = sum(w.y0 for w in actual) / len(actual)
            if abs(p.y0 - y_prom) <= 8:
                actual.append(p)
            else:
                lineas.append(self.crear_linea(actual))
                actual = [p]
        if actual:
            lineas.append(self.crear_linea(actual))
        return lineas

    def crear_linea(self, palabras):
        palabras = sorted(palabras, key=lambda w: w.x0)
        return LineBox(
            min(w.x0 for w in palabras), min(w.y0 for w in palabras),
            max(w.x1 for w in palabras), max(w.y1 for w in palabras),
            " ".join(w.text for w in palabras), palabras
        )

    def parse_decimal(self, t):
        t = str(t).replace("$", "").replace(" ", "").strip()
        if not t:
            return None
        if re.match(r"^-?\d{1,3}(\.\d{3})*,\d+$", t):
            t = t.replace(".", "").replace(",", ".")
        elif re.match(r"^-?\d+,\d+$", t):
            t = t.replace(",", ".")
        try:
            return float(t)
        except Exception:
            return None

    def es_numero(self, t):
        t = str(t).replace("$", "").strip()
        if self.is_cuit(t):
            return False
        return bool(re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2,4}$", t)) or bool(re.match(r"^-?\d+([,.]\d+)?$", t))

    def cantidad_columnas(self, linea):
        if len(linea.words) < 3:
            return 0
        gaps = 0
        for a, b in zip(linea.words, linea.words[1:]):
            if b.x0 - a.x1 > 25:
                gaps += 1
        return gaps + 1

    def puntuar_linea(self, linea):
        text = linea.text.lower()
        if any(self.is_cuit(w.text) for w in linea.words):
            linea.score = -10
            return linea.score
        if any(p in text for p in self.stop_words):
            linea.score = -5
            return linea.score
        nums = sum(1 for w in linea.words if self.es_numero(w.text))
        cols = self.cantidad_columnas(linea)
        score = 0
        if cols >= 3: score += 2
        if cols >= 5: score += 2
        if nums >= 2: score += 2
        if nums >= 4: score += 2
        if len(linea.words) >= 5: score += 1
        linea.score = score
        return score

    def detectar_zonas(self, lineas):
        for l in lineas:
            self.puntuar_linea(l)
        candidatas = [l for l in lineas if l.score >= 4]
        if not candidatas:
            return []
        grupos, grupo = [], [candidatas[0]]
        for l in candidatas[1:]:
            if l.y0 - grupo[-1].y1 < 75:
                grupo.append(l)
            else:
                grupos.append(grupo)
                grupo = [l]
        grupos.append(grupo)
        zonas = []
        for g in grupos:
            if len(g) < 2:
                continue
            zonas.append((min(l.x0 for l in g), min(l.y0 for l in g), max(l.x1 for l in g), max(l.y1 for l in g), g))
        return zonas

    def detectar_items(self, lineas):
        zonas = self.detectar_zonas(lineas)
        items = []
        for x0, y0, x1, y1, grupo in zonas:
            for linea in grupo:
                if linea.score < 4 or any(self.is_cuit(w.text) for w in linea.words):
                    continue
                item = self.linea_a_item(linea)
                if item:
                    items.append(item)
        return items, zonas

    def linea_a_item(self, linea):
        codigo_proveedor = ""
        descripcion_parts = []
        numeros_detectados = []
        for w in linea.words:
            t = w.text.strip()
            if self.is_cuit(t):
                return None
            if self.es_numero(t):
                numeros_detectados.append(t)
                continue
            if not codigo_proveedor and re.match(r"^[A-Z0-9][A-Z0-9.\-/]{2,}$", t, re.IGNORECASE):
                codigo_proveedor = t[:15]
                continue
            if t.upper() in ["U", "UN", "KG", "M", "L", "%"]:
                continue
            if re.match(r"^-?\d+([,.]\d+)?%?$", t):
                continue
            descripcion_parts.append(t)

        vals = [(n, self.parse_decimal(n)) for n in numeros_detectados]
        vals = [(n, v) for n, v in vals if v is not None]

        iva = ""
        if len(vals) >= 4:
            posibles = [(n, v) for n, v in vals if abs(v - 21.0) < 0.01 or abs(v - 10.5) < 0.01 or abs(v - 27.0) < 0.01 or abs(v - 0.0) < 0.01]
            if posibles:
                no_cero = [x for x in posibles if abs(x[1]) > 0.01]
                elegido = no_cero[0] if no_cero else posibles[0]
                iva = elegido[0]
                vals = [x for x in vals if x[0] != elegido[0]]

        cantidad = vals[0][0] if vals else ""
        precio = vals[-2][0] if len(vals) >= 2 else ""
        importe = vals[-1][0] if vals else ""
        descripcion = re.sub(r"\s+", " ", " ".join(descripcion_parts)).replace(" /", "").replace("/", "").strip()

        if not descripcion and not codigo_proveedor:
            return None

        advertencia = self.validar_calculo(cantidad, precio, importe)
        return {
            "usar": True, "existe_tango": "No", "accion": "Alta artículo", "codigo_tango": "",
            "descripcion": descripcion[:50], "desc_adic": "", "sinonimo": codigo_proveedor[:15],
            "codigo_barras": "", "cantidad": cantidad, "precio": precio, "importe": importe,
            "iva": iva, "advertencia": advertencia, "tipo": "Simple", "escala": "No usa",
            "linea_detectada": linea.text, "score": linea.score, "coincidencia": "", "match_score": 0,
        }

    def validar_calculo(self, cantidad, precio, importe):
        c, p, i = self.parse_decimal(cantidad), self.parse_decimal(precio), self.parse_decimal(importe)
        if c is None or p is None or i is None:
            return ""
        tolerancia = max(1.0, abs(i) * 0.01)
        if abs((c * p) - i) > tolerancia:
            return "Revisar cantidad/precio/importe"
        return "OK"

    def extraer_texto_documento(self, ruta):
        ext = os.path.splitext(ruta)[1].lower()
        if ext == ".pdf":
            try:
                doc = fitz.open(ruta)
                return "\n".join(page.get_text("text") for page in doc)
            except Exception:
                return ""
        try:
            img = self.preparar_imagen(Image.open(ruta))
            try:
                return pytesseract.image_to_string(img, lang="spa")
            except Exception:
                return pytesseract.image_to_string(img)
        except Exception:
            return ""

    def detectar_proveedor(self, texto):
        lineas = [l.strip() for l in texto.splitlines() if l.strip()]
        for l in lineas[:20]:
            may = l.upper()
            if "S.R.L" in may or "SRL" in may or "S.A" in may or " SA" in may:
                return l
        for l in lineas[:10]:
            if len(l) > 4 and not re.search(r"\d{2}/\d{2}/\d{2,4}", l):
                return l
        return ""

    def detectar_cuit_proveedor(self, texto):
        for patron in [
            r"CUIT\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
            r"C\.?U\.?I\.?T\.?\s*[:\-]?\s*(\d{2}[-\s]?\d{8}[-\s]?\d)",
        ]:
            for m in re.findall(patron, texto, re.IGNORECASE):
                cuit = self.normalizar_cuit(m)
                if len(cuit) == 11 and cuit[:2] in ["20", "23", "24", "27", "30", "33", "34"]:
                    return cuit
        m = re.search(r"\b((?:20|23|24|27|30|33|34)\d{9})\b", texto)
        return m.group(1) if m else ""

    def detectar_tipo_comprobante(self, texto):
        may = texto.upper()
        if "FACTURA" in may:
            if re.search(r"FACTURA\s*A", may) or re.search(r"\bCOD\.?\s*0?1\b", may): return "FACTURA A"
            if re.search(r"FACTURA\s*B", may) or re.search(r"\bCOD\.?\s*0?6\b", may): return "FACTURA B"
            if re.search(r"FACTURA\s*C", may) or re.search(r"\bCOD\.?\s*0?11\b", may): return "FACTURA C"
            return "FACTURA"
        if "REMITO" in may:
            return "REMITO"
        return ""

    def detectar_numero_comprobante(self, texto):
        patrones = [
            r"Punto\s+de\s+Venta\s*[:\-]?\s*(\d{4,5}).{0,80}?Comp\.?\s*Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"Pto\.?\s*Vta\.?\s*[:\-]?\s*(\d{4,5}).{0,80}?Nro\.?\s*[:\-]?\s*(\d{6,8})",
            r"(?:FACTURA|REMITO)\s*(?:N[°º]?)?\s*[:\-]?\s*(\d{4,5})[-\s]?(\d{6,8})",
            r"\b(\d{4,5})[-\s](\d{6,8})\b",
        ]
        for p in patrones:
            m = re.search(p, texto, re.IGNORECASE | re.DOTALL)
            if m:
                pv, nro = m.group(1).zfill(4), m.group(2).zfill(8)
                return pv, nro, f"{pv}-{nro}"
        return "", "", ""

    def detectar_fecha(self, texto):
        m = re.search(r"(?:Fecha\s+de\s+Emisi[oó]n|Fecha)\s*[:\-]?\s*(\d{1,2}/\d{1,2}/\d{2,4})", texto, re.IGNORECASE)
        if m: return m.group(1)
        m = re.search(r"\b(\d{1,2}/\d{1,2}/\d{2,4})\b", texto)
        return m.group(1) if m else ""

    def detectar_total(self, texto):
        posibles = re.findall(r"(?:TOTAL|Total)\s*[:$ ]+\s*([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})", texto)
        return posibles[-1] if posibles else ""

    def clave_factura(self, datos):
        cuit = self.normalizar_cuit(datos.get("cuit_proveedor", ""))
        proveedor = str(datos.get("proveedor", "")).strip().upper()
        tipo = str(datos.get("tipo", "")).strip().upper()
        punto = str(datos.get("punto_venta", "")).strip().zfill(4) if datos.get("punto_venta") else ""
        numero = str(datos.get("numero", "")).strip().zfill(8) if datos.get("numero") else ""
        if cuit and punto and numero:
            return f"CUIT:{cuit}|TIPO:{tipo}|PV:{punto}|N:{numero}"
        if proveedor and punto and numero:
            return f"PROV:{re.sub(r'\\s+', ' ', proveedor)}|TIPO:{tipo}|PV:{punto}|N:{numero}"
        return f"ARCHIVO:{datos.get('archivo','')}"

    def detectar_datos_factura(self, ruta):
        texto = self.extraer_texto_documento(ruta)
        archivo = os.path.basename(ruta)
        proveedor = self.detectar_proveedor(texto)
        cuit = self.detectar_cuit_proveedor(texto)
        tipo = self.detectar_tipo_comprobante(texto)
        pv, nro, completo = self.detectar_numero_comprobante(texto)
        fecha = self.detectar_fecha(texto)
        total = self.detectar_total(texto)
        partes = []
        if proveedor: partes.append(proveedor[:30])
        if completo: partes.append(completo)
        if fecha: partes.append(fecha)
        etiqueta = " | ".join(partes) if partes else archivo
        datos = {
            "archivo": archivo, "ruta": ruta, "proveedor": proveedor, "cuit_proveedor": cuit,
            "tipo": tipo, "punto_venta": pv, "numero": nro, "numero_completo": completo,
            "fecha": fecha, "total": total, "etiqueta": etiqueta,
        }
        datos["clave"] = self.clave_factura(datos)
        return datos

    def analizar(self, ruta):
        paginas = self.abrir_documento(ruta)
        datos = self.detectar_datos_factura(ruta)
        resultados, todos_items = [], []
        for page_num, (img, palabras) in enumerate(paginas, start=1):
            lineas = self.agrupar_lineas(palabras)
            items, zonas = self.detectar_items(lineas)
            debug_path = self.dibujar_debug(img, lineas, zonas, page_num, datos["archivo"])
            for item in items:
                item.update({
                    "pagina": page_num, "factura_archivo": datos["archivo"], "factura_etiqueta": datos["etiqueta"],
                    "proveedor_factura": datos["proveedor"], "cuit_proveedor": datos["cuit_proveedor"],
                    "tipo_factura": datos["tipo"], "punto_venta": datos["punto_venta"],
                    "nro_factura": datos["numero"], "numero_completo": datos["numero_completo"],
                    "fecha_factura": datos["fecha"], "total_factura": datos["total"],
                    "ruta_factura": datos["ruta"], "clave_factura": datos["clave"],
                })
                todos_items.append(item)
            resultados.append((datos["archivo"], page_num, debug_path, lineas, zonas))
        return todos_items, resultados, datos

    def dibujar_debug(self, img, lineas, zonas, page_num, nombre_archivo):
        debug = img.convert("RGB")
        draw = ImageDraw.Draw(debug)
        for linea in lineas:
            if linea.score >= 4:
                draw.rectangle([linea.x0, linea.y0, linea.x1, linea.y1], outline="blue", width=2)
                draw.text((linea.x0, max(0, linea.y0 - 14)), f"S{linea.score}", fill="blue")
        for x0, y0, x1, y1, grupo in zonas:
            draw.rectangle([x0, y0, x1, y1], outline="red", width=4)
        base = re.sub(r"[^A-Za-z0-9_-]+", "_", nombre_archivo)
        out = os.path.join(tempfile.gettempdir(), f"debug_{base}_pagina_{page_num}.png")
        debug.save(out)
        return out


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Facturas múltiples → Tango V8.1")
        self.root.geometry("1760x860")

        self.base = BaseLocal()
        self.reader = FacturaTableReader()
        self.catalogo = CatalogoTango()

        self.ruta_modelo_tango = None
        self.ruta_catalogo_tango = None
        self.resultados_debug = []
        self.facturas_cargadas = []
        self.claves_facturas = set()
        self.filtro_actual = tk.StringVar(value="Ver todos")
        self._todos_items = []

        self.crear_ui()

    def configurar_estilo(self):
        style = ttk.Style()
        try: style.theme_use("clam")
        except Exception: pass
        style.configure("TButton", padding=5)
        style.configure("Treeview", rowheight=24)
        style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        style.configure("TLabel", font=("Segoe UI", 9))

    def crear_ui(self):
        self.configurar_estilo()
        main = ttk.Frame(self.root, padding=8)
        main.pack(fill="both", expand=True)

        top = ttk.Frame(main)
        top.pack(fill="x", pady=(0, 8))
        ttk.Button(top, text="1. Artículos Tango", command=self.abrir_catalogo).pack(side="left", padx=3)
        ttk.Button(top, text="2. Modelo Tango", command=self.abrir_modelo).pack(side="left", padx=3)
        ttk.Button(top, text="3. Agregar factura(s)", command=self.agregar_facturas).pack(side="left", padx=3)
        ttk.Button(top, text="Sacar foto", command=self.sacar_foto).pack(side="left", padx=3)
        ttk.Button(top, text="Ver debug", command=self.ver_debug).pack(side="left", padx=3)
        ttk.Button(top, text="Abrir factura", command=self.abrir_factura_original).pack(side="left", padx=3)
        ttk.Button(top, text="Guardar sesión", command=self.guardar_sesion).pack(side="left", padx=3)
        ttk.Button(top, text="Cargar sesión", command=self.cargar_sesion).pack(side="left", padx=3)
        ttk.Button(top, text="Aprender equivalencia", command=self.aprender_equivalencia_seleccionada).pack(side="left", padx=3)

        filtro_frame = ttk.Frame(main)
        filtro_frame.pack(fill="x", pady=(0, 6))
        ttk.Label(filtro_frame, text="Filtro:").pack(side="left")
        combo = ttk.Combobox(
            filtro_frame, textvariable=self.filtro_actual, state="readonly", width=24,
            values=["Ver todos", "Solo seleccionados", "Solo nuevos", "Solo existentes",
                    "Sin código Tango", "Baja confianza", "Destildados", "Con advertencias"]
        )
        combo.pack(side="left", padx=5)
        combo.bind("<<ComboboxSelected>>", lambda e: self.aplicar_filtro())

        self.lbl_catalogo = ttk.Label(main, text="Artículos Tango: no seleccionado")
        self.lbl_catalogo.pack(fill="x")
        self.lbl_modelo = ttk.Label(main, text="Modelo Tango: no seleccionado")
        self.lbl_modelo.pack(fill="x")
        self.lbl_facturas = ttk.Label(main, text="Facturas cargadas: 0")
        self.lbl_facturas.pack(fill="x", pady=(0, 6))

        self.columns = (
            "usar", "factura", "proveedor", "cuit", "tipo", "numero", "fecha",
            "existe", "accion", "codigo_tango", "descripcion", "sinonimo",
            "codigo_barras", "cantidad", "precio", "importe", "iva", "advertencia",
            "match", "match_score", "tipo_art", "escala", "pagina", "ruta", "linea"
        )
        headers = {
            "usar": "✓", "factura": "Factura", "proveedor": "Proveedor", "cuit": "CUIT",
            "tipo": "Tipo", "numero": "N°", "fecha": "Fecha", "existe": "Existe",
            "accion": "Acción", "codigo_tango": "Código Tango", "descripcion": "Descripción",
            "sinonimo": "Cod. proveedor", "codigo_barras": "Cód. barras", "cantidad": "Cantidad",
            "precio": "Precio", "importe": "Importe", "iva": "IVA %", "advertencia": "Advertencia",
            "match": "Coincidencia", "match_score": "Conf.", "tipo_art": "Tipo art.",
            "escala": "Escala", "pagina": "Pág.", "ruta": "Ruta factura", "linea": "Línea detectada",
        }
        widths = {
            "usar": 45, "factura": 160, "proveedor": 200, "cuit": 105, "tipo": 90,
            "numero": 110, "fecha": 85, "existe": 65, "accion": 145, "codigo_tango": 125,
            "descripcion": 320, "sinonimo": 130, "codigo_barras": 125, "cantidad": 85,
            "precio": 90, "importe": 95, "iva": 70, "advertencia": 190, "match": 210,
            "match_score": 60, "tipo_art": 80, "escala": 80, "pagina": 50, "ruta": 260,
            "linea": 420,
        }

        frame_tabla = ttk.Frame(main)
        frame_tabla.pack(fill="both", expand=True, pady=6)
        scroll_y = ttk.Scrollbar(frame_tabla, orient="vertical")
        scroll_y.pack(side="right", fill="y")
        scroll_x = ttk.Scrollbar(frame_tabla, orient="horizontal")
        scroll_x.pack(side="bottom", fill="x")

        self.tabla = ttk.Treeview(
            frame_tabla, columns=self.columns, show="headings", height=25,
            yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set
        )
        scroll_y.config(command=self.tabla.yview)
        scroll_x.config(command=self.tabla.xview)

        for c in self.columns:
            self.tabla.heading(c, text=headers[c])
            self.tabla.column(c, width=widths[c], stretch=False)

        self.tabla.tag_configure("verde", background="#d9f5d6")
        self.tabla.tag_configure("amarillo", background="#fff3bf")
        self.tabla.tag_configure("rojo", background="#ffd6d6")
        self.tabla.tag_configure("gris", background="#e9ecef")
        self.tabla.tag_configure("azul", background="#d6eaff")
        self.tabla.tag_configure("advertencia", background="#ffd8a8")
        self.tabla.pack(side="left", fill="both", expand=True)
        self.tabla.bind("<Double-1>", self.doble_click)

        bottom = ttk.Frame(main)
        bottom.pack(fill="x", pady=(8, 0))
        ttk.Button(bottom, text="Seleccionar todo", command=self.seleccionar_todo).pack(side="left", padx=3)
        ttk.Button(bottom, text="Borrar selección", command=self.borrar_seleccion).pack(side="left", padx=3)
        ttk.Button(bottom, text="Invertir selección", command=self.invertir_seleccion).pack(side="left", padx=3)
        ttk.Button(bottom, text="Agregar manual", command=self.agregar_manual).pack(side="left", padx=3)
        ttk.Button(bottom, text="Eliminar fila", command=self.eliminar_fila).pack(side="left", padx=3)
        ttk.Button(bottom, text="Limpiar todo", command=self.limpiar_todo).pack(side="left", padx=3)
        ttk.Button(bottom, text="Recomparar", command=self.recomparar).pack(side="left", padx=3)
        ttk.Button(bottom, text="Reporte revisión", command=self.generar_reporte_revision).pack(side="right", padx=3)
        ttk.Button(bottom, text="Alta artículos", command=self.generar_excel_alta_articulos).pack(side="right", padx=3)
        ttk.Button(bottom, text="Resumen compra/stock", command=self.generar_resumen_compra_stock).pack(side="right", padx=3)

        self.lbl_estado = ttk.Label(
            main,
            text="V8.1: filtro corregido. Historial, equivalencias, colores, sesión, IVA y validación cantidad × precio.",
            foreground="#0b5ed7"
        )
        self.lbl_estado.pack(fill="x", pady=(6, 0))

    def refrescar_indice_items(self):
        visibles = list(self.tabla.get_children(""))
        existentes = []
        for item_id in self._todos_items:
            try:
                self.tabla.item(item_id)
                existentes.append(item_id)
            except Exception:
                pass
        for item_id in visibles:
            if item_id not in existentes:
                existentes.append(item_id)
        self._todos_items = existentes

    def abrir_catalogo(self):
        ruta = filedialog.askopenfilename(title="Seleccionar Artículos.xlsx exportado desde Tango", filetypes=[("Excel", "*.xlsx")])
        if not ruta: return
        try:
            self.catalogo.cargar(ruta)
            self.ruta_catalogo_tango = ruta
            self.lbl_catalogo.config(text=f"Artículos Tango: {ruta} ({len(self.catalogo.articulos)} artículos)")
            self.recomparar(silencioso=True)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def abrir_modelo(self):
        ruta = filedialog.askopenfilename(title="Seleccionar Excel modelo exportado desde Tango", filetypes=[("Excel", "*.xlsx")])
        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def factura_duplicada_en_sesion(self, datos):
        return datos.get("clave", "") in self.claves_facturas

    def confirmar_factura_duplicada(self, datos, fuente):
        msg = (
            f"Esta factura ya fue cargada ({fuente}).\n\n"
            f"Proveedor: {datos.get('proveedor','')}\nCUIT: {datos.get('cuit_proveedor','')}\n"
            f"Tipo: {datos.get('tipo','')}\nNúmero: {datos.get('numero_completo','')}\n"
            f"Fecha: {datos.get('fecha','')}\nArchivo: {datos.get('archivo','')}\n\n"
            "¿Querés cargarla igualmente?"
        )
        return messagebox.askyesno("Factura duplicada", msg)

    def agregar_facturas(self):
        rutas = filedialog.askopenfilenames(
            title="Seleccionar una o varias facturas",
            filetypes=[("Facturas", "*.pdf *.jpg *.jpeg *.png *.bmp *.tif *.tiff"), ("Todos", "*.*")]
        )
        if not rutas: return
        total_items, cargadas, salteadas = 0, 0, 0
        for ruta in rutas:
            try:
                datos_previos = self.reader.detectar_datos_factura(ruta)
                if self.factura_duplicada_en_sesion(datos_previos):
                    if not self.confirmar_factura_duplicada(datos_previos, "en esta sesión"):
                        salteadas += 1
                        continue
                historial = self.base.factura_existe(datos_previos.get("clave", ""))
                if historial:
                    if not self.confirmar_factura_duplicada(datos_previos, "historial permanente"):
                        salteadas += 1
                        continue
                items, debug, datos = self.reader.analizar(ruta)
                self.resultados_debug.extend(debug)
                self.facturas_cargadas.append(datos)
                self.claves_facturas.add(datos["clave"])
                for item in items:
                    self.insertar_item(item)
                total_items += len(items)
                cargadas += 1
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo procesar:\n{ruta}\n\n{e}")
        if self.catalogo.articulos:
            self.recomparar(silencioso=True)
        else:
            self.actualizar_tags()
            self.aplicar_filtro()
        self.actualizar_label_facturas()
        messagebox.showinfo("Lectura terminada", f"Facturas agregadas: {cargadas}\nFacturas salteadas por duplicadas: {salteadas}\nLíneas candidatas agregadas: {total_items}")

    def sacar_foto(self):
        if cv2 is None:
            messagebox.showerror("Error", "Falta instalar opencv-python")
            return
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "No se pudo abrir la cámara.")
            return
        messagebox.showinfo("Cámara", "ESPACIO captura / ESC cancela")
        ruta_foto = None
        while True:
            ret, frame = cap.read()
            if not ret: break
            cv2.imshow("Capturar factura", frame)
            key = cv2.waitKey(1)
            if key == 27: break
            if key == 32:
                ruta_foto = os.path.join(tempfile.gettempdir(), f"factura_capturada_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg")
                cv2.imwrite(ruta_foto, frame)
                break
        cap.release()
        cv2.destroyAllWindows()
        if ruta_foto:
            self.procesar_ruta_unica(ruta_foto)

    def procesar_ruta_unica(self, ruta):
        try:
            datos_previos = self.reader.detectar_datos_factura(ruta)
            if self.factura_duplicada_en_sesion(datos_previos):
                if not self.confirmar_factura_duplicada(datos_previos, "en esta sesión"): return
            historial = self.base.factura_existe(datos_previos.get("clave", ""))
            if historial:
                if not self.confirmar_factura_duplicada(datos_previos, "historial permanente"): return
            items, debug, datos = self.reader.analizar(ruta)
            self.resultados_debug.extend(debug)
            self.facturas_cargadas.append(datos)
            self.claves_facturas.add(datos["clave"])
            for item in items:
                self.insertar_item(item)
            if self.catalogo.articulos:
                self.recomparar(silencioso=True)
            else:
                self.actualizar_tags()
                self.aplicar_filtro()
            self.actualizar_label_facturas()
            messagebox.showinfo("Foto procesada", f"Líneas candidatas agregadas: {len(items)}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def insertar_item(self, item):
        valores = (
            CHECK_ON if item["usar"] else CHECK_OFF,
            item.get("factura_etiqueta", ""), item.get("proveedor_factura", ""), item.get("cuit_proveedor", ""),
            item.get("tipo_factura", ""), item.get("numero_completo", ""), item.get("fecha_factura", ""),
            item["existe_tango"], item["accion"], item["codigo_tango"], item["descripcion"], item["sinonimo"],
            item["codigo_barras"], item["cantidad"], item["precio"], item["importe"], item.get("iva", ""),
            item.get("advertencia", ""), item.get("coincidencia", ""), item.get("match_score", ""),
            item["tipo"], item["escala"], item.get("pagina", ""), item.get("ruta_factura", ""), item["linea_detectada"],
        )
        tag = self.tag_para_valores(valores)
        item_id = self.tabla.insert("", "end", values=valores, tags=(tag,))
        if item_id not in self._todos_items:
            self._todos_items.append(item_id)

    def tag_para_valores(self, v):
        usar, existe, codigo, advertencia, match, conf = v[0], v[7], v[9], v[17], v[18], v[19]
        if usar == CHECK_OFF: return "gris"
        if advertencia and advertencia != "OK": return "advertencia"
        if match == "Equivalencia aprendida": return "azul"
        try: conf = int(conf)
        except Exception: conf = 0
        if existe == "Sí" and conf >= 90: return "verde"
        if existe == "Sí" and conf < 90: return "amarillo"
        if existe == "No" or not codigo: return "rojo"
        return ""

    def actualizar_tags(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
                if v:
                    self.tabla.item(item_id, tags=(self.tag_para_valores(v),))
            except Exception:
                pass

    def recomparar(self, silencioso=False):
        if not self.catalogo.articulos:
            if not silencioso:
                messagebox.showinfo("Catálogo no cargado", "Primero cargá el Excel Artículos.xlsx exportado desde Tango.")
            return
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try: v = list(self.tabla.item(item_id, "values"))
            except Exception: continue
            if not v: continue
            cuit, codigo_proveedor, codigo_barras, descripcion = v[3], v[11], v[12], v[10]
            eq = self.base.buscar_equivalencia(cuit, codigo_proveedor)
            if eq:
                v[7], v[8], v[9], v[10], v[12], v[18], v[19] = "Sí", "Cargar stock / compra", eq["codigo_tango"], eq["descripcion_tango"] or descripcion, eq["codigo_barras"], "Equivalencia aprendida", "100"
            else:
                art, metodo, score = self.catalogo.buscar(codigo_proveedor, codigo_barras, descripcion)
                if art:
                    v[7], v[8], v[9], v[10], v[12], v[18], v[19] = "Sí", "Cargar stock / compra", art["codigo"], art["descripcion"], art["codigo_barras"], metodo, str(score)
                else:
                    v[7], v[8], v[18], v[19] = "No", "Alta artículo", "", ""
            self.tabla.item(item_id, values=v)
        self.actualizar_tags()
        self.aplicar_filtro()
        if not silencioso:
            messagebox.showinfo("Listo", "Comparación terminada.")

    def doble_click(self, event):
        item_id = self.tabla.focus()
        if not item_id: return
        col = self.tabla.identify_column(event.x)
        idx = int(col.replace("#", "")) - 1
        valores = list(self.tabla.item(item_id, "values"))
        if idx == 0:
            valores[0] = CHECK_OFF if valores[0] == CHECK_ON else CHECK_ON
            self.tabla.item(item_id, values=valores)
            self.actualizar_tags()
            self.aplicar_filtro()
            return
        nombres = [
            "Usar", "Factura", "Proveedor", "CUIT", "Tipo", "Número", "Fecha", "Existe", "Acción",
            "Código Tango", "Descripción", "Cod. proveedor", "Cód. barras", "Cantidad", "Precio",
            "Importe", "IVA %", "Advertencia", "Coincidencia", "Confianza", "Tipo art.", "Escala",
            "Página", "Ruta factura", "Línea detectada"
        ]
        nuevo = simpledialog.askstring("Editar", nombres[idx], initialvalue=valores[idx])
        if nuevo is not None:
            valores[idx] = nuevo
            self.tabla.item(item_id, values=valores)
            self.actualizar_tags()
            self.aplicar_filtro()

    def seleccionar_todo(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = list(self.tabla.item(item_id, "values"))
                if v:
                    v[0] = CHECK_ON
                    self.tabla.item(item_id, values=v)
            except Exception:
                pass
        self.actualizar_tags(); self.aplicar_filtro()

    def borrar_seleccion(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = list(self.tabla.item(item_id, "values"))
                if v:
                    v[0] = CHECK_OFF
                    self.tabla.item(item_id, values=v)
            except Exception:
                pass
        self.actualizar_tags(); self.aplicar_filtro()

    def invertir_seleccion(self):
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = list(self.tabla.item(item_id, "values"))
                if v:
                    v[0] = CHECK_OFF if v[0] == CHECK_ON else CHECK_ON
                    self.tabla.item(item_id, values=v)
            except Exception:
                pass
        self.actualizar_tags(); self.aplicar_filtro()

    def agregar_manual(self):
        self.insertar_item({
            "usar": True, "factura_etiqueta": "Manual", "proveedor_factura": "", "cuit_proveedor": "",
            "tipo_factura": "", "numero_completo": "", "fecha_factura": "", "existe_tango": "No",
            "accion": "Alta artículo", "codigo_tango": "", "descripcion": "", "desc_adic": "",
            "sinonimo": "", "codigo_barras": "", "cantidad": "", "precio": "", "importe": "", "iva": "",
            "advertencia": "", "tipo": "Simple", "escala": "No usa", "pagina": "", "ruta_factura": "",
            "linea_detectada": "Manual", "coincidencia": "", "match_score": "",
        })
        self.aplicar_filtro()

    def eliminar_fila(self):
        for item_id in self.tabla.selection():
            try: self.tabla.delete(item_id)
            except Exception: pass
            if item_id in self._todos_items:
                self._todos_items.remove(item_id)

    def limpiar_todo(self):
        if not messagebox.askyesno("Confirmar", "¿Querés borrar todas las facturas y filas cargadas?"):
            return
        self.limpiar_todo_sin_preguntar()

    def limpiar_todo_sin_preguntar(self):
        self.refrescar_indice_items()
        for item_id in list(self._todos_items):
            try: self.tabla.delete(item_id)
            except Exception: pass
        self.facturas_cargadas, self.resultados_debug, self.claves_facturas, self._todos_items = [], [], set(), []
        self.actualizar_label_facturas()

    def aplicar_filtro(self):
        filtro = self.filtro_actual.get()
        self.refrescar_indice_items()
        for item_id in self._todos_items:
            try:
                v = self.tabla.item(item_id, "values")
                if not v: continue
                if self.pasa_filtro(v, filtro):
                    self.tabla.move(item_id, "", "end")
                else:
                    self.tabla.detach(item_id)
            except Exception:
                pass

    def pasa_filtro(self, v, filtro):
        usar, existe, codigo, advertencia, conf = v[0], v[7], v[9], v[17], v[19]
        try: conf_int = int(conf)
        except Exception: conf_int = 0
        if filtro == "Ver todos": return True
        if filtro == "Solo seleccionados": return usar == CHECK_ON
        if filtro == "Solo nuevos": return existe == "No"
        if filtro == "Solo existentes": return existe == "Sí"
        if filtro == "Sin código Tango": return not codigo
        if filtro == "Baja confianza": return existe == "Sí" and conf_int < 90
        if filtro == "Destildados": return usar == CHECK_OFF
        if filtro == "Con advertencias": return bool(advertencia and advertencia != "OK")
        return True

    def actualizar_label_facturas(self):
        nombres = []
        for f in self.facturas_cargadas:
            etiqueta = f.get("numero_completo") or f.get("archivo", "")
            proveedor = f.get("proveedor", "")
            if proveedor:
                etiqueta = f"{proveedor[:25]} {etiqueta}"
            nombres.append(etiqueta)
        detalle = " | ".join(nombres[:4]) + (f" | ... (+{len(nombres)-4})" if len(nombres) > 4 else "")
        self.lbl_facturas.config(text=f"Facturas cargadas: {len(self.facturas_cargadas)}   {detalle}")

    def ver_debug(self):
        if not self.resultados_debug:
            messagebox.showerror("Error", "Primero agregá una factura.")
            return
        for archivo, pagina, debug_path, lineas, zonas in self.resultados_debug:
            try: os.startfile(debug_path)
            except Exception as e:
                messagebox.showerror("Error", str(e)); return

    def abrir_factura_original(self):
        item_id = self.tabla.focus()
        if not item_id:
            messagebox.showerror("Error", "Seleccioná una fila.")
            return
        v = self.tabla.item(item_id, "values")
        ruta = v[23]
        if not ruta or not os.path.exists(ruta):
            messagebox.showerror("Error", "No se encontró la ruta de la factura original.")
            return
        os.startfile(ruta)

    def leer_filas(self, solo_visibles=False):
        ids = self.tabla.get_children("") if solo_visibles else self._todos_items
        filas = []
        for item_id in ids:
            try: v = self.tabla.item(item_id, "values")
            except Exception: continue
            if not v: continue
            filas.append({
                "usar": v[0], "factura": v[1], "proveedor": v[2], "cuit": v[3],
                "tipo_comprobante": v[4], "numero_comprobante": v[5], "fecha": v[6],
                "existe": v[7], "accion": v[8], "codigo_tango": v[9].strip(),
                "descripcion": v[10].strip(), "sinonimo": v[11].strip(), "codigo_barras": v[12].strip(),
                "cantidad": v[13].strip(), "precio": v[14].strip(), "importe": v[15].strip(),
                "iva": v[16].strip(), "advertencia": v[17].strip(), "coincidencia": v[18].strip(),
                "match_score": v[19].strip(), "tipo": v[20].strip() or "Simple", "escala": v[21].strip() or "No usa",
                "pagina": v[22], "ruta": v[23], "linea": v[24],
            })
        return filas

    def deduplicar_nuevos(self, nuevos):
        normalizador = CatalogoTango()
        vistos, resultado, duplicados = {}, [], []
        for f in nuevos:
            clave = (normalizador.normalizar(f["sinonimo"]), normalizador.normalizar(f["codigo_barras"]), normalizador.normalizar(f["descripcion"]))
            if clave in vistos:
                duplicados.append((f, vistos[clave])); continue
            vistos[clave] = f; resultado.append(f)
        return resultado, duplicados

    def aprender_equivalencia_seleccionada(self):
        sel = self.tabla.selection()
        if not sel:
            messagebox.showerror("Error", "Seleccioná una fila para aprender la equivalencia.")
            return
        aprendidas = 0
        for item_id in sel:
            v = self.tabla.item(item_id, "values")
            cuit, proveedor, cod_prov, linea, cod_tango, desc, barras = v[3], v[2], v[11], v[24], v[9], v[10], v[12]
            if not cuit or not cod_prov or not cod_tango:
                continue
            self.base.guardar_equivalencia(cuit, proveedor, cod_prov, linea, cod_tango, desc, barras)
            aprendidas += 1
        messagebox.showinfo("Listo", f"Equivalencias aprendidas: {aprendidas}")
        self.recomparar(silencioso=True)

    def aprender_equivalencias_automaticas(self):
        for f in self.leer_filas():
            if f["usar"] != CHECK_ON: continue
            if not f["cuit"] or not f["sinonimo"] or not f["codigo_tango"]: continue
            self.base.guardar_equivalencia(
                f["cuit"], f["proveedor"], f["sinonimo"], f["linea"],
                f["codigo_tango"], f["descripcion"], f["codigo_barras"]
            )

    def generar_excel_alta_articulos(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo Tango.")
            return
        filas = self.leer_filas()
        nuevos = [f for f in filas if f["usar"] == CHECK_ON and f["existe"] == "No"]
        if not nuevos:
            messagebox.showinfo("Sin artículos nuevos", "No hay artículos nuevos para dar de alta.")
            return
        if [f for f in nuevos if not f["codigo_tango"]]:
            messagebox.showerror("Faltan códigos Tango", "Hay artículos nuevos sin Código Tango.\n\nCompletalos antes de generar el Excel.")
            return
        nuevos_unicos, duplicados = self.deduplicar_nuevos(nuevos)
        if duplicados:
            messagebox.showinfo("Duplicados agrupados", f"Se encontraron {len(duplicados)} artículos nuevos repetidos.\nSe exportará solo una vez cada artículo nuevo.")
        salida = filedialog.asksaveasfilename(title="Guardar Excel Alta Artículos Tango", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Alta_Articulos_Tango.xlsx")
        if not salida: return
        try:
            wb = load_workbook(self.ruta_modelo_tango)
            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja 'Artículos'."); return
            ws = wb["Artículos"]
            if ws.max_row > 1: ws.delete_rows(2, ws.max_row - 1)
            for fila, item in enumerate(nuevos_unicos, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = ""
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]
            wb.save(salida)
            self.aprender_equivalencias_automaticas()
            messagebox.showinfo("Listo", f"Excel de alta de artículos generado:\n\n{salida}\n\nArtículos nuevos únicos: {len(nuevos_unicos)}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def generar_resumen_compra_stock(self):
        filas = [f for f in self.leer_filas() if f["usar"] == CHECK_ON]
        if not filas:
            messagebox.showerror("Error", "No hay filas seleccionadas."); return
        salida = filedialog.asksaveasfilename(title="Guardar resumen compra/stock", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Resumen_Compra_Stock.xlsx")
        if not salida: return
        self.crear_excel_resumen(salida, filas)
        for f in self.facturas_cargadas:
            self.base.guardar_factura_procesada(f, archivo_resumen=salida)
        self.aprender_equivalencias_automaticas()
        messagebox.showinfo("Listo", f"Resumen de compra/stock generado:\n\n{salida}\n\nFacturas registradas en historial permanente: {len(self.facturas_cargadas)}")

    def crear_excel_resumen(self, salida, filas):
        wb = Workbook()
        bold = Font(bold=True)
        ws = wb.active
        ws.title = "Compra_Stock"
        headers = [
            "Factura", "Proveedor", "CUIT", "Tipo comprobante", "Número comprobante", "Fecha",
            "Existe en Tango", "Acción", "Código Tango", "Descripción Tango/OCR",
            "Código proveedor", "Código barras", "Cantidad", "Precio", "Importe", "IVA %",
            "Advertencia", "Coincidencia", "Confianza", "Línea detectada"
        ]
        for col, h in enumerate(headers, start=1):
            ws.cell(1, col).value = h
            ws.cell(1, col).font = bold
        for row, f in enumerate(filas, start=2):
            values = [
                f["factura"], f["proveedor"], f["cuit"], f["tipo_comprobante"], f["numero_comprobante"],
                f["fecha"], f["existe"], f["accion"], f["codigo_tango"], f["descripcion"], f["sinonimo"],
                f["codigo_barras"], f["cantidad"], f["precio"], f["importe"], f["iva"], f["advertencia"],
                f["coincidencia"], f["match_score"], f["linea"]
            ]
            for col, value in enumerate(values, start=1):
                ws.cell(row, col).value = value
        widths = [25, 35, 15, 18, 18, 12, 16, 22, 18, 50, 18, 18, 12, 12, 12, 10, 30, 30, 12, 80]
        for idx, width in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(1, idx).column_letter].width = width

        ws2 = wb.create_sheet("Facturas_Cargadas")
        h2 = ["Clave", "Archivo", "Proveedor detectado", "CUIT proveedor", "Tipo", "Punto de venta", "Número", "Número completo", "Fecha", "Total", "Ruta"]
        for col, h in enumerate(h2, start=1):
            ws2.cell(1, col).value = h; ws2.cell(1, col).font = bold
        for row, f in enumerate(self.facturas_cargadas, start=2):
            ws2.cell(row, 1).value = f["clave"]; ws2.cell(row, 2).value = f["archivo"]
            ws2.cell(row, 3).value = f["proveedor"]; ws2.cell(row, 4).value = f["cuit_proveedor"]
            ws2.cell(row, 5).value = f["tipo"]; ws2.cell(row, 6).value = f["punto_venta"]
            ws2.cell(row, 7).value = f["numero"]; ws2.cell(row, 8).value = f["numero_completo"]
            ws2.cell(row, 9).value = f["fecha"]; ws2.cell(row, 10).value = f.get("total", "")
            ws2.cell(row, 11).value = f["ruta"]
        for col in range(1, 12):
            ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 25
        wb.save(salida)

    def generar_reporte_revision(self):
        filas = self.leer_filas()
        if not filas:
            messagebox.showerror("Error", "No hay filas para reportar."); return
        salida = filedialog.asksaveasfilename(title="Guardar reporte de revisión", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")], initialfile="Reporte_Revision.xlsx")
        if not salida: return
        wb = Workbook()
        bold = Font(bold=True)
        ws = wb.active; ws.title = "Resumen"
        resumen = [
            ("Filas totales", len(filas)),
            ("Filas seleccionadas", len([f for f in filas if f["usar"] == CHECK_ON])),
            ("Artículos existentes", len([f for f in filas if f["existe"] == "Sí"])),
            ("Artículos nuevos", len([f for f in filas if f["existe"] == "No"])),
            ("Sin código Tango", len([f for f in filas if not f["codigo_tango"]])),
            ("Advertencias cantidad/precio/importe", len([f for f in filas if f["advertencia"] and f["advertencia"] != "OK"])),
            ("Facturas cargadas", len(self.facturas_cargadas)),
        ]
        for row, (k, v) in enumerate(resumen, start=1):
            ws.cell(row, 1).value = k; ws.cell(row, 1).font = bold; ws.cell(row, 2).value = v
        ws.column_dimensions["A"].width = 40; ws.column_dimensions["B"].width = 20
        ws2 = wb.create_sheet("Detalle")
        headers = ["Usar", "Factura", "Proveedor", "CUIT", "Número", "Existe", "Acción", "Código Tango", "Descripción", "Código proveedor", "Cantidad", "Precio", "Importe", "IVA", "Advertencia", "Coincidencia", "Confianza"]
        for col, h in enumerate(headers, start=1):
            ws2.cell(1, col).value = h; ws2.cell(1, col).font = bold
        for row, f in enumerate(filas, start=2):
            vals = [f["usar"], f["factura"], f["proveedor"], f["cuit"], f["numero_comprobante"], f["existe"], f["accion"], f["codigo_tango"], f["descripcion"], f["sinonimo"], f["cantidad"], f["precio"], f["importe"], f["iva"], f["advertencia"], f["coincidencia"], f["match_score"]]
            for col, val in enumerate(vals, start=1):
                ws2.cell(row, col).value = val
        for col in range(1, len(headers)+1):
            ws2.column_dimensions[ws2.cell(1, col).column_letter].width = 22
        wb.save(salida)
        messagebox.showinfo("Listo", f"Reporte de revisión generado:\n\n{salida}")

    def guardar_sesion(self):
        salida = filedialog.asksaveasfilename(title="Guardar sesión", defaultextension=".json", filetypes=[("JSON", "*.json")], initialfile=f"Sesion_Revision_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        if not salida: return
        data = {
            "version": "V8.1", "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "ruta_modelo_tango": self.ruta_modelo_tango, "ruta_catalogo_tango": self.ruta_catalogo_tango,
            "facturas_cargadas": self.facturas_cargadas, "filas": self.leer_filas()
        }
        with open(salida, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("Listo", f"Sesión guardada:\n\n{salida}")

    def cargar_sesion(self):
        ruta = filedialog.askopenfilename(title="Cargar sesión", filetypes=[("JSON", "*.json")])
        if not ruta: return
        with open(ruta, "r", encoding="utf-8") as f:
            data = json.load(f)
        self.limpiar_todo_sin_preguntar()
        self.ruta_modelo_tango = data.get("ruta_modelo_tango")
        self.ruta_catalogo_tango = data.get("ruta_catalogo_tango")
        if self.ruta_modelo_tango:
            self.lbl_modelo.config(text=f"Modelo Tango: {self.ruta_modelo_tango}")
        if self.ruta_catalogo_tango and os.path.exists(self.ruta_catalogo_tango):
            try:
                self.catalogo.cargar(self.ruta_catalogo_tango)
                self.lbl_catalogo.config(text=f"Artículos Tango: {self.ruta_catalogo_tango} ({len(self.catalogo.articulos)} artículos)")
            except Exception:
                pass
        self.facturas_cargadas = data.get("facturas_cargadas", [])
        self.claves_facturas = set(f.get("clave", "") for f in self.facturas_cargadas)
        for f in data.get("filas", []):
            item = {
                "usar": f.get("usar") == CHECK_ON, "factura_etiqueta": f.get("factura", ""),
                "proveedor_factura": f.get("proveedor", ""), "cuit_proveedor": f.get("cuit", ""),
                "tipo_factura": f.get("tipo_comprobante", ""), "numero_completo": f.get("numero_comprobante", ""),
                "fecha_factura": f.get("fecha", ""), "existe_tango": f.get("existe", "No"),
                "accion": f.get("accion", ""), "codigo_tango": f.get("codigo_tango", ""),
                "descripcion": f.get("descripcion", ""), "sinonimo": f.get("sinonimo", ""),
                "codigo_barras": f.get("codigo_barras", ""), "cantidad": f.get("cantidad", ""),
                "precio": f.get("precio", ""), "importe": f.get("importe", ""), "iva": f.get("iva", ""),
                "advertencia": f.get("advertencia", ""), "tipo": f.get("tipo", "Simple"),
                "escala": f.get("escala", "No usa"), "pagina": f.get("pagina", ""),
                "ruta_factura": f.get("ruta", ""), "linea_detectada": f.get("linea", ""),
                "coincidencia": f.get("coincidencia", ""), "match_score": f.get("match_score", ""),
            }
            self.insertar_item(item)
        self.actualizar_label_facturas(); self.actualizar_tags(); self.aplicar_filtro()
        messagebox.showinfo("Listo", "Sesión cargada correctamente.")


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
