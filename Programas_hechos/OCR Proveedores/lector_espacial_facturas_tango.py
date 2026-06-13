import os
import re
import tempfile
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
from dataclasses import dataclass
from typing import List, Optional
from PIL import Image, ImageOps, ImageEnhance, ImageFilter
from openpyxl import load_workbook
import pytesseract
import fitz

try:
    import cv2
except ImportError:
    cv2 = None


# Ajustar si Tesseract está instalado en otra ruta
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str
    page: int = 1


@dataclass
class VisualLine:
    y: float
    x0: float
    x1: float
    words: List[WordBox]

    @property
    def text(self):
        return " ".join(w.text for w in self.words).strip()


class FacturaParserEspacial:
    def __init__(self):
        self.header_keywords = [
            "codigo", "código", "cant", "cantidad", "producto", "servicio",
            "descripcion", "descripción", "detalle", "precio", "unit", "unitario",
            "subtotal", "total", "importe", "iva", "u.m", "um"
        ]

        self.stop_keywords = [
            "importe neto", "total:", "total $", "subtotal:", "iva 21", "iva 10",
            "iva 27", "cae", "fecha de vto", "fecha cae", "son pesos",
            "observaciones", "bonificacion", "bonificación", "descuento:",
            "la mercaderia", "la mercadería", "qr"
        ]

        self.exclude_keywords = [
            "factura", "original", "duplicado", "cuit", "domicilio", "responsable",
            "fecha de emisión", "punto de venta", "comp. nro", "señor", "cliente",
            "condición de venta", "condicion de venta", "vendedor", "remito",
            "codigo de cuenta", "código de cuenta", "capital federal", "argentina",
            "ingresos brutos", "inicio de actividades"
        ]

    def extract_words_from_pdf(self, path: str) -> List[WordBox]:
        words = []
        doc = fitz.open(path)

        for page_index, page in enumerate(doc, start=1):
            raw_words = page.get_text("words")

            for w in raw_words:
                x0, y0, x1, y1, text = w[:5]
                text = str(text).strip()

                if text:
                    words.append(WordBox(x0, y0, x1, y1, text, page_index))

        return words

    def extract_words_from_image(self, path: str) -> List[WordBox]:
        img = Image.open(path)
        img = self.prepare_image(img)

        data = pytesseract.image_to_data(
            img,
            lang="spa",
            output_type=pytesseract.Output.DICT,
            config="--psm 6"
        )

        words = []

        for i, text in enumerate(data["text"]):
            text = str(text).strip()
            conf = int(data["conf"][i]) if str(data["conf"][i]).isdigit() else -1

            if not text or conf < 25:
                continue

            x = float(data["left"][i])
            y = float(data["top"][i])
            w = float(data["width"][i])
            h = float(data["height"][i])

            words.append(WordBox(x, y, x + w, y + h, text, 1))

        return words

    def prepare_image(self, img: Image.Image) -> Image.Image:
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        img = ImageEnhance.Contrast(img).enhance(2.2)
        img = ImageEnhance.Sharpness(img).enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)
        return img

    def extract_words(self, path: str) -> List[WordBox]:
        ext = os.path.splitext(path)[1].lower()

        if ext == ".pdf":
            words = self.extract_words_from_pdf(path)

            if len(words) > 30:
                return words

            # PDF escaneado: renderizar a imagen y OCR
            temp_images = self.pdf_to_temp_images(path)
            all_words = []

            for img_path in temp_images:
                all_words.extend(self.extract_words_from_image(img_path))

            return all_words

        return self.extract_words_from_image(path)

    def pdf_to_temp_images(self, path: str) -> List[str]:
        doc = fitz.open(path)
        output = []

        for page_index, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img_path = os.path.join(tempfile.gettempdir(), f"factura_page_{page_index}.png")
            pix.save(img_path)
            output.append(img_path)

        return output

    def group_words_into_lines(self, words: List[WordBox]) -> List[VisualLine]:
        if not words:
            return []

        lines = []

        for page in sorted(set(w.page for w in words)):
            page_words = [w for w in words if w.page == page]
            page_words.sort(key=lambda w: (w.y0, w.x0))

            current = []

            for w in page_words:
                if not current:
                    current.append(w)
                    continue

                avg_y = sum(c.y0 for c in current) / len(current)

                if abs(w.y0 - avg_y) <= 5:
                    current.append(w)
                else:
                    lines.append(self.make_line(current))
                    current = [w]

            if current:
                lines.append(self.make_line(current))

        lines.sort(key=lambda l: (l.words[0].page, l.y, l.x0))
        return lines

    def make_line(self, words: List[WordBox]) -> VisualLine:
        words.sort(key=lambda w: w.x0)
        return VisualLine(
            y=sum(w.y0 for w in words) / len(words),
            x0=min(w.x0 for w in words),
            x1=max(w.x1 for w in words),
            words=words
        )

    def normalize(self, text: str) -> str:
        return text.lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")

    def is_money(self, text: str) -> bool:
        text = text.strip().replace("$", "")
        return bool(re.match(r"^-?\d{1,3}(\.\d{3})*,\d{2,4}$", text)) or bool(re.match(r"^-?\d+,\d{2,4}$", text))

    def is_quantity(self, text: str) -> bool:
        text = text.strip()
        return bool(re.match(r"^\d+([,.]\d+)?$", text))

    def looks_like_item_code(self, text: str) -> bool:
        text = text.strip()
        if len(text) < 3:
            return False
        if re.match(r"^[A-Z0-9][A-Z0-9.\-/]{2,}$", text, re.IGNORECASE):
            return True
        return False

    def find_table_header_y(self, lines: List[VisualLine]) -> Optional[float]:
        best_score = 0
        best_y = None

        for line in lines:
            txt = self.normalize(line.text)
            score = sum(1 for k in self.header_keywords if k in txt)

            if score > best_score:
                best_score = score
                best_y = line.y

        if best_score >= 3:
            return best_y

        return None

    def is_stop_line(self, line: VisualLine) -> bool:
        txt = self.normalize(line.text)
        return any(k in txt for k in self.stop_keywords)

    def is_excluded_line(self, line: VisualLine) -> bool:
        txt = self.normalize(line.text)
        return any(k in txt for k in self.exclude_keywords)

    def line_has_multiple_columns(self, line: VisualLine) -> bool:
        xs = [w.x0 for w in line.words]

        if len(xs) < 4:
            return False

        gaps = []

        for i in range(len(xs) - 1):
            gap = xs[i + 1] - line.words[i].x1
            if gap > 20:
                gaps.append(gap)

        return len(gaps) >= 2

    def analyze_line(self, line: VisualLine):
        words = line.words
        nums = []
        monies = []

        for w in words:
            t = w.text.strip()
            if self.is_money(t):
                monies.append(w)
            elif self.is_quantity(t):
                nums.append(w)

        return nums, monies

    def parse(self, path: str):
        words = self.extract_words(path)
        lines = self.group_words_into_lines(words)

        header_y = self.find_table_header_y(lines)

        if header_y is None:
            candidate_lines = lines
        else:
            candidate_lines = [l for l in lines if l.y > header_y + 3]

        items = []
        current_item = None
        inside_table = False

        for line in candidate_lines:
            if self.is_stop_line(line):
                if inside_table:
                    break
                continue

            if self.is_excluded_line(line):
                continue

            nums, monies = self.analyze_line(line)
            has_columns = self.line_has_multiple_columns(line)

            item_line = False

            # Regla universal:
            # artículo probable = tiene estructura de columnas y al menos:
            # cantidad + precio/importe, o dos valores monetarios.
            if has_columns and (len(monies) >= 2 or (len(nums) >= 1 and len(monies) >= 1)):
                item_line = True

            # caso alternativo: línea con código fuerte + importes
            if not item_line and len(line.words) >= 4 and len(monies) >= 2:
                if self.looks_like_item_code(line.words[0].text):
                    item_line = True

            if item_line:
                inside_table = True
                item = self.build_item_from_line(line)
                items.append(item)
                current_item = item
                continue

            # línea de continuación dentro de la tabla
            if inside_table and current_item is not None:
                txt = line.text.strip()

                if not txt:
                    continue

                # Si es solo numérica/precios, no agregar a descripción
                if re.match(r"^[0-9.,\-\sU%$]+$", txt):
                    continue

                # Si parece texto descriptivo, agregarlo al item anterior
                if len(txt) > 2 and not self.is_excluded_line(line):
                    current_item["descripcion"] = (current_item["descripcion"] + " " + txt)[:50]
                    current_item["linea_ocr"] += " | " + txt

        return items, "\n".join(l.text for l in lines)

    def build_item_from_line(self, line: VisualLine):
        words = line.words
        text = line.text

        # Detectar código: primera palabra alfanumérica útil
        codigo_proveedor = ""
        for w in words[:4]:
            if self.looks_like_item_code(w.text) and not self.is_quantity(w.text) and not self.is_money(w.text):
                codigo_proveedor = w.text.strip()
                break

        # Armar descripción: palabras no numéricas, no importes, no columnas finales
        desc_parts = []

        for w in words:
            t = w.text.strip()

            if t == codigo_proveedor:
                continue
            if self.is_money(t):
                continue
            if self.is_quantity(t):
                continue
            if t.upper() in ["U", "UN", "KG", "M", "L", "%"]:
                continue
            if re.match(r"^-?\d+([,.]\d+)?%?$", t):
                continue

            desc_parts.append(t)

        descripcion = " ".join(desc_parts)
        descripcion = re.sub(r"\s+", " ", descripcion).strip()
        descripcion = descripcion.replace(" /", "").replace("/", "")

        return {
            "codigo_tango": "",
            "descripcion": descripcion[:50],
            "desc_adic": "",
            "sinonimo": codigo_proveedor[:15],
            "codigo_barras": "",
            "tipo": "Simple",
            "escala": "No usa",
            "linea_ocr": text,
        }


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Lector espacial de facturas → Tango")
        self.root.geometry("1250x670")

        self.parser = FacturaParserEspacial()
        self.ruta_factura = None
        self.ruta_modelo_tango = None
        self.texto_leido = ""

        self.crear_interfaz()

    def crear_interfaz(self):
        frame = tk.Frame(self.root)
        frame.pack(fill="x", padx=10, pady=10)

        tk.Button(frame, text="1. Abrir imagen/PDF", command=self.abrir_factura, width=24).pack(side="left", padx=4)
        tk.Button(frame, text="Sacar foto", command=self.sacar_foto, width=14).pack(side="left", padx=4)
        tk.Button(frame, text="2. Modelo Tango", command=self.abrir_modelo, width=20).pack(side="left", padx=4)
        tk.Button(frame, text="3. Generar Excel Tango", command=self.generar_excel_tango, width=24).pack(side="left", padx=4)

        self.lbl_factura = tk.Label(self.root, text="Factura: no seleccionada", anchor="w")
        self.lbl_factura.pack(fill="x", padx=15)

        self.lbl_modelo = tk.Label(self.root, text="Modelo Tango: no seleccionado", anchor="w")
        self.lbl_modelo.pack(fill="x", padx=15)

        columns = (
            "codigo_tango",
            "descripcion",
            "desc_adic",
            "sinonimo",
            "codigo_barras",
            "tipo",
            "escala",
            "linea_ocr",
        )

        self.tabla = ttk.Treeview(self.root, columns=columns, show="headings", height=22)

        headers = {
            "codigo_tango": "Código Tango",
            "descripcion": "Descripción",
            "desc_adic": "Desc. adicional",
            "sinonimo": "Código proveedor",
            "codigo_barras": "Código barras",
            "tipo": "Tipo",
            "escala": "Escala",
            "linea_ocr": "Línea detectada",
        }

        widths = {
            "codigo_tango": 120,
            "descripcion": 320,
            "desc_adic": 120,
            "sinonimo": 140,
            "codigo_barras": 150,
            "tipo": 90,
            "escala": 90,
            "linea_ocr": 500,
        }

        for c in columns:
            self.tabla.heading(c, text=headers[c])
            self.tabla.column(c, width=widths[c])

        self.tabla.pack(fill="both", expand=True, padx=10, pady=10)
        self.tabla.bind("<Double-1>", self.editar_celda)

        bottom = tk.Frame(self.root)
        bottom.pack(fill="x", padx=10, pady=5)

        tk.Button(bottom, text="Agregar manual", command=self.agregar_manual, width=18).pack(side="left", padx=5)
        tk.Button(bottom, text="Eliminar seleccionado", command=self.eliminar, width=22).pack(side="left", padx=5)
        tk.Button(bottom, text="Ver texto detectado", command=self.ver_texto, width=20).pack(side="left", padx=5)

        tk.Label(
            self.root,
            text="Regla nueva: detecta estructura de tabla por coordenadas. Doble click para editar.",
            fg="blue"
        ).pack(pady=5)

    def abrir_factura(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar factura",
            filetypes=[
                ("Facturas", "*.pdf *.jpg *.jpeg *.png *.bmp *.tif *.tiff"),
                ("Todos", "*.*")
            ]
        )

        if ruta:
            self.ruta_factura = ruta
            self.lbl_factura.config(text=f"Factura: {ruta}")
            self.procesar_factura(ruta)

    def sacar_foto(self):
        if cv2 is None:
            messagebox.showerror("Error", "Falta instalar opencv-python")
            return

        cap = cv2.VideoCapture(0)

        if not cap.isOpened():
            messagebox.showerror("Error", "No se pudo abrir la cámara.")
            return

        messagebox.showinfo("Cámara", "ESPACIO captura / ESC cancela")

        ruta_foto = None

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            cv2.imshow("Capturar factura", frame)
            tecla = cv2.waitKey(1)

            if tecla == 27:
                break
            if tecla == 32:
                ruta_foto = os.path.join(tempfile.gettempdir(), "factura_capturada.jpg")
                cv2.imwrite(ruta_foto, frame)
                break

        cap.release()
        cv2.destroyAllWindows()

        if ruta_foto:
            self.ruta_factura = ruta_foto
            self.lbl_factura.config(text=f"Factura: {ruta_foto}")
            self.procesar_factura(ruta_foto)

    def abrir_modelo(self):
        ruta = filedialog.askopenfilename(
            title="Seleccionar Excel modelo exportado desde Tango",
            filetypes=[("Excel", "*.xlsx")]
        )

        if ruta:
            self.ruta_modelo_tango = ruta
            self.lbl_modelo.config(text=f"Modelo Tango: {ruta}")

    def procesar_factura(self, ruta):
        try:
            items, texto = self.parser.parse(ruta)
            self.texto_leido = texto
            self.cargar_items(items)

            messagebox.showinfo(
                "Lectura terminada",
                f"Se detectaron {len(items)} artículos probables."
            )

        except Exception as e:
            messagebox.showerror("Error", str(e))

    def cargar_items(self, items):
        for item in self.tabla.get_children():
            self.tabla.delete(item)

        for item in items:
            self.tabla.insert(
                "",
                "end",
                values=(
                    item["codigo_tango"],
                    item["descripcion"],
                    item["desc_adic"],
                    item["sinonimo"],
                    item["codigo_barras"],
                    item["tipo"],
                    item["escala"],
                    item["linea_ocr"],
                )
            )

    def editar_celda(self, event):
        item_id = self.tabla.focus()
        if not item_id:
            return

        col = self.tabla.identify_column(event.x)
        index = int(col.replace("#", "")) - 1

        campos = [
            "Código Tango",
            "Descripción",
            "Desc. adicional",
            "Código proveedor",
            "Código barras",
            "Tipo",
            "Escala",
            "Línea detectada",
        ]

        valores = list(self.tabla.item(item_id, "values"))
        actual = valores[index]

        nuevo = simpledialog.askstring("Editar", campos[index], initialvalue=actual)

        if nuevo is not None:
            valores[index] = nuevo
            self.tabla.item(item_id, values=valores)

    def leer_items_tabla(self):
        items = []

        for item_id in self.tabla.get_children():
            v = self.tabla.item(item_id, "values")

            items.append({
                "codigo_tango": v[0].strip(),
                "descripcion": v[1].strip(),
                "desc_adic": v[2].strip(),
                "sinonimo": v[3].strip(),
                "codigo_barras": v[4].strip(),
                "tipo": v[5].strip() or "Simple",
                "escala": v[6].strip() or "No usa",
            })

        return items

    def agregar_manual(self):
        self.tabla.insert("", "end", values=("", "", "", "", "", "Simple", "No usa", "Manual"))

    def eliminar(self):
        for item in self.tabla.selection():
            self.tabla.delete(item)

    def ver_texto(self):
        w = tk.Toplevel(self.root)
        w.title("Texto detectado")
        w.geometry("900x600")

        txt = tk.Text(w, wrap="word")
        txt.pack(fill="both", expand=True)
        txt.insert("1.0", self.texto_leido)

    def generar_excel_tango(self):
        if not self.ruta_modelo_tango:
            messagebox.showerror("Error", "Primero seleccioná el modelo Tango.")
            return

        items = self.leer_items_tabla()

        if not items:
            messagebox.showerror("Error", "No hay artículos.")
            return

        faltan = [i for i in items if not i["codigo_tango"]]

        if faltan:
            messagebox.showerror(
                "Faltan códigos Tango",
                "Hay artículos sin Código Tango. Completalos antes de exportar."
            )
            return

        salida = filedialog.asksaveasfilename(
            title="Guardar Excel Tango",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile="Articulos_Para_Importar_Tango.xlsx"
        )

        if not salida:
            return

        try:
            wb = load_workbook(self.ruta_modelo_tango)

            if "Artículos" not in wb.sheetnames:
                messagebox.showerror("Error", "El modelo no tiene hoja 'Artículos'.")
                return

            ws = wb["Artículos"]

            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)

            for fila, item in enumerate(items, start=2):
                ws.cell(fila, 1).value = item["codigo_tango"][:15]
                ws.cell(fila, 2).value = item["descripcion"][:50]
                ws.cell(fila, 3).value = item["desc_adic"][:20]
                ws.cell(fila, 4).value = item["sinonimo"][:15]
                ws.cell(fila, 5).value = item["codigo_barras"][:40]
                ws.cell(fila, 6).value = item["tipo"]
                ws.cell(fila, 7).value = item["escala"]

            wb.save(salida)
            messagebox.showinfo("Listo", f"Excel generado:\n\n{salida}")

        except Exception as e:
            messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()